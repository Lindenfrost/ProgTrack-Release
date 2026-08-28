# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright © 2026 Dimitri L. Lindenwald and Deutsches Primatenzentrum GmbH
# Part of: ProgTrack 0.2.2
# Required ProgTrack version: see plugin manifest.
# Required Launcher version: 0.1.0 RC or newer.
# Module: Flow Track embryo-flow visualization widget.

import os
import json
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

from Plugins.core.animal_identity import animal_base_name
from Plugins.core.animal_roles import canonical_role_value, role_color_for_record
from Plugins.core.platform_helpers import default_save_path
from Plugins.core.backend_store import BackendJsonStore
from Plugins.core.ui_icons import apply_icon
from Plugins.core.dialog_geometry import install_dialog_geometry_guard

# Set up paths
PLUGIN_DIR = os.path.dirname(os.path.abspath(__file__))
PLUGINS_ROOT = os.path.dirname(PLUGIN_DIR)
PROGTRACK_ROOT = os.path.dirname(PLUGINS_ROOT)
ICON_DIR = os.path.join(PROGTRACK_ROOT, 'icons')

# Date format constant
DATE_FORMAT = "%d.%m.%Y"

# Flow_Track 3.0 Constants
FREEZER_NODE_NAME = "FREEZER"
FREEZER_TRANSFER_ID = "freezer_main"

# Stage naming constants (Flow_Track 3.0)
STAGE_IN_VIVO_M2 = "in_vivo_m2"
STAGE_IN_VITRO_M2 = "in_vitro_m2"
VALID_STAGES = {STAGE_IN_VIVO_M2, STAGE_IN_VITRO_M2}

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('FlowTrack')


def _animal_role_value(animal_data: Optional[Dict[str, Any]]) -> str:
    """Return the stable internal role ID for a ProgTrack animal record."""
    return canonical_role_value((animal_data or {}).get('rolle'))


def _flow_datetime(value: Any) -> Optional[datetime]:
    """Normalize backend ISO/date values before Flow Track builds IDs."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        for fmt in ("%d.%m.%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


class FlowTrackWidget:
    """
    Main widget for Flow Track plugin.
    
    This plugin visualizes embryo transfer flow between:
    - Egg donors (left rail)
    - Sperm donors (right rail)  
    - Surrogates (bottom rail)
    
    Uses ProgTrack's LazyLoader pattern - all Qt/matplotlib/numpy accessed via parent_app.
    Returns a QWidget suitable for tab integration.
    """
    
    def __init__(self, parent_app, messages: Optional[Dict] = None):
        """
        Initialize Flow Track widget.
        
        Args:
            parent_app: Main ProgTrack application instance (provides LazyLoader modules)
            messages: Localization dictionary (optional)
        """
        # Store parent app reference for LazyLoader access
        self.parent_app = parent_app
        self.messages = messages or {}
        self._data_store = BackendJsonStore(
            parent_app.backend, "reproduction", "flow"
        )
        self._settings_store = BackendJsonStore(
            parent_app.backend, "configuration", "flow-track"
        )
        
        # Access Qt modules via parent_app LazyLoader
        QtWidgets = parent_app.QtWidgets
        QtCore = parent_app.QtCore
        
        # Create main widget (for tab integration, not a dialog)
        self.widget = QtWidgets.QWidget()
        
        # No window icon needed for tab widget
        
        # Initialize data structures (Flow_Track 3.0 schema)
        self.manual_data = {
            'sperm_donors': {},  # donor_name -> {donations: {donation_id: {...}}}
            'egg_donors': {},    # donor_name -> {surgeries: {surgery_date: {...}}}
            'transfers_by_id': {},  # transfer_id -> {surrogate_name, transfer_date, embryos: []}
            'total_embryo_count': 0  # Global counter for embryo ID generation
        }
        self.settings = {
            'egg_donor_metric': 'usable_per_gv',  # DEFAULT: Overall usable yield per GV
            'egg_donor_aggregation': 'total',  # DEFAULT: total across all surgeries
            'sperm_donor_metric': 'usable_per_ivm',  # DEFAULT: Usable embryo yield per IVM M2
            'sperm_donor_aggregation': 'total',  # DEFAULT: total across all donations
            'surrogate_metric': 'total_implantation',  # DEFAULT: Total Implantation rate
            'surrogate_aggregation': 'total',  # DEFAULT: total across all transfers
            'render_full_graph': True,
            'show_animal_names': True,
            'show_lifebars': True,
            'show_grid': False,
            'snap_to_grid': False,
            'auto_save_enabled': True,
            'embryo_visibility': {}  # Per-node embryo visibility state
        }
        self.ui_preferences = {}
        self.lifebar_data = {}  # animal_name -> [{event_id, event_date, metric_name, efficiency_pct, patches, position, bounds}]
        
        # Event lists from ProgTrack
        self.egg_surgeries = []  # List of egg donation surgeries
        self.sperm_donations = []  # List of sperm donations
        self.embryo_transfers = []  # List of embryo transfers
        
        # Artist registries for click handling
        self.node_artists = {}  # artist -> animal_name
        self.embryo_artists = {}  # artist -> (transfer_id, embryo_id)
        
        # Temporary node positions (reset on redraw)
        self.temp_positions = {}  # (node_type, node_id) -> (x, y)
        
        # Pan/zoom state
        self.pan_active = False
        self.pan_start = None
        self.current_xlim = None
        self.current_ylim = None
        self.auto_fit_bounds = None  # Cache for auto-calculated bounds
        
        # Drag state
        self.drag_active = False
        self.drag_artist = None
        self.drag_offset = (0, 0)
        self.click_start_pos = None  # Track initial click position
        self.is_dragging = False  # True only after movement threshold exceeded
        self.drag_threshold = 0.05  # Minimum movement to be considered a drag (increased from 0.01)
        
        # Embryo visibility toggle per node (freezer + surrogates)
        # Key: node name (animal name or FREEZER_NODE_NAME), Value: boolean (True=visible, False=hidden)
        self.embryo_visibility = {}
        
        # Pending menu action for click vs drag detection
        self.pending_menu_action = None
        
        # Undo/Redo system
        self.undo_manager = UndoManager()
        
        # Initialize UI
        self._init_ui()
        
        # Load data
        self._load_settings()
        # Load embryo visibility state from settings
        self.embryo_visibility = self.settings.get('embryo_visibility', {})
        
        # Initialize timeline visibility after settings are loaded
        if hasattr(self, 'timeline_widget'):
            self.timeline_widget.setVisible(self.settings.get('show_timeline', False))
        
        self._load_flow_track_data()
        self._ensure_freezer_transfer_exists()  # Flow_Track 3.0: Initialize freezer
        self._populate_events_from_progtrack()
        
        # Initial render
        self._redraw_canvas()
    
    def _init_ui(self):
        """Initialize the user interface."""
        QtWidgets = self.parent_app.QtWidgets
        QtCore = self.parent_app.QtCore
        Qt = self.parent_app.QtCore.Qt
        QStackedWidget = self.parent_app.QtWidgets.QStackedWidget
        QLabel = self.parent_app.QtWidgets.QLabel
        QPixmap = self.parent_app.QtGui.QPixmap
        
        # Main layout
        main_layout = QtWidgets.QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create stacked widget to hold splash and content
        self.stack = QStackedWidget()
        main_layout.addWidget(self.stack)
        
        # === SPLASH WIDGET (index 0) ===
        self.splash_widget = QtWidgets.QWidget()
        splash_layout = QtWidgets.QVBoxLayout(self.splash_widget)
        splash_layout.addStretch(1)
        
        # Add disclaimer
        disclaimer_label = QLabel(
            self.messages.get("footer.rights", "ProgTrack").format(year=datetime.now().year)
        )
        disclaimer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        splash_layout.addWidget(disclaimer_label, alignment=Qt.AlignmentFlag.AlignCenter)
        splash_layout.addSpacing(20)
        
        # Add splash image
        img_label = QLabel()
        pix_path = Path(ICON_DIR) / 'Splash.png'
        if pix_path.exists():
            pix = QPixmap(str(pix_path))
            pix = pix.scaled(800, 800, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            img_label.setPixmap(pix)
        img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        splash_layout.addWidget(img_label, alignment=Qt.AlignmentFlag.AlignCenter)
        splash_layout.addStretch(1)
        self.stack.addWidget(self.splash_widget)
        
        # === CONTENT WIDGET (index 1) ===
        self.content_widget = QtWidgets.QWidget()
        content_layout = QtWidgets.QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # === TOOLBAR ===
        toolbar_layout = QtWidgets.QHBoxLayout()
        
        # Add stretch to center the title
        toolbar_layout.addStretch()
        
        # Title label in the center (store reference for language updates)
        title_text = self.messages.get("flow_track.visualization.title", "Embryo Transfer Flow")
        subtitle_text = self.messages.get("flow_track.visualization.subtitle", "Donor-Surrogate relationship visualization")
        combined_text = f"<b>{title_text}</b>: {subtitle_text}"
        
        self.title_label = QtWidgets.QLabel(combined_text)
        self.title_label.setStyleSheet("color: black; font-size: 11pt;")
        self.title_label.setAlignment(self.parent_app.QtCore.Qt.AlignmentFlag.AlignCenter)
        toolbar_layout.addWidget(self.title_label)
        
        # Stretch to push buttons to the right
        toolbar_layout.addStretch()
        
        # Fit Automatically button (store reference for language updates)
        self.fit_btn = QtWidgets.QPushButton()
        apply_icon(self.fit_btn, "action.refresh", fallback="Refresh")
        # Match the 30px icon treatment used by the adjacent toolbar buttons.
        self.fit_btn.setIconSize(QtCore.QSize(30, 30))
        self.fit_btn.setToolTip(self.messages.get("flow_track.button.fit_auto", "Fit Automatically"))
        self.fit_btn.clicked.connect(self._fit_automatically)
        toolbar_layout.addWidget(self.fit_btn)
        
        # Freezer toggle button (Flow_Track 3.0)
        self.freezer_btn = QtWidgets.QPushButton()
        apply_icon(self.freezer_btn, "flow.freezer", fallback="Freezer")
        self.freezer_btn.setIconSize(QtCore.QSize(30, 30))
        self.freezer_btn.setToolTip(self.messages.get("flow_track.button.freezer", "Show/Hide Freezer"))
        self.freezer_btn.clicked.connect(self._toggle_freezer_visibility)
        toolbar_layout.addWidget(self.freezer_btn)
        
        # Settings button (store reference for language updates)
        self.settings_btn = QtWidgets.QPushButton()
        apply_icon(self.settings_btn, "action.settings", fallback="Settings")
        self.settings_btn.setIconSize(QtCore.QSize(30, 30))
        self.settings_btn.setToolTip(self.messages.get("flow_track.button.settings", "Settings"))
        self.settings_btn.clicked.connect(self._open_settings_dialog)
        toolbar_layout.addWidget(self.settings_btn)
        
        content_layout.addLayout(toolbar_layout)
        
        # === CANVAS AREA ===
        canvas_container = QtWidgets.QWidget()
        canvas_layout = QtWidgets.QVBoxLayout(canvas_container)
        
        # Access matplotlib via parent_app LazyLoader
        matplotlib = self.parent_app.matplotlib
        Figure = matplotlib.figure.Figure
        FigureCanvasQTAgg = matplotlib.backends.backend_qtagg.FigureCanvasQTAgg
        
        # Create matplotlib figure that adapts to window size
        self.figure = Figure(dpi=100)
        self.canvas = FigureCanvasQTAgg(self.figure)
        self.ax = self.figure.add_subplot(111)
        
        # Set canvas to expand and fill available space
        self.canvas.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Expanding
        )
        
        # Maximize subplot to use entire figure area - no margins
        self.figure.subplots_adjust(left=0, right=1, top=1, bottom=0)
        
        # Store annotation reference for tooltips (like ProgTrack weight plot)
        self._hover_annotation = None
        
        # Connect canvas events
        self.canvas.mpl_connect('pick_event', self._on_pick)
        self.canvas.mpl_connect('motion_notify_event', self._on_hover)
        self.canvas.mpl_connect('scroll_event', self._on_scroll)
        self.canvas.mpl_connect('button_press_event', self._on_mouse_press)
        self.canvas.mpl_connect('button_release_event', self._on_mouse_release)
        self.canvas.mpl_connect('motion_notify_event', self._on_mouse_move)
        self.canvas.mpl_connect('resize_event', self._on_resize)
        
        canvas_layout.addWidget(self.canvas)
        
        # === TIMELINE CHART AREA ===
        self.timeline_widget = QtWidgets.QWidget()
        self.timeline_layout = QtWidgets.QVBoxLayout(self.timeline_widget)
        
        # Set timeline widget to expand horizontally but have fixed vertical height
        self.timeline_widget.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Preferred
        )
        
        # Timeline canvas
        self.timeline_figure = Figure(dpi=100, figsize=(12, 3))
        self.timeline_canvas = FigureCanvasQTAgg(self.timeline_figure)
        self.timeline_ax = self.timeline_figure.add_subplot(111)
        
        # Set timeline canvas to expand horizontally
        self.timeline_canvas.setSizePolicy(
            QtWidgets.QSizePolicy.Policy.Expanding,
            QtWidgets.QSizePolicy.Policy.Preferred
        )
        
        # Adjust subplot margins to prevent label cropping (extra bottom margin for rotated labels)
        self.timeline_figure.subplots_adjust(left=0.1, right=0.95, top=0.9, bottom=0.25)
        
        # Initially hidden, controlled by settings
        self.timeline_widget.setVisible(False)
        
        # Add timeline canvas to timeline layout
        self.timeline_layout.addWidget(self.timeline_canvas)
        
        canvas_layout.addWidget(self.timeline_widget)
        content_layout.addWidget(canvas_container)
        
        # Add content widget to stack
        self.stack.addWidget(self.content_widget)
        
        # Show splash by default
        self.stack.setCurrentWidget(self.splash_widget)
        
        self.widget.setLayout(main_layout)
    
    def get_widget(self):
        """Return the main widget for tab integration."""
        return self.widget
    
    def _translate_stage(self, stage_value):
        """
        Translate stage constant for UI display.
        Internal data stays English, but UI shows localized text.
        """
        if stage_value == STAGE_IN_VIVO_M2:
            return self.messages.get('flow_track.stage.in_vivo_m2', 'in vivo M2')
        elif stage_value == STAGE_IN_VITRO_M2:
            return self.messages.get('flow_track.stage.in_vitro_m2', 'IVM M2')
        else:
            return stage_value  # Fallback to original value
    
    def _translate_freezer_label(self):
        """Translate FREEZER label for UI display."""
        return self.messages.get('flow_track.freezer_label', 'FREEZER')
    
    def update_language(self, messages):
        """Update all text when language is changed."""
        self.messages = messages
        
        # Update title label
        title_text = self.messages.get("flow_track.visualization.title", "Embryo Transfer Flow")
        subtitle_text = self.messages.get("flow_track.visualization.subtitle", "Donor-Surrogate relationship visualization")
        combined_text = f"<b>{title_text}</b>: {subtitle_text}"
        self.title_label.setText(combined_text)
        
        # Update button tooltips
        if hasattr(self, 'fit_btn'):
            self.fit_btn.setToolTip(self.messages.get("flow_track.button.fit_auto", "Fit Automatically"))
        if hasattr(self, 'settings_btn'):
            self.settings_btn.setToolTip(self.messages.get("flow_track.button.settings", "Settings"))
        if hasattr(self, 'freezer_btn'):
            self.freezer_btn.setToolTip(self.messages.get("flow_track.button.freezer", "Show/Hide Freezer"))
        
        # Redraw canvas to update any visible text (node labels, etc.)
        self._redraw_canvas()
        
        # Update timeline if visible
        if self.settings.get('show_timeline', False):
            self._render_timeline_chart()
    
    def _load_settings(self):
        """Load plugin settings from the configured backend (Flow_Track 3.0 schema)."""
        config = self._settings_store.load({})
        self.settings = config.get('settings', self._get_default_settings())
        self.ui_preferences = config.get('ui_preferences', {})
    
    def _get_default_settings(self):
        """Get default settings (Flow_Track 3.0 schema)."""
        return {
            'egg_lifebar': {
                'metric': 'fertilized_to_implantation',
                'stage': STAGE_IN_VITRO_M2,
                'aggregation': 'total'
            },
            'sperm_lifebar': {
                'metric': 'transferred_to_implantation',
                'stage': STAGE_IN_VITRO_M2,
                'aggregation': 'total'
            },
            'freezer_visible': True,
            'show_timeline': False,
            'max_suggestions_per_transfer': 3,
            'suggestion_tolerance_days': 7,
            'show_egg_donors': True,
            'show_sperm_donors': True,
            'show_surrogates': True,
            'render_full_graph': True,
            'show_animal_names': True,
            'show_lifebars': True
        }
    
    def _save_settings(self):
        """Save settings to the configured backend (Flow_Track 3.0 schema)."""
        try:
            config = {
                'settings': self.settings,
                'ui_preferences': self.ui_preferences
            }
            self._settings_store.save(config)
            logger.info("Flow Track settings saved through backend")
        except Exception as e:
            logger.error(f"Failed to save settings: {e}")

    def _set_scrollable_dialog_layout(self, dialog, content_layout):
        """Attach a layout through a scroll area and keep tall dialogs on screen."""
        QtWidgets = self.parent_app.QtWidgets
        scroll = QtWidgets.QScrollArea(dialog)
        scroll.setWidgetResizable(True)
        content = QtWidgets.QWidget(scroll)
        content.setLayout(content_layout)
        scroll.setWidget(content)

        outer_layout = QtWidgets.QVBoxLayout(dialog)
        outer_layout.addWidget(scroll)
        dialog.setSizeGripEnabled(True)

        screen = QtWidgets.QApplication.primaryScreen()
        if screen is not None:
            available = screen.availableGeometry()
            max_width = max(360, int(available.width() * 0.9))
            max_height = max(320, int(available.height() * 0.9))
            dialog.setMaximumSize(max_width, max_height)
            if dialog.width() > max_width or dialog.height() > max_height:
                dialog.resize(min(dialog.width(), max_width), min(dialog.height(), max_height))
        install_dialog_geometry_guard(dialog)
    
    def _load_flow_track_data(self):
        """Load Flow Track records from the configured backend (3.0 schema)."""
        try:
            data = self._data_store.load({
                "version": "3.0",
                "manual_data": {},
            })
            
            version = data.get('version', '1.0')
            if version != '3.0':
                logger.warning(f"Loading data with version {version}, expected 3.0")
            
            # Load nested manual_data structure (v3.0 schema)
            loaded_manual_data = data.get('manual_data', {})
            self.manual_data = {
                'sperm_donors': loaded_manual_data.get('sperm_donors', {}),
                'egg_donors': loaded_manual_data.get('egg_donors', {}),
                'transfers_by_id': loaded_manual_data.get('transfers_by_id', {}),
                'total_embryo_count': loaded_manual_data.get('total_embryo_count', 0)
            }
            
            # Load per-animal transfer data (for embryo counts from surrogate dialog)
            for animal_name, animal_data in loaded_manual_data.items():
                if animal_name not in ['sperm_donors', 'egg_donors', 'transfers_by_id', 'total_embryo_count']:
                    self.manual_data[animal_name] = animal_data
            
            logger.info("Flow Track data loaded through backend (v%s)", version)
        except Exception as e:
            logger.error(f"Failed to load Flow Track records from backend: {e}")
            self.manual_data = {
                'sperm_donors': {},
                'egg_donors': {},
                'transfers_by_id': {},
                'total_embryo_count': 0
            }
    
    def _validate_data(self) -> tuple[bool, str]:
        """Validate data before saving (Flow_Track 3.0 section 11 - hard-blocking validations).
        
        Returns:
            (is_valid, error_message) tuple
        """
        errors = []
        
        # Check all embryos across all transfers
        embryo_ids_seen = set()
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            for embryo in transfer_data.get('embryos', []):
                embryo_id = embryo.get('embryo_id')
                
                # Validation 1: embryo_id must be non-empty
                if not embryo_id:
                    errors.append(f"Embryo in {transfer_id} has empty embryo_id")
                    continue
                
                # Validation 2: embryo_id must be globally unique
                if embryo_id in embryo_ids_seen:
                    errors.append(f"Duplicate embryo_id: {embryo_id}")
                embryo_ids_seen.add(embryo_id)
                
                # Validation 3: stage must be valid (in_vivo_m2 or in_vitro_m2)
                stage = embryo.get('stage')
                if stage not in VALID_STAGES:
                    errors.append(f"Embryo {embryo_id} has invalid stage: {stage}")
                
                # Validation 4: cryopreserved=True requires freeze_date
                if embryo.get('cryopreserved', False) and not embryo.get('freeze_date'):
                    errors.append(f"Frozen embryo {embryo_id} missing freeze_date")
        
        if errors:
            error_msg = self.messages.get(
                "flow_track.validation.errors",
                "Data validation failed:\n\n"
            ) + "\n".join(errors)
            return False, error_msg
        
        return True, ""
    
    def _can(self, action: str) -> bool:
        """Check permission using namespaced flow_track permissions."""
        fn = getattr(self.parent_app, '_master_can', None)
        if fn is None:
            return True
        # Map internal action names to full permission names
        perm_map = {
            'edit': 'flow_track.edit',
            'create': 'flow_track.create',
            'delete': 'flow_track.delete',
            'use': 'flow_track.use',
            'open': 'flow_track.open',
        }
        perm = perm_map.get(action, f'flow_track.{action}')
        return fn(perm)

    def _deny(self) -> None:
        if hasattr(self.parent_app, '_show_permission_denied'):
            self.parent_app._show_permission_denied()

    def _save_flow_track_data(self):
        """Save Flow Track records to the configured backend (3.0 schema with validations)."""
        if not self._can('edit'):
            self._deny()
            return
        # Hard-blocking validation (section 11)
        is_valid, error_msg = self._validate_data()
        if not is_valid:
            QtWidgets = self.parent_app.QtWidgets
            QtWidgets.QMessageBox.critical(
                self.widget,
                self.messages.get("flow_track.validation.title", "Validation Error"),
                error_msg
            )
            logger.error(f"Validation failed: {error_msg}")
            return  # Block save
        
        # v3.0 schema: nested manual_data (includes per-animal transfer data for persistence)
        data = {
            'version': '3.0',
            'last_updated': datetime.now().isoformat(),
            'manual_data': self.manual_data  # This now includes both core data and per-animal data
        }
        
        try:
            self._data_store.save(data)
            logger.info("Flow Track data saved through backend (v3.0)")
        except Exception as e:
            logger.error(f"Failed to save Flow Track records to backend: {e}")
            QtWidgets = self.parent_app.QtWidgets
            QtWidgets.QMessageBox.warning(
                self.widget,
                self.messages.get("flow_track.error.save_failed.title", "Save Error"),
                self.messages.get("flow_track.error.save_failed.message", f"Failed to save: {e}")
            )
    
    def _is_embryo_id_unique(self, embryo_id: str, exclude_transfer_id: str = None, exclude_embryo_id: str = None) -> bool:
        """Check if embryo_id is globally unique across all transfers (v3.0 hard requirement).
        
        Args:
            embryo_id: The embryo ID to check
            exclude_transfer_id: Optional transfer to exclude from check (for editing)
            exclude_embryo_id: Optional embryo to exclude from check (for editing)
        
        Returns:
            True if unique, False if duplicate exists
        """
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            # Skip excluded transfer when editing
            if transfer_id == exclude_transfer_id:
                continue
            
            for embryo in transfer_data.get('embryos', []):
                # Skip excluded embryo when editing
                if embryo.get('embryo_id') == exclude_embryo_id:
                    continue
                
                if embryo.get('embryo_id') == embryo_id:
                    return False  # Duplicate found
        
        return True  # Unique
    
    def _generate_next_embryo_id(self, transfer_id: str, freeze_date: str = None) -> str:
        """Generate next unique embryo ID using total_embryo_count (v3.0 spec section 6.1).
        
        Args:
            transfer_id: The transfer this embryo belongs to
            freeze_date: If provided, embryo is being created in freezer
        
        Returns:
            Proposed unique embryo ID
        """
        # Increment counter
        self.manual_data['total_embryo_count'] += 1
        counter = self.manual_data['total_embryo_count']
        
        if freeze_date:
            # Freezer embryo: FZ_{freeze_date}_E{counter:05d}
            embryo_id = f"FZ_{freeze_date}_E{counter:05d}"
        else:
            # Surrogate transfer embryo: use a sanitized short-name token, never raw IPID.
            transfer_data = self.manual_data.get('transfers_by_id', {}).get(transfer_id, {})
            surrogate_name = transfer_data.get('surrogate_name', 'UNKNOWN')
            transfer_date = transfer_data.get('transfer_date', 'UNKNOWN')
            embryo_id = f"{self._safe_identity_token(surrogate_name)}_{transfer_date}_E{counter:05d}"
        
        return embryo_id

    @staticmethod
    def _safe_identity_token(value: str) -> str:
        token = animal_base_name(value) or "UNKNOWN"
        # Legacy references may omit the origin component while still using
        # the human-readable ``name | species | birth`` form.  Embryo IDs must
        # remain short and stable; never embed the complete identity string.
        if " | " in token:
            token = token.split(" | ", 1)[0].strip()
        token = re.sub(r"[^A-Za-z0-9._-]+", "_", token).strip("_")
        return token or "UNKNOWN"

    def _format_embryo_id(self, surrogate_name: str, transfer_date: str, embryo_num: int) -> str:
        return f"{self._safe_identity_token(surrogate_name)}_{transfer_date}_embryo_{embryo_num}"
    
    def _collect_timeline_data(self):
        """Collect transfer and implantation data for timeline chart."""
        transfers_by_date = {}
        implantations_by_date = {}
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_id == FREEZER_TRANSFER_ID:
                continue  # Skip freezer transfers
                
            transfer_date_str = transfer_data.get('transfer_date')
            if not transfer_date_str:
                continue
                
            # Parse date (assuming ISO format from existing code)
            try:
                transfer_date = datetime.fromisoformat(transfer_date_str.replace('Z', '+00:00'))
                date_key = transfer_date.strftime('%Y-%m-%d')
                
                # Count transferred embryos
                embryo_count = len(transfer_data.get('embryos', []))
                transfers_by_date[date_key] = transfers_by_date.get(date_key, 0) + embryo_count
                
                # Count implanted embryos (use same transfer date)
                implanted_count = sum(1 for embryo in transfer_data.get('embryos', []) 
                                    if embryo.get('implanted', False))
                if implanted_count > 0:
                    implantations_by_date[date_key] = implantations_by_date.get(date_key, 0) + implanted_count
                    
            except (ValueError, AttributeError):
                continue  # Skip invalid dates
        
        return transfers_by_date, implantations_by_date
    
    def _render_timeline_chart(self):
        """Render the transfer/implantation timeline chart."""
        if not self.settings.get('show_timeline', False):
            return
            
        # Clear previous chart
        self.timeline_figure.clear()
        ax = self.timeline_figure.add_subplot(111)
        
        # Collect data
        transfers_by_date, implantations_by_date = self._collect_timeline_data()
        
        if not transfers_by_date and not implantations_by_date:
            ax.text(0.5, 0.5, 
                    self.messages.get("flow_track.timeline.no_data", "No transfer data available"),
                    ha='center', va='center', transform=ax.transAxes)
            self.timeline_canvas.draw()
            return
        
        # Sort dates and prepare cumulative data
        all_dates = sorted(set(transfers_by_date.keys()) | set(implantations_by_date.keys()))
        
        cumulative_transfers = []
        cumulative_implantations = []
        transfer_total = 0
        implantation_total = 0
        
        for date in all_dates:
            transfer_total += transfers_by_date.get(date, 0)
            implantation_total += implantations_by_date.get(date, 0)
            cumulative_transfers.append(transfer_total)
            cumulative_implantations.append(implantation_total)
        
        # Convert dates for matplotlib
        date_objects = [datetime.strptime(date, '%Y-%m-%d') for date in all_dates]
        
        # Create stair-step data for transfers
        if len(date_objects) > 0:
            # Create stair-step coordinates
            transfer_x = []
            transfer_y = []
            implantation_x = []
            implantation_y = []
            
            # Start from first date with zero
            if len(date_objects) > 0:
                transfer_x.extend([date_objects[0], date_objects[0]])
                transfer_y.extend([0, cumulative_transfers[0]])
                
                implantation_x.extend([date_objects[0], date_objects[0]])
                implantation_y.extend([0, cumulative_implantations[0]])
            
            # Add stair steps for each subsequent date
            for i in range(len(date_objects)):
                if i > 0:
                    # Horizontal line from previous date to current date
                    transfer_x.extend([date_objects[i-1], date_objects[i]])
                    transfer_y.extend([cumulative_transfers[i-1], cumulative_transfers[i-1]])
                    
                    # Vertical line up to new value
                    transfer_x.extend([date_objects[i], date_objects[i]])
                    transfer_y.extend([cumulative_transfers[i-1], cumulative_transfers[i]])
                    
                    # Same for implantations
                    implantation_x.extend([date_objects[i-1], date_objects[i]])
                    implantation_y.extend([cumulative_implantations[i-1], cumulative_implantations[i-1]])
                    
                    implantation_x.extend([date_objects[i], date_objects[i]])
                    implantation_y.extend([cumulative_implantations[i-1], cumulative_implantations[i]])
            
            # Plot stair-step lines
            ax.plot(transfer_x, transfer_y, 'k-', linewidth=2, 
                    label=self.messages.get("flow_track.timeline.transfers", "Transfers"))
            ax.plot(implantation_x, implantation_y, 'g-', linewidth=2,
                    label=self.messages.get("flow_track.timeline.implantations", "Implantations"))
        else:
            # No data case
            ax.plot([], [], 'k-', linewidth=2, 
                    label=self.messages.get("flow_track.timeline.transfers", "Transfers"))
            ax.plot([], [], 'g-', linewidth=2,
                    label=self.messages.get("flow_track.timeline.implantations", "Implantations"))
        
        # Auto-scale axes
        ax.relim()
        ax.autoscale_view()
        
        # Format x-axis for dates - only show dates with actual transfers
        import matplotlib.dates as mdates
        
        # Set ticks only for actual transfer dates (not automatic intervals)
        ax.set_xticks(date_objects)
        
        # Format date labels and rotate 45° CCW for better readability
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m.%Y'))
        import matplotlib.pyplot as plt
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Labels and legend
        ax.set_ylabel(self.messages.get("flow_track.timeline.count", "Cumulative Count"))
        ax.set_title(self.messages.get("flow_track.timeline.title", "Transfer Timeline"))
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        self.timeline_canvas.draw()
    
    def _toggle_timeline_visibility(self, state):
        """Toggle timeline chart visibility."""
        show_timeline = state == 2  # Qt.Checked
        self.settings['show_timeline'] = show_timeline
        self.timeline_widget.setVisible(show_timeline)
        
        if show_timeline:
            self._render_timeline_chart()
        
        self._save_settings()
    
    def _ensure_freezer_transfer_exists(self):
        """Ensure the freezer transfer exists (Flow_Track 3.0 section 7.1)."""
        if FREEZER_TRANSFER_ID not in self.manual_data.get('transfers_by_id', {}):
            self.manual_data['transfers_by_id'][FREEZER_TRANSFER_ID] = {
                'surrogate_name': FREEZER_NODE_NAME,
                'transfer_date': None,  # Freezer has no transfer date
                'embryos': []
            }
            logger.info("Initialized freezer transfer")
    
    def _open_freezer_menu(self):
        """Show freezer menu (Flow_Track 3.0 section 7.3 - only entry point for freezer actions)."""
        QtWidgets = self.parent_app.QtWidgets
        
        menu = QtWidgets.QMenu(self.widget)
        
        # Menu items per spec section 7.3
        open_action = menu.addAction(self.messages.get("flow_track.freezer.open", "Open Freezer"))
        open_action.triggered.connect(self._open_freezer_embryo_list)
        
        create_action = menu.addAction(self.messages.get("flow_track.freezer.create_embryo", "Create Embryo (Freezer)"))
        create_action.triggered.connect(self._create_freezer_embryo)
        
        refresh_action = menu.addAction(self.messages.get("flow_track.freezer.refresh", "Refresh"))
        refresh_action.triggered.connect(self._redraw_canvas)
        
        menu.addSeparator()
        
        # Toggle visibility
        toggle_text = self.messages.get("flow_track.freezer.hide", "Hide Freezer") if self.settings.get('freezer_visible', True) else self.messages.get("flow_track.freezer.show", "Show Freezer")
        toggle_action = menu.addAction(toggle_text)
        toggle_action.triggered.connect(self._toggle_freezer_visibility)
        
        # Show menu at freezer button position
        menu.exec(self.freezer_btn.mapToGlobal(self.freezer_btn.rect().bottomLeft()))
    
    def _toggle_freezer_visibility(self):
        """Toggle freezer embryo visibility (Flow_Track 3.0 section 7.2)."""
        self.settings['freezer_visible'] = not self.settings.get('freezer_visible', True)
        self._save_settings()
        self._redraw_canvas()
        logger.info(f"Freezer visibility: {self.settings['freezer_visible']}")
    
    def _open_freezer_embryo_list(self):
        """Show Freezer Embryo List dialog (Flow_Track 3.0 section 7.4)."""
        QtWidgets = self.parent_app.QtWidgets
        
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(self.messages.get("flow_track.freezer.list_title", "Freezer Embryos"))
        dialog.setModal(True)
        dialog.resize(600, 400)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Get freezer embryos
        freezer_transfer = self.manual_data.get('transfers_by_id', {}).get(FREEZER_TRANSFER_ID, {})
        embryos = freezer_transfer.get('embryos', [])
        
        # Info label
        info_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.list_info", f"Frozen embryos: {len(embryos)}"))
        layout.addWidget(info_label)
        
        # List widget
        embryo_list = QtWidgets.QListWidget()
        for embryo in embryos:
            unknown_text = self.messages.get('flow_track.unknown', 'Unknown')
            embryo_id = embryo.get('embryo_id', unknown_text)
            stage = self._translate_stage(embryo.get('stage', unknown_text))
            freeze_date = embryo.get('freeze_date', unknown_text)
            frozen_label = self.messages.get('flow_track.freezer.frozen_prefix', 'Frozen:')
            item_text = f"{embryo_id} - {stage} - {frozen_label} {freeze_date}"
            embryo_list.addItem(item_text)
        
        layout.addWidget(embryo_list)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        create_btn = QtWidgets.QPushButton(self.messages.get("flow_track.freezer.create_embryo", "Create Embryo"))
        create_btn.clicked.connect(lambda: self._create_freezer_embryo_from_dialog(dialog, info_label, embryo_list))
        button_layout.addWidget(create_btn)
        
        open_btn = QtWidgets.QPushButton(self.messages.get("flow_track.freezer.open_details", "Open Details"))
        open_btn.clicked.connect(lambda: self._open_freezer_embryo_details(embryo_list, embryos, dialog))
        button_layout.addWidget(open_btn)
        
        move_btn = QtWidgets.QPushButton(self.messages.get("flow_track.freezer.move_to_surrogate", "Move to Surrogate"))
        move_btn.clicked.connect(lambda: self._move_freezer_embryo_to_surrogate(embryo_list, embryos, dialog))
        button_layout.addWidget(move_btn)
        
        delete_btn = QtWidgets.QPushButton(self.messages.get("flow_track.freezer.delete", "Delete"))
        delete_btn.clicked.connect(lambda: self._delete_freezer_embryo(embryo_list, embryos, info_label))
        button_layout.addWidget(delete_btn)
        
        button_layout.addStretch()
        
        close_btn = QtWidgets.QPushButton(self.messages.get("button.close", "Close"))
        close_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        self._set_scrollable_dialog_layout(dialog, layout)
        dialog.exec()
    
    def _open_freezer_embryo_details(self, list_widget, embryos, parent_dialog):
        """Open embryo details dialog from freezer list."""
        current_row = list_widget.currentRow()
        if current_row < 0 or current_row >= len(embryos):
            return
        
        embryo = embryos[current_row]
        # Open standard embryo edit dialog
        self._edit_embryo(FREEZER_TRANSFER_ID, embryo.get('embryo_id'))
        parent_dialog.accept()
        self._redraw_canvas()
    
    def _move_freezer_embryo_to_surrogate(self, list_widget, embryos, parent_dialog):
        """Move embryo from freezer to surrogate transfer (Flow_Track 3.0 section 7.7)."""
        if not self._can('edit'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        current_row = list_widget.currentRow()
        if current_row < 0 or current_row >= len(embryos):
            QtWidgets.QMessageBox.warning(
                parent_dialog,
                self.messages.get("error.title", "Error"),
                self.messages.get("flow_track.freezer.select_embryo", "Please select an embryo")
            )
            return
        
        embryo = embryos[current_row]
        
        # Show transfer selection dialog
        transfer_dialog = QtWidgets.QDialog(parent_dialog)
        transfer_dialog.setWindowTitle(self.messages.get("flow_track.freezer.select_transfer", "Select Target Transfer"))
        transfer_dialog.setModal(True)
        
        t_layout = QtWidgets.QVBoxLayout()
        
        info_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.move_info", "Select surrogate transfer for embryo:") + f"\n{embryo.get('embryo_id')}")
        t_layout.addWidget(info_label)
        
        # List available transfers (most recent first per spec 7.4)
        transfer_list = QtWidgets.QListWidget()
        transfer_rows = []

        def _transfer_sort_key(transfer_data):
            raw_date = transfer_data.get('transfer_date')
            if isinstance(raw_date, datetime):
                return raw_date
            if isinstance(raw_date, str) and raw_date.strip():
                for parser in (
                    lambda value: datetime.fromisoformat(value.replace('Z', '+00:00')),
                    lambda value: datetime.strptime(value, DATE_FORMAT),
                    lambda value: datetime.strptime(value, '%Y-%m-%d'),
                ):
                    try:
                        parsed = parser(raw_date.strip())
                        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
                    except ValueError:
                        continue
            return datetime.min
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_id == FREEZER_TRANSFER_ID:
                continue  # Skip freezer itself

            surrogate_name = transfer_data.get('surrogate_name', 'Unknown')
            transfer_date = transfer_data.get('transfer_date', 'Unknown')
            item_text = f"{self._animal_export_name(surrogate_name)} - {transfer_date}"
            transfer_rows.append((_transfer_sort_key(transfer_data), item_text, transfer_id))

        available_transfers = []
        for _sort_key, item_text, transfer_id in sorted(transfer_rows, key=lambda row: row[0], reverse=True):
            transfer_list.addItem(item_text)
            available_transfers.append(transfer_id)

        t_layout.addWidget(transfer_list)
        
        button_layout = QtWidgets.QHBoxLayout()
        ok_btn = QtWidgets.QPushButton(self.messages.get("button.ok", "OK"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        ok_btn.clicked.connect(transfer_dialog.accept)
        cancel_btn.clicked.connect(transfer_dialog.reject)
        
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        t_layout.addLayout(button_layout)
        
        transfer_dialog.setLayout(t_layout)
        
        if transfer_dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            selected_row = transfer_list.currentRow()
            if selected_row < 0:
                return
            
            target_transfer_id = available_transfers[selected_row]
            
            # Move embryo (preserve cryopreserved=True and freeze_date per spec 7.7)
            freezer_transfer = self.manual_data['transfers_by_id'][FREEZER_TRANSFER_ID]
            target_transfer = self.manual_data['transfers_by_id'][target_transfer_id]
            
            # Remove from freezer
            freezer_transfer['embryos'].remove(embryo)
            
            # Add to target (preserving cryopreserved and freeze_date)
            target_transfer['embryos'].append(embryo)
            
            self._save_flow_track_data()
            parent_dialog.accept()
            self._redraw_canvas()
            
            logger.info(f"Moved embryo {embryo.get('embryo_id')} from freezer to {target_transfer_id}")
    
    def _delete_freezer_embryo(self, list_widget, embryos, info_label):
        """Delete embryo from freezer."""
        if not self._can('delete'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        current_row = list_widget.currentRow()
        if current_row < 0 or current_row >= len(embryos):
            return
        
        embryo = embryos[current_row]
        
        reply = QtWidgets.QMessageBox.question(
            self.widget,
            self.messages.get("flow_track.confirm_delete.title", "Confirm Delete"),
            self.messages.get("flow_track.confirm_delete.message", f"Delete embryo {embryo.get('embryo_id')}?"),
            QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
            QtWidgets.QMessageBox.StandardButton.No
        )
        
        if reply == QtWidgets.QMessageBox.StandardButton.Yes:
            freezer_transfer = self.manual_data['transfers_by_id'][FREEZER_TRANSFER_ID]
            freezer_transfer['embryos'].remove(embryo)
            list_widget.takeItem(current_row)
            frozen_embryos_text = self.messages.get("flow_track.freezer.list_info", "Frozen embryos: {count}")
            info_label.setText(frozen_embryos_text.format(count=len(freezer_transfer['embryos'])))
            self._save_flow_track_data()
            self._redraw_canvas()
            logger.info(f"Deleted embryo {embryo.get('embryo_id')} from freezer")
    
    def _create_freezer_embryo_from_dialog(self, parent_dialog, info_label, list_widget):
        """Create embryo from within Freezer Embryos dialog and refresh the list."""
        if not self._can('create'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        dialog = QtWidgets.QDialog(parent_dialog)
        dialog.setWindowTitle(self.messages.get("flow_track.freezer.create_title", "Create Embryo (Freezer)"))
        dialog.setModal(True)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Info label
        info = QtWidgets.QLabel(self.messages.get("flow_track.freezer.create_info", 
            "Create embryo with minimal data. Full details can be edited afterwards."))
        info.setWordWrap(True)
        layout.addWidget(info)
        layout.addSpacing(10)
        
        # Freeze date (required)
        freeze_date_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.freeze_date", "Freeze Date (required):"))
        layout.addWidget(freeze_date_label)
        
        freeze_date_edit = QtWidgets.QDateEdit()
        freeze_date_edit.setCalendarPopup(True)
        freeze_date_edit.setDate(self.parent_app.QtCore.QDate.currentDate())
        layout.addWidget(freeze_date_edit)
        
        layout.addSpacing(10)
        
        # Embryo ID (auto-proposed, editable)
        embryo_id_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.embryo_id", "Embryo ID (editable):"))
        layout.addWidget(embryo_id_label)
        
        # Generate proposed ID
        freeze_date_str = freeze_date_edit.date().toString("yyyy-MM-dd")
        proposed_id = self._generate_next_embryo_id(FREEZER_TRANSFER_ID, freeze_date_str)
        
        embryo_id_edit = QtWidgets.QLineEdit(proposed_id)
        layout.addWidget(embryo_id_edit)
        
        layout.addSpacing(20)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        create_btn = QtWidgets.QPushButton(self.messages.get("button.create", "Create"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        create_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(create_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            embryo_id = embryo_id_edit.text().strip()
            
            # Validate: ID must be non-empty and unique
            if not embryo_id:
                QtWidgets.QMessageBox.warning(
                    parent_dialog,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.embryo_id_empty", "Embryo ID cannot be empty")
                )
                return
            
            if not self._is_embryo_id_unique(embryo_id):
                QtWidgets.QMessageBox.warning(
                    parent_dialog,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.embryo_id_duplicate", "Embryo ID must be unique")
                )
                return
            
            freeze_date_str = freeze_date_edit.date().toString("yyyy-MM-dd")
            
            # Create embryo with minimal data
            new_embryo = {
                'embryo_id': embryo_id,
                'egg_donor_name': None,
                'egg_donation_date': None,
                'sperm_donor_name': None,
                'sperm_donation_id': None,
                'stage': STAGE_IN_VITRO_M2,
                'implanted': False,
                'cryopreserved': True,
                'freeze_date': freeze_date_str,
                'comment': ''
            }
            
            # Add to freezer
            freezer_transfer = self.manual_data['transfers_by_id'][FREEZER_TRANSFER_ID]
            freezer_transfer['embryos'].append(new_embryo)
            
            self._save_flow_track_data()
            self._redraw_canvas()
            
            # Refresh the list widget
            embryos = freezer_transfer.get('embryos', [])
            list_widget.clear()
            for embryo in embryos:
                embryo_id_display = embryo.get('embryo_id', 'Unknown')
                stage = embryo.get('stage', 'Unknown')
                freeze_date = embryo.get('freeze_date', 'Unknown')
                item_text = f"{embryo_id_display} - {stage} - Frozen: {freeze_date}"
                list_widget.addItem(item_text)
            
            info_label.setText(self.messages.get("flow_track.freezer.list_info", f"Frozen embryos: {len(embryos)}"))
            
            logger.info(f"Created freezer embryo {embryo_id}")
    
    def _create_freezer_embryo(self):
        """Create embryo directly in freezer (Flow_Track 3.0 section 7.5)."""
        if not self._can('create'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(self.messages.get("flow_track.freezer.create_title", "Create Embryo (Freezer)"))
        dialog.setModal(True)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Info label
        info_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.create_info", 
            "Create embryo with minimal data. Full details can be edited afterwards."))
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        layout.addSpacing(10)
        
        # Freeze date (required)
        freeze_date_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.freeze_date", "Freeze Date (required):"))
        layout.addWidget(freeze_date_label)
        
        freeze_date_edit = QtWidgets.QDateEdit()
        freeze_date_edit.setCalendarPopup(True)
        freeze_date_edit.setDate(self.parent_app.QtCore.QDate.currentDate())
        layout.addWidget(freeze_date_edit)
        
        layout.addSpacing(10)
        
        # Embryo ID (auto-proposed, editable)
        embryo_id_label = QtWidgets.QLabel(self.messages.get("flow_track.freezer.embryo_id", "Embryo ID (editable):"))
        layout.addWidget(embryo_id_label)
        
        # Generate proposed ID
        freeze_date_str = freeze_date_edit.date().toString("yyyy-MM-dd")
        proposed_id = self._generate_next_embryo_id(FREEZER_TRANSFER_ID, freeze_date_str)
        
        embryo_id_edit = QtWidgets.QLineEdit(proposed_id)
        layout.addWidget(embryo_id_edit)
        
        layout.addSpacing(20)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        create_btn = QtWidgets.QPushButton(self.messages.get("button.create", "Create"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        create_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(create_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            embryo_id = embryo_id_edit.text().strip()
            
            # Validate: ID must be non-empty and unique
            if not embryo_id:
                QtWidgets.QMessageBox.warning(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.embryo_id_empty", "Embryo ID cannot be empty")
                )
                return
            
            if not self._is_embryo_id_unique(embryo_id):
                QtWidgets.QMessageBox.warning(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.embryo_id_duplicate", "Embryo ID must be unique")
                )
                return
            
            freeze_date_str = freeze_date_edit.date().toString("yyyy-MM-dd")
            
            # Create embryo with minimal data (per spec 7.5)
            new_embryo = {
                'embryo_id': embryo_id,
                'egg_donor_name': None,
                'egg_donation_date': None,
                'sperm_donor_name': None,
                'sperm_donation_id': None,
                'stage': STAGE_IN_VITRO_M2,  # Default stage
                'implanted': False,
                'cryopreserved': True,
                'freeze_date': freeze_date_str,
                'comment': ''
            }
            
            # Add to freezer
            freezer_transfer = self.manual_data['transfers_by_id'][FREEZER_TRANSFER_ID]
            freezer_transfer['embryos'].append(new_embryo)
            
            self._save_flow_track_data()
            self._redraw_canvas()
            
            # Inform user to edit full details
            QtWidgets.QMessageBox.information(
                self.widget,
                self.messages.get("flow_track.freezer.created_title", "Embryo Created"),
                self.messages.get("flow_track.freezer.created_message", 
                    "Embryo created. Open Freezer to edit full details (donors, stage, etc.)")
            )
            
            logger.info(f"Created freezer embryo {embryo_id}")
    
    def _populate_events_from_progtrack(self):
        """
        Populate events from ProgTrack's animal data.
        Extracts surgeries, sperm donations, and transfers.
        """
        # Access Role enum from parent_app
        Role = self.parent_app.Role
        
        self.egg_surgeries = []
        self.sperm_donations = []
        self.embryo_transfers = []
        
        detected_surgery_ids = set()
        detected_sperm_ids = set()
        detected_transfer_ids = set()
        
        # Scan animals
        for name, rec in self.parent_app.animals.items():
            rolle = _animal_role_value(rec)
            
            # Egg donation surgeries (SPENDER role)
            if rolle == Role.SPENDER.value:
                # From unified events
                for ev in rec.get('events', []):
                    if ev.get('typ') == 'surgery':
                        event_date = _flow_datetime(ev.get('datum'))
                        if event_date is None:
                            continue
                        surgery_id = f"surgery_{name}_{event_date.isoformat()}"
                        if surgery_id not in detected_surgery_ids:
                            self.egg_surgeries.append({
                                'id': surgery_id,
                                'animal_name': name,
                                'date': event_date,
                                'event_type': 'egg_donation_surgery'
                            })
                            detected_surgery_ids.add(surgery_id)
                
            # Sperm donations (SAMENSP role)
            if rolle == Role.SAMENSP.value:
                for sperm_entry in rec.get('sperm', []):
                    event_date = _flow_datetime(sperm_entry.get('datum'))
                    if event_date is None:
                        continue
                    sperm_id = f"sperm_{name}_{event_date.isoformat()}"
                    if sperm_id not in detected_sperm_ids:
                        self.sperm_donations.append({
                            'id': sperm_id,
                            'animal_name': name,
                            'date': event_date,
                            'event_type': 'sperm_donation',
                            'motility': sperm_entry.get('motility'),
                            'progressive': sperm_entry.get('progressive'),
                            'count': sperm_entry.get('count')
                        })
                        detected_sperm_ids.add(sperm_id)
            
            # Embryo transfers (AMME role)
            if rolle == Role.AMME.value:
                # From unified events
                for ev in rec.get('events', []):
                    if ev.get('typ') == 'embryo_transfer':
                        event_date = _flow_datetime(ev.get('datum'))
                        if event_date is None:
                            continue
                        transfer_id = f"transfer_{name}_{event_date.isoformat()}"
                        if transfer_id not in detected_transfer_ids:
                            self.embryo_transfers.append({
                                'id': transfer_id,
                                'animal_name': name,
                                'date': event_date,
                                'event_type': 'embryo_transfer'
                            })
                            detected_transfer_ids.add(transfer_id)
                
        # Sort by date
        self.egg_surgeries.sort(key=lambda x: x['date'])
        self.sperm_donations.sort(key=lambda x: x['date'])
        self.embryo_transfers.sort(key=lambda x: x['date'])
        
        logger.info(f"Detected {len(self.egg_surgeries)} surgeries, "
                   f"{len(self.sperm_donations)} sperm donations, "
                   f"{len(self.embryo_transfers)} transfers")

    def _get_selected_scope_animals(self) -> set[str]:
        """Return selected animal names that currently exist in the dataset."""
        selected = getattr(self.parent_app, 'selected_animals', None)
        if not selected:
            return set()
        return {name for name in selected if name in self.parent_app.animals}

    def _to_date_key(self, value: Any) -> Optional[str]:
        """Normalize datetime-like values to a YYYY-MM-DD key for de-duplication."""
        if isinstance(value, datetime):
            return value.date().isoformat()

        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                return datetime.fromisoformat(text.replace('Z', '+00:00')).date().isoformat()
            except ValueError:
                return text.split('T', 1)[0]

        if hasattr(value, 'isoformat'):
            try:
                iso_text = value.isoformat()
                if isinstance(iso_text, str) and iso_text:
                    return iso_text.split('T', 1)[0]
            except Exception:
                return None

        return None

    def _is_embryo_connected_to_selection(
        self,
        embryo: Dict[str, Any],
        surrogate_name: str,
        selected_animals: set[str]
    ) -> bool:
        """Check whether an embryo belongs to the currently selected context."""
        if not selected_animals:
            return True
        if surrogate_name in selected_animals:
            return True

        egg_donor = embryo.get('egg_donor_name')
        sperm_donor = embryo.get('sperm_donor_name')
        return egg_donor in selected_animals or sperm_donor in selected_animals

    def _get_flow_animals_to_show(self) -> set[str]:
        """Return animals that should participate in Flow Track rendering.

        Full-graph mode shows all animals. In selected-only mode, include the
        selected animals plus context nodes connected through selected embryos:
        surrogate nodes (ammen) and sperm donors.
        """
        if self.settings.get('render_full_graph', True):
            return set(self.parent_app.animals.keys())

        selected_animals = self._get_selected_scope_animals()
        if not selected_animals:
            return set(self.parent_app.animals.keys())

        animals_to_show = set(selected_animals)

        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_id == FREEZER_TRANSFER_ID:
                continue

            surrogate_name = transfer_data.get('surrogate_name')
            embryos = transfer_data.get('embryos', [])
            connected_embryos = [
                embryo for embryo in embryos
                if self._is_embryo_connected_to_selection(embryo, surrogate_name, selected_animals)
            ]

            if not connected_embryos:
                continue

            if surrogate_name in self.parent_app.animals:
                animals_to_show.add(surrogate_name)

            for embryo in connected_embryos:
                sperm_donor = embryo.get('sperm_donor_name')
                if sperm_donor in self.parent_app.animals:
                    animals_to_show.add(sperm_donor)

        return animals_to_show

    def _get_transfers_for_surrogate(self, surrogate: str) -> List[Dict[str, Any]]:
        """Collect transfers for a surrogate from events plus persisted manual data."""
        transfers: List[Dict[str, Any]] = []
        seen_transfer_ids: set[str] = set()
        seen_transfer_date_keys: set[str] = set()

        # Primary source: transfers parsed from ProgTrack events
        for transfer in self.embryo_transfers:
            if transfer.get('animal_name') != surrogate:
                continue

            transfer_id = transfer.get('id')
            transfer_date_key = self._to_date_key(transfer.get('date'))

            if transfer_id and transfer_id in seen_transfer_ids:
                continue
            if transfer_date_key and transfer_date_key in seen_transfer_date_keys:
                continue

            if transfer_id:
                seen_transfer_ids.add(transfer_id)
            if transfer_date_key:
                seen_transfer_date_keys.add(transfer_date_key)
            transfers.append(transfer)

        # Fallback source: persisted manual transfers (covers event desync cases)
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_id == FREEZER_TRANSFER_ID:
                continue
            if transfer_data.get('surrogate_name') != surrogate:
                continue
            if transfer_id in seen_transfer_ids:
                continue

            transfer_date = transfer_data.get('transfer_date')
            parsed_date = None
            if isinstance(transfer_date, datetime):
                parsed_date = transfer_date
            elif isinstance(transfer_date, str) and transfer_date:
                try:
                    parsed_date = datetime.fromisoformat(transfer_date.replace('Z', '+00:00'))
                except ValueError:
                    parsed_date = None

            transfer_date_key = self._to_date_key(parsed_date) or self._to_date_key(transfer_date)
            if transfer_date_key and transfer_date_key in seen_transfer_date_keys:
                continue

            transfers.append({
                'id': transfer_id,
                'animal_name': surrogate,
                'date': parsed_date,
                'event_type': 'embryo_transfer'
            })
            seen_transfer_ids.add(transfer_id)
            if transfer_date_key:
                seen_transfer_date_keys.add(transfer_date_key)

        transfers.sort(key=lambda x: x.get('date') or datetime.min)
        return transfers
    
    def _redraw_canvas(self):
        """Redraw the entire canvas with current data."""
        # Check if database has any animals at all
        has_animals = hasattr(self.parent_app, 'animals') and bool(self.parent_app.animals)
        
        if not has_animals:
            # No animals in database - show splash
            self.stack.setCurrentWidget(self.splash_widget)
            return

        # Keep event-derived transfers in sync with current ProgTrack records.
        # This ensures newly saved embryo_transfer events immediately appear.
        self._populate_events_from_progtrack()
        
        # Show content widget (we have animals to display)
        self.stack.setCurrentWidget(self.content_widget)
        
        self.ax.clear()
        
        # Clear lifebar data to prevent stale annotation references
        self.lifebar_data.clear()
        
        # Calculate or use stored view limits
        if self.current_xlim is None or self.current_ylim is None:
            # Initial view - calculate bounds from content
            bounds = self._calculate_content_bounds()
            if bounds:
                xlim = (bounds[0], bounds[1])
                ylim = (bounds[2], bounds[3])
                # Apply aspect fill to eliminate gutters
                xlim, ylim = self._apply_aspect_fill(xlim, ylim)
                self.current_xlim = xlim
                self.current_ylim = ylim
            else:
                # No content fallback - minimal default view
                self.current_xlim = (0, 1)
                self.current_ylim = (0, 1)
        else:
            # Re-apply aspect fill on existing limits (handles window resize)
            self.current_xlim, self.current_ylim = self._apply_aspect_fill(
                self.current_xlim, self.current_ylim
            )
        
        # Apply view limits
        self.ax.set_xlim(self.current_xlim)
        self.ax.set_ylim(self.current_ylim)
        
        # Use adjustable='datalim' to fill canvas without gutters
        self.ax.set_aspect('equal', adjustable='datalim')
        self.ax.axis('off')
        
        # Draw grid if enabled (dynamically based on current view)
        if self.settings.get('show_grid', False):
            self._draw_grid()
        
        # DON'T clear temp positions - they should persist after dragging
        # Only clear artist registries for re-registration
        self.node_artists.clear()
        self.embryo_artists.clear()
        
        # Render the flow track visualization
        self._render_flow_track()
        
        # Update timeline if visible
        if self.settings.get('show_timeline', False):
            self._render_timeline_chart()
        
        self.canvas.draw()
    
    def _render_flow_track(self):
        """Render the complete flow track visualization."""
        # Layout constants - tight to edges to maximize canvas usage
        EGG_RAIL_X = 0.02
        SPERM_RAIL_X = 0.98
        SURROGATE_RAIL_Y = 0.02
        
        
        # Determine which animals to show
        animals_to_show = self._get_flow_animals_to_show()
        render_full_graph = self.settings.get('render_full_graph', True)
        selected_scope_animals = set()
        if not render_full_graph:
            selected_scope_animals = self._get_selected_scope_animals()

        if render_full_graph:
            logger.info(f"Flow Track: Rendering full graph with {len(animals_to_show)} animals")
        else:
            logger.info(
                "Flow Track: Rendering selected context with %d animals (selected=%d)",
                len(animals_to_show),
                len(selected_scope_animals)
            )
        
        Role = self.parent_app.Role
        
        # Categorize animals by role
        egg_donors = []
        sperm_donors = []
        surrogates = []
        
        for animal_name in animals_to_show:
            animal_data = self.parent_app.animals.get(animal_name, {})
            role = _animal_role_value(animal_data)
            
            if role == Role.SPENDER.value:
                egg_donors.append(animal_name)
            elif role == Role.SAMENSP.value:
                sperm_donors.append(animal_name)
            elif role == Role.AMME.value:
                surrogates.append(animal_name)
        
        # Calculate positions
        positions = {}
        
        # Egg donors: vertical distribution on left rail - expanded range
        n_egg = len(egg_donors)
        for i, donor in enumerate(egg_donors):
            y = 0.1 + (i / max(1, n_egg - 1)) * 0.8 if n_egg > 1 else 0.5
            positions[donor] = (EGG_RAIL_X, y)
        
        # Sperm donors: vertical distribution on right rail - expanded range
        n_sperm = len(sperm_donors)
        for i, donor in enumerate(sperm_donors):
            y = 0.1 + (i / max(1, n_sperm - 1)) * 0.8 if n_sperm > 1 else 0.5
            positions[donor] = (SPERM_RAIL_X, y)
        
        # Surrogates: horizontal distribution on bottom rail - expanded range
        n_surr = len(surrogates)
        for i, surrogate in enumerate(surrogates):
            x = 0.1 + (i / max(1, n_surr - 1)) * 0.8 if n_surr > 1 else 0.5
            positions[surrogate] = (x, SURROGATE_RAIL_Y)
        
        # Freezer node (v3.0): always add, positioned at top center
        if self.settings.get('freezer_visible', True):
            positions[FREEZER_NODE_NAME] = (0.5, 0.98)
        
        # Draw animal nodes (including freezer)
        for animal_name, (x, y) in positions.items():
            # Check if there's a temporary dragged position
            if ('animal', animal_name) in self.temp_positions:
                x, y = self.temp_positions[('animal', animal_name)]
                positions[animal_name] = (x, y)  # Update positions dict for connections
            
            # Special handling for freezer node (v3.0)
            if animal_name == FREEZER_NODE_NAME:
                color = 'cyan'
                role = None
            else:
                animal_record = self.parent_app.animals.get(animal_name, {})
                role = _animal_role_value(animal_record)
                color = role_color_for_record(
                    animal_record,
                    getattr(self.parent_app, "animal_role_registry", None),
                )
            
            # Draw node (clip_on=False allows drawing beyond axis limits)
            artist, = self.ax.plot(x, y, 'o', markersize=12, 
                                  markerfacecolor=color,
                                  markeredgecolor='black',
                                  markeredgewidth=1.5,
                                  picker=10, zorder=20, clip_on=False)
            
            # Register for picking
            self.node_artists[artist] = animal_name
            
            # Draw lifebars for this animal (NEW in v2.0)
            num_lifebars = 0
            if self.settings.get('show_lifebars', True):
                num_lifebars = self._render_lifebars_for_animal(animal_name, x, y, role)
            
            # Add label below lifebars if enabled in settings
            if self.settings.get('show_animal_names', True):
                # Special label for freezer
                if animal_name == FREEZER_NODE_NAME:
                    label_text = self._translate_freezer_label()
                    y_offset_points = 10  # Above the node for freezer
                    va = 'bottom'
                else:
                    label_text = self._animal_export_name(animal_name)
                    # Calculate offset: negative values go downward
                    if num_lifebars > 0:
                        # Initial offset - (bars * spacing) - gap = further down
                        y_offset_points = -(5 + num_lifebars * 4 + 5)
                    else:
                        y_offset_points = -10  # Standard offset when no lifebars
                    va = 'top'
                
                # Use annotation for consistent offset behavior
                self.ax.annotate(
                    label_text,
                    xy=(x, y),
                    xytext=(0, y_offset_points),
                    textcoords='offset points',
                    ha='center',
                    va=va,
                    fontsize=9,
                    weight='bold',
                    clip_on=False
                )
        
        # Draw embryo transfers for selected surrogates
        for surrogate in surrogates:
            # Find transfers for this surrogate
            transfers = self._get_transfers_for_surrogate(surrogate)

            if selected_scope_animals:
                filtered_transfers = []
                for transfer in transfers:
                    transfer_id = transfer.get('id')
                    if surrogate in selected_scope_animals:
                        filtered_transfers.append(transfer)
                        continue
                    transfer_data = self.manual_data.get('transfers_by_id', {}).get(transfer_id, {})
                    embryos = transfer_data.get('embryos', [])
                    has_connected_embryo = any(
                        self._is_embryo_connected_to_selection(embryo, surrogate, selected_scope_animals)
                        for embryo in embryos
                    )
                    if has_connected_embryo:
                        filtered_transfers.append(transfer)
                transfers = filtered_transfers

            if not transfers:
                continue
            
            # Calculate positions for each transfer group around the surrogate
            surrogate_x, surrogate_y = positions[surrogate]
            n_transfers = len(transfers)
            
            # Check if embryos are visible for this surrogate
            surrogate_embryos_visible = self.embryo_visibility.get(surrogate, True)
            
            for transfer_idx, transfer in enumerate(transfers):
                transfer_id = transfer['id']
                transfer_date = transfer.get('date')
                if not isinstance(transfer_date, datetime):
                    transfer_date = None
                
                # Check if user has entered embryo count manually (always preserve user input)
                user_embryo_count = None
                if surrogate in self.manual_data:
                    if 'transfers' in self.manual_data[surrogate]:
                        if transfer_id in self.manual_data[surrogate]['transfers']:
                            user_embryo_count = self.manual_data[surrogate]['transfers'][transfer_id].get('embryo_count')
                
                # Also check if already exists in transfers_by_id and sync the count
                if transfer_id in self.manual_data.get('transfers_by_id', {}):
                    existing_count = len(self.manual_data['transfers_by_id'][transfer_id].get('embryos', []))
                    # If no user count but transfer exists, preserve existing count
                    if user_embryo_count is None and existing_count > 0:
                        user_embryo_count = existing_count
                
                # Get or create transfer data (v3.0: from manual_data)
                if transfer_id not in self.manual_data.get('transfers_by_id', {}):
                    # Determine number of embryos to create
                    if user_embryo_count is not None:
                        # Use user-specified count
                        num_embryos = user_embryo_count
                    else:
                        # Use auto-suggestions, but limit by max_suggestions setting
                        suggestions = self._generate_suggestions(transfer_date, surrogate) if transfer_date else []
                        num_embryos = min(len(suggestions), self.settings.get('max_suggestions_per_transfer', 3))
                        if num_embryos == 0:
                            num_embryos = 1  # At least one embryo
                    
                    # Generate suggestions for donor assignment
                    suggestions = self._generate_suggestions(transfer_date, surrogate) if transfer_date else []
                    
                    embryos = []
                    # Format date as yyyy-mm-dd
                    date_str = transfer_date.strftime('%Y-%m-%d') if transfer_date else transfer_id
                    for i in range(num_embryos):
                        # Try to use suggestion if available
                        if i < len(suggestions):
                            egg_donor, sperm_donor, confidence = suggestions[i]
                        else:
                            egg_donor, sperm_donor, confidence = None, None, 0
                        
                        # Embryo numbering starts at 1
                        embryo_num = i + 1
                        
                        # Determine egg_donation_date from egg donor's surgery events
                        egg_donation_date = None
                        if egg_donor:
                            donor_data = self.parent_app.animals.get(egg_donor, {})
                            surgeries = []
                            for ev in donor_data.get('events', []):
                                if ev.get('typ') == 'surgery':
                                    event_date = _flow_datetime(ev.get('datum'))
                                    if event_date is not None:
                                        surgeries.append(event_date)
                            
                            # Find closest surgery before transfer
                            for surg_date in sorted(surgeries, reverse=True):
                                if transfer_date and surg_date <= transfer_date:
                                    if hasattr(surg_date, 'date'):
                                        egg_donation_date = surg_date.date().isoformat()
                                    elif hasattr(surg_date, 'isoformat'):
                                        egg_donation_date = surg_date.isoformat()
                                    else:
                                        egg_donation_date = str(surg_date)
                                    break
                        
                        embryos.append({
                            'embryo_id': self._format_embryo_id(surrogate, date_str, embryo_num),
                            'egg_donor_name': egg_donor,
                            'egg_donation_date': egg_donation_date,
                            'sperm_donor_name': sperm_donor,
                            'sperm_donation_id': None,
                            'stage': STAGE_IN_VITRO_M2,  # v3.0: default stage - IVM M2
                            'implanted': False,
                            'cryopreserved': False,
                            'freeze_date': None,  # v3.0: required field
                            'comment': '',  # v3.0: required field
                            'confidence': confidence
                        })
                    
                    # v3.0: Store in manual_data structure
                    self.manual_data['transfers_by_id'][transfer_id] = {
                        'surrogate_name': surrogate,
                        'transfer_date': transfer_date.isoformat() if transfer_date else None,
                        'embryos': embryos
                    }
                else:
                    # Transfer exists - check if we need to adjust embryo count
                    transfer_data = self.manual_data['transfers_by_id'][transfer_id]
                    current_embryos = transfer_data.get('embryos', [])
                    
                    if user_embryo_count is not None and len(current_embryos) != user_embryo_count:
                        # User changed embryo count - adjust
                        if user_embryo_count > len(current_embryos):
                            # Add more embryos
                            suggestions = self._generate_suggestions(transfer_date, surrogate) if transfer_date else []
                            # Format date as yyyy-mm-dd
                            date_str = transfer_date.strftime('%Y-%m-%d') if transfer_date else transfer_id
                            for i in range(len(current_embryos), user_embryo_count):
                                if i < len(suggestions):
                                    egg_donor, sperm_donor, confidence = suggestions[i]
                                else:
                                    egg_donor, sperm_donor, confidence = None, None, 0
                                
                                # Embryo numbering starts at 1
                                embryo_num = i + 1
                                
                                # Determine egg_donation_date from egg donor's surgery events
                                egg_donation_date = None
                                if egg_donor:
                                    donor_data = self.parent_app.animals.get(egg_donor, {})
                                    surgeries = []
                                    for ev in donor_data.get('events', []):
                                        if ev.get('typ') == 'surgery':
                                            event_date = _flow_datetime(ev.get('datum'))
                                            if event_date is not None:
                                                surgeries.append(event_date)

                                    # Find closest surgery before transfer
                                    for surg_date in sorted(surgeries, reverse=True):
                                        if transfer_date and surg_date <= transfer_date:
                                            if hasattr(surg_date, 'date'):
                                                egg_donation_date = surg_date.date().isoformat()
                                            elif hasattr(surg_date, 'isoformat'):
                                                egg_donation_date = surg_date.isoformat()
                                            else:
                                                egg_donation_date = str(surg_date)
                                            break
                                
                                current_embryos.append({
                                    'embryo_id': self._format_embryo_id(surrogate, date_str, embryo_num),
                                    'egg_donor_name': egg_donor,
                                    'egg_donation_date': egg_donation_date,
                                    'sperm_donor_name': sperm_donor,
                                    'sperm_donation_id': None,
                                    'stage': STAGE_IN_VITRO_M2,  # v3.0: default stage - IVM M2
                                    'implanted': False,
                                    'cryopreserved': False,
                                    'freeze_date': None,  # v3.0: required field
                                    'comment': '',  # v3.0: required field
                                    'confidence': confidence
                                })
                        else:
                            # Remove excess embryos
                            transfer_data['embryos'] = current_embryos[:user_embryo_count]
                
                transfer_data = self.manual_data['transfers_by_id'][transfer_id]
                embryos = transfer_data.get('embryos', [])

                if selected_scope_animals:
                    embryos = [
                        embryo for embryo in embryos
                        if self._is_embryo_connected_to_selection(embryo, surrogate, selected_scope_animals)
                    ]
                    if not embryos:
                        continue
                
                # Position transfer groups in an arc around the surrogate
                # Each transfer gets its own position offset
                if n_transfers == 1:
                    # Single transfer - center above surrogate
                    group_offset_x = 0
                    group_offset_y = 0.12  # Increased from 0.08 for more vertical space
                else:
                    # Multiple transfers - distribute in arc with more separation
                    angle_range = 90  # degrees - increased from 60 for wider arc
                    angle_step = angle_range / max(1, n_transfers - 1)
                    angle = -angle_range/2 + transfer_idx * angle_step
                    radius = 0.18  # Increased from 0.1 for more separation between groups
                    
                    import math
                    group_offset_x = radius * math.sin(math.radians(angle))
                    group_offset_y = radius * math.cos(math.radians(angle))
                
                # Position embryos based on visibility
                n_embryos = len(embryos)
                
                if surrogate_embryos_visible:
                    # Visible: arc above surrogate
                    embryo_y = surrogate_y + group_offset_y
                    embryo_spacing = 0.04
                    start_x = surrogate_x + group_offset_x - (n_embryos - 1) * embryo_spacing / 2
                    
                    for i, embryo in enumerate(embryos):
                        x = start_x + i * embryo_spacing
                        embryo_id = embryo['embryo_id']
                        
                        # Check for dragged position
                        if ('embryo', embryo_id) in self.temp_positions:
                            x, embryo_y = self.temp_positions[('embryo', embryo_id)]
                        
                        embryo_pos = (x, embryo_y)
                        
                        # Draw connections to donors (always visible)
                        egg_donor = embryo.get('egg_donor_name')
                        sperm_donor = embryo.get('sperm_donor_name')
                        
                        if egg_donor and egg_donor in positions:
                            donor_pos = positions[egg_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'deeppink', 
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
                        
                        if sperm_donor and sperm_donor in positions:
                            donor_pos = positions[sperm_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'black',
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
                        
                        # Connection to surrogate (only when visible)
                        self._draw_bezier_curve(embryo_pos, (surrogate_x, surrogate_y),
                                               'mediumpurple', linestyle='-', linewidth=1.5, alpha=0.5)
                        
                        # Circle appearance based on implantation and cryopreservation
                        implanted = embryo.get('implanted', embryo.get('pregnant', False))
                        cryopreserved = embryo.get('cryopreserved', False)
                        
                        if implanted:
                            facecolor = 'green'
                            edgecolor = 'darkgreen'
                        elif cryopreserved:
                            facecolor = 'none'
                            edgecolor = 'blue'
                        else:
                            facecolor = 'none'
                            edgecolor = 'gray'
                        
                        # Draw embryo circle (only when visible)
                        artist, = self.ax.plot(x, embryo_y, 'o', markersize=10,
                                              markerfacecolor=facecolor,
                                              markeredgecolor=edgecolor,
                                              markeredgewidth=2,
                                              picker=5, zorder=15, clip_on=False)
                        
                        # Add snowflake symbol if cryopreserved
                        if cryopreserved:
                            snowflake_color = 'white' if implanted else 'blue'
                            self.ax.text(x, embryo_y, 'Frozen', ha='center', va='center',
                                        fontsize=8, color=snowflake_color,
                                        weight='bold', zorder=16, clip_on=False)
                        
                        # Register for picking
                        self.embryo_artists[artist] = (transfer_id, embryo_id)
                else:
                    # Hidden: all embryos at surrogate position (behind surrogate dot)
                    for embryo in embryos:
                        embryo_id = embryo['embryo_id']
                        embryo_pos = (surrogate_x, surrogate_y)
                        
                        # Draw connections to donors (always visible even when embryos hidden)
                        egg_donor = embryo.get('egg_donor_name')
                        sperm_donor = embryo.get('sperm_donor_name')
                        
                        if egg_donor and egg_donor in positions:
                            donor_pos = positions[egg_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'deeppink', 
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
                        
                        if sperm_donor and sperm_donor in positions:
                            donor_pos = positions[sperm_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'black',
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
        
        # === DRAW FREEZER EMBRYOS (always draw connections, hide dots when toggled off) ===
        if FREEZER_NODE_NAME in positions:
            freezer_x, freezer_y = positions[FREEZER_NODE_NAME]
            freezer_transfer = self.manual_data.get('transfers_by_id', {}).get(FREEZER_TRANSFER_ID)
            
            if freezer_transfer:
                embryos = freezer_transfer.get('embryos', [])
                n_embryos = len(embryos)
                
                if n_embryos > 0:
                    # Check if embryos are visible for freezer
                    freezer_embryos_visible = self.embryo_visibility.get(FREEZER_NODE_NAME, False)
                    
                    # Position embryos based on visibility
                    if freezer_embryos_visible:
                        # Visible: arc below freezer node
                        embryo_spacing = 0.04
                        start_x = freezer_x - (n_embryos - 1) * embryo_spacing / 2
                        embryo_y = freezer_y - 0.12
                    else:
                        # Hidden: all embryos at freezer position (behind freezer dot)
                        start_x = freezer_x
                        embryo_y = freezer_y
                        embryo_spacing = 0  # All at same position
                    
                    for i, embryo in enumerate(embryos):
                        if self.embryo_visibility.get(FREEZER_NODE_NAME, False):
                            x = start_x + i * embryo_spacing
                        else:
                            x = freezer_x  # Behind freezer
                        
                        embryo_id = embryo['embryo_id']
                        
                        # Check for dragged position (only when visible)
                        if self.embryo_visibility.get(FREEZER_NODE_NAME, False) and ('embryo', embryo_id) in self.temp_positions:
                            x, embryo_y = self.temp_positions[('embryo', embryo_id)]
                        
                        embryo_pos = (x, embryo_y)
                        
                        # Draw connections to donors (always visible)
                        egg_donor = embryo.get('egg_donor_name')
                        sperm_donor = embryo.get('sperm_donor_name')
                        
                        if egg_donor and egg_donor in positions:
                            donor_pos = positions[egg_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'deeppink',
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
                        
                        if sperm_donor and sperm_donor in positions:
                            donor_pos = positions[sperm_donor]
                            self._draw_bezier_curve(embryo_pos, donor_pos, 'black',
                                                   linestyle='-', linewidth=1.5, alpha=0.6)
                        
                        # Connection to freezer node (only when visible)
                        if self.embryo_visibility.get(FREEZER_NODE_NAME, False):
                            self._draw_bezier_curve(embryo_pos, (freezer_x, freezer_y),
                                                   'cyan', linestyle='-', linewidth=1.5, alpha=0.5)
                        
                        # Draw embryo circle (only when visible, behind freezer when hidden)
                        if self.embryo_visibility.get(FREEZER_NODE_NAME, False):
                            facecolor = 'none'
                            edgecolor = 'blue'
                            
                            artist, = self.ax.plot(x, embryo_y, 'o', markersize=10,
                                                  markerfacecolor=facecolor,
                                                  markeredgecolor=edgecolor,
                                                  markeredgewidth=2,
                                                  picker=10, zorder=15, clip_on=False)
                            
                            # Register for picking
                            self.embryo_artists[artist] = (FREEZER_TRANSFER_ID, embryo_id)
    
    def _calculate_content_bounds(self):
        """Calculate tight bounds around all plotted content including dragged positions.
        
        Returns:
            Tuple of (x_min, x_max, y_min, y_max) or None if no content
        """
        # Determine which animals to show (same logic as _render_flow_track)
        animals_to_show = self._get_flow_animals_to_show()
        
        if not animals_to_show:
            return None
        
        Role = self.parent_app.Role
        
        # Categorize animals by role
        egg_donors = []
        sperm_donors = []
        surrogates = []

        selected_scope_animals = set()
        if not self.settings.get('render_full_graph', True):
            selected_scope_animals = self._get_selected_scope_animals()
        
        for animal_name in animals_to_show:
            animal_data = self.parent_app.animals.get(animal_name, {})
            role = _animal_role_value(animal_data)
            
            if role == Role.SPENDER.value:
                egg_donors.append(animal_name)
            elif role == Role.SAMENSP.value:
                sperm_donors.append(animal_name)
            elif role == Role.AMME.value:
                surrogates.append(animal_name)
        
        # Layout constants - match _render_flow_track
        EGG_RAIL_X = 0.02
        SPERM_RAIL_X = 0.98
        SURROGATE_RAIL_Y = 0.02
        
        # Collect all positions
        x_coords = []
        y_coords = []
        
        # Egg donors positions (check temp_positions first)
        n_egg = len(egg_donors)
        for i, donor in enumerate(egg_donors):
            if ('animal', donor) in self.temp_positions:
                x, y = self.temp_positions[('animal', donor)]
            else:
                x = EGG_RAIL_X
                y = 0.1 + (i / max(1, n_egg - 1)) * 0.8 if n_egg > 1 else 0.5
            x_coords.append(x)
            y_coords.append(y)
        
        # Sperm donors positions (check temp_positions first)
        n_sperm = len(sperm_donors)
        for i, donor in enumerate(sperm_donors):
            if ('animal', donor) in self.temp_positions:
                x, y = self.temp_positions[('animal', donor)]
            else:
                x = SPERM_RAIL_X
                y = 0.1 + (i / max(1, n_sperm - 1)) * 0.8 if n_sperm > 1 else 0.5
            x_coords.append(x)
            y_coords.append(y)
        
        # Surrogates positions (check temp_positions first)
        n_surr = len(surrogates)
        for i, surrogate in enumerate(surrogates):
            if ('animal', surrogate) in self.temp_positions:
                x, y = self.temp_positions[('animal', surrogate)]
            else:
                x = 0.1 + (i / max(1, n_surr - 1)) * 0.8 if n_surr > 1 else 0.5
                y = SURROGATE_RAIL_Y
            x_coords.append(x)
            y_coords.append(y)
        
        # Embryo positions (including temp_positions)
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            surrogate_name = transfer_data.get('surrogate_name')
            embryos = transfer_data.get('embryos', [])

            if selected_scope_animals:
                embryos = [
                    embryo for embryo in embryos
                    if self._is_embryo_connected_to_selection(embryo, surrogate_name, selected_scope_animals)
                ]
            
            if surrogate_name in surrogates and embryos:
                # Get surrogate position (may be dragged)
                if ('animal', surrogate_name) in self.temp_positions:
                    surr_x, surr_y = self.temp_positions[('animal', surrogate_name)]
                else:
                    surr_idx = surrogates.index(surrogate_name)
                    surr_x = 0.1 + (surr_idx / max(1, n_surr - 1)) * 0.8 if n_surr > 1 else 0.5
                    surr_y = SURROGATE_RAIL_Y
                
                # Embryos are placed above surrogate
                n_embryos = len(embryos)
                for j, embryo in enumerate(embryos):
                    embryo_id = embryo.get('id')
                    if ('embryo', transfer_id, embryo_id) in self.temp_positions:
                        embryo_x, embryo_y = self.temp_positions[('embryo', transfer_id, embryo_id)]
                    else:
                        embryo_x = surr_x + (j - n_embryos / 2) * 0.015
                        embryo_y = surr_y + 0.12
                    x_coords.append(embryo_x)
                    y_coords.append(embryo_y)
        
        if not x_coords or not y_coords:
            return None
        
        # Calculate bounds
        x_min, x_max = min(x_coords), max(x_coords)
        y_min, y_max = min(y_coords), max(y_coords)
        
        # Add generous padding to prevent clipping of names and lifebars
        # Names are positioned above nodes, lifebars below
        # Side rails have names that extend horizontally beyond node centers
        x_range = x_max - x_min
        y_range = y_max - y_min
        
        # Ensure minimum range to avoid too tight bounds
        x_range = max(x_range, 0.2)
        y_range = max(y_range, 0.2)
        
        # Horizontal padding: generous for names extending from edge rails
        padding_x = max(x_range * 0.15, 0.15)  # At least 15% or 0.15 units
        
        # Vertical padding: extra room for names above and lifebars below
        padding_top = max(y_range * 0.12, 0.1)  # More space for names above
        padding_bottom = max(y_range * 0.15, 0.12)  # More space for lifebars below
        
        x_min -= padding_x
        x_max += padding_x
        y_min -= padding_bottom
        y_max += padding_top
        
        return (x_min, x_max, y_min, y_max)
    
    def _apply_aspect_fill(self, xlim, ylim):
        """Adjust data limits to fill canvas while maintaining equal aspect ratio.
        
        Args:
            xlim: Tuple (x_min, x_max)
            ylim: Tuple (y_min, y_max)
            
        Returns:
            Tuple of (new_xlim, new_ylim) that fills canvas without gutters
        """
        # Get canvas dimensions
        canvas_width = self.canvas.get_width_height()[0]
        canvas_height = self.canvas.get_width_height()[1]
        
        if canvas_width <= 0 or canvas_height <= 0:
            return xlim, ylim
        
        # Calculate aspect ratios
        fig_ratio = canvas_width / canvas_height
        
        x_range = xlim[1] - xlim[0]
        y_range = ylim[1] - ylim[0]
        
        if x_range <= 0 or y_range <= 0:
            return xlim, ylim
        
        data_ratio = x_range / y_range
        
        # Adjust limits to match canvas aspect ratio
        x_center = (xlim[0] + xlim[1]) / 2
        y_center = (ylim[0] + ylim[1]) / 2
        
        if data_ratio > fig_ratio:
            # Data is wider - expand y range
            new_y_range = x_range / fig_ratio
            new_ylim = (y_center - new_y_range / 2, y_center + new_y_range / 2)
            new_xlim = xlim
        else:
            # Data is taller - expand x range
            new_x_range = y_range * fig_ratio
            new_xlim = (x_center - new_x_range / 2, x_center + new_x_range / 2)
            new_ylim = ylim
        
        return new_xlim, new_ylim
    
    def _draw_grid(self):
        """Draw a light grid on the canvas for positioning reference."""
        import math
        grid_spacing = 0.05  # 5% increments
        
        # Get current axis limits to draw grid across entire visible area
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        
        # Calculate grid line positions using floor for correct negative coordinate handling
        x_start = math.floor(xlim[0] / grid_spacing) * grid_spacing
        x_end = math.ceil(xlim[1] / grid_spacing) * grid_spacing
        y_start = math.floor(ylim[0] / grid_spacing) * grid_spacing
        y_end = math.ceil(ylim[1] / grid_spacing) * grid_spacing
        
        # Draw vertical grid lines
        x = x_start
        while x <= x_end:
            self.ax.axvline(x, color='lightgray', linewidth=0.5, alpha=0.3, zorder=1)
            x += grid_spacing
        
        # Draw horizontal grid lines
        y = y_start
        while y <= y_end:
            self.ax.axhline(y, color='lightgray', linewidth=0.5, alpha=0.3, zorder=1)
            y += grid_spacing
    
    def _snap_to_grid(self, x, y):
        """Snap coordinates to nearest grid point."""
        if not self.settings.get('snap_to_grid', False):
            return x, y
        
        grid_spacing = 0.05
        snapped_x = round(x / grid_spacing) * grid_spacing
        snapped_y = round(y / grid_spacing) * grid_spacing
        return snapped_x, snapped_y
    
    def _draw_bezier_curve(self, start_pos, end_pos, color, linestyle='-', linewidth=1.5, alpha=0.6):
        """Draw a smooth Bézier curve between two points."""
        # Access matplotlib via parent_app
        Path = self.parent_app.matplotlib.path.Path
        mpatches = self.parent_app.matplotlib.patches
        
        x_start, y_start = start_pos
        x_end, y_end = end_pos
        
        # Calculate control point for quadratic Bézier curve
        if abs(x_end - x_start) > abs(y_end - y_start):
            # Mostly horizontal - arc upward
            ctrl_x = (x_start + x_end) / 2
            ctrl_y = max(y_start, y_end) + 0.1
        else:
            # Mostly vertical - arc toward center
            ctrl_x = (x_start + x_end) / 2
            ctrl_y = (y_start + y_end) / 2
        
        # Create Bézier path
        verts = [(x_start, y_start), (ctrl_x, ctrl_y), (x_end, y_end)]
        codes = [Path.MOVETO, Path.CURVE3, Path.CURVE3]
        path = Path(verts, codes)
        
        # Create patch (clip_on=False allows drawing beyond axis limits)
        patch = mpatches.PathPatch(path, facecolor='none', edgecolor=color,
                                   linewidth=linewidth, linestyle=linestyle,
                                   alpha=alpha, zorder=5)
        patch.set_clip_on(False)
        self.ax.add_patch(patch)
    
    def _get_lifebar_color(self, efficiency_pct: Optional[float]) -> str:
        """Get lifebar color based on efficiency percentage."""
        if efficiency_pct is None:
            return '#CCCCCC'  # Grey for no data
        
        if efficiency_pct < 15:
            return '#FF4444'  # Red
        elif efficiency_pct < 35:
            return '#FFAA00'  # Yellow
        elif efficiency_pct < 55:
            return '#88DD88'  # Light green
        else:
            return '#00AA00'  # Dark green
    
    def _render_lifebars_for_animal(self, animal_name: str, node_x: float, node_y: float, role: int) -> int:
        """Render efficiency lifebars for an animal based on selected metric.
        
        Returns:
            Number of lifebars drawn
        """
        Role = self.parent_app.Role
        
        # Determine which metric to show based on role
        if role == Role.SPENDER.value:
            metric = self.settings.get('egg_donor_metric', 'in_vitro_embryo_to_implanted')
            aggregation = self.settings.get('egg_donor_aggregation', 'total')
            stats = self._calculate_egg_donor_efficiency(animal_name)
        elif role == Role.SAMENSP.value:
            metric = self.settings.get('sperm_donor_metric', 'in_vitro_fert_to_implanted')
            aggregation = self.settings.get('sperm_donor_aggregation', 'total')
            stats = self._calculate_sperm_donor_efficiency(animal_name)
        elif role == Role.AMME.value:
            metric = self.settings.get('surrogate_metric', 'total_implantation')
            aggregation = self.settings.get('surrogate_aggregation', 'total')
            stats = self._calculate_surrogate_efficiency(animal_name)
        else:
            return 0
        
        # Build list of lifebars to draw
        lifebars_to_draw = []
        
        if aggregation in ['per_surgery', 'per_donation', 'per_transfer']:
            # Per-event metric: one bar per event
            if role == Role.SPENDER.value:
                for surg in stats.get('per_surgery', []):
                    efficiency = self._extract_egg_efficiency(surg, metric)
                    lifebars_to_draw.append({
                        'event_date': surg['date'],
                        'efficiency_pct': efficiency
                    })
            elif role == Role.SAMENSP.value:
                for donation in stats.get('per_donation', []):
                    efficiency = self._extract_sperm_efficiency(donation, metric)
                    lifebars_to_draw.append({
                        'event_date': donation['date'],
                        'efficiency_pct': efficiency
                    })
            elif role == Role.AMME.value:
                for transfer in stats.get('per_transfer', []):
                    # Extract appropriate metric based on selection
                    if metric == 'total_implantation':
                        efficiency = transfer.get('implantation_pct_all')
                    elif metric == 'transfer_success_rate':
                        # Per-transfer success is binary: 100% if any implanted, 0% otherwise
                        efficiency = 100.0 if transfer.get('implanted_count_all', 0) > 0 else 0.0
                    else:
                        efficiency = None
                    lifebars_to_draw.append({
                        'event_date': transfer['date'],
                        'efficiency_pct': efficiency
                    })
        else:
            # Total metric: single bar
            total = stats.get('total', {})
            efficiency = self._extract_total_efficiency(total, metric, role)
            lifebars_to_draw.append({
                'event_date': None,
                'efficiency_pct': efficiency
            })
        
        # Position and draw lifebars
        has_name = self.settings.get('show_animal_names', True)
        
        # Initialize lifebar data for this animal if not exists
        if animal_name not in self.lifebar_data:
            self.lifebar_data[animal_name] = []
        
        # Clear old lifebars
        for old_bar in self.lifebar_data[animal_name]:
            if 'annotation' in old_bar:
                # Annotations need to be removed from the axes differently
                try:
                    old_bar['annotation'].set_visible(False)
                    if old_bar['annotation'] in self.ax.texts:
                        self.ax.texts.remove(old_bar['annotation'])
                except (ValueError, AttributeError):
                    pass  # Already removed or doesn't exist
            elif 'patches' in old_bar:  # Backward compatibility
                outline, fill = old_bar['patches']
                if outline in self.ax.patches:
                    outline.remove()
                if fill in self.ax.patches:
                    fill.remove()
        self.lifebar_data[animal_name] = []
        
        # Draw each lifebar using annotations for proper positioning
        LIFEBAR_SPACING_PX = 4  # pixels between bars (3px bar + 1px gap)
        
        for i, bar_data in enumerate(lifebars_to_draw):
            efficiency = bar_data['efficiency_pct']
            
            # Calculate y offset in points (negative = below node)
            y_offset = -5 if has_name else -3
            y_offset -= (i * LIFEBAR_SPACING_PX)
            
            # Determine fill percentage and color
            if efficiency is not None:
                fill_pct = efficiency / 100.0
                color = self._get_lifebar_color(efficiency)
            else:
                fill_pct = 1.0
                color = '#CCCCCC'
            
            # Create lifebar as Unicode block characters
            num_blocks = 10  # Total blocks for full bar
            filled_blocks = int(fill_pct * num_blocks)
            empty_blocks = num_blocks - filled_blocks
            
            # Use block characters: filled ▓, empty ░
            bar_text = '▓' * filled_blocks + '░' * empty_blocks
            
            # Create annotation for lifebar with box
            lifebar_ann = self.ax.annotate(
                bar_text,
                xy=(node_x, node_y),
                xytext=(0, y_offset),
                textcoords='offset points',
                ha='center',
                va='top',
                fontsize=7,
                color=color,
                fontfamily='monospace',
                weight='bold',
                bbox=dict(
                    boxstyle='square,pad=0.1',
                    facecolor='white',
                    edgecolor='black',
                    linewidth=1
                ),
                zorder=19,
                clip_on=False
            )
            
            # Store lifebar data for hover detection
            self.lifebar_data[animal_name].append({
                'event_date': bar_data['event_date'],
                'efficiency_pct': efficiency,
                'metric_name': self._get_metric_display_name(metric),
                'annotation': lifebar_ann,
                'node_pos': (node_x, node_y),
                'offset_points': (0, y_offset)
            })
        
        # Return number of lifebars drawn
        return len(lifebars_to_draw)
    
    def _extract_egg_efficiency(self, surg_data: Dict, metric: str) -> Optional[float]:
        """Extract efficiency value from egg donor surgery data (6 IVM-focused metrics)."""
        # Metrics matching _calculate_egg_donor_efficiency
        if metric == 'ivm_rate':
            return surg_data.get('ivm_rate')
        elif metric == 'fert_rate_after_ivm':
            return surg_data.get('fert_rate_after_ivm')
        elif metric == 'usable_from_fert':
            return surg_data.get('usable_from_fert')
        elif metric == 'implantation_rate':
            return surg_data.get('implantation_rate')
        elif metric == 'usable_per_gv':
            # This is a ratio, convert to percentage for display
            ratio = surg_data.get('usable_per_gv')
            return (ratio * 100) if ratio is not None else None
        elif metric == 'fert_per_gv':
            # This is a ratio, convert to percentage for display
            ratio = surg_data.get('fert_per_gv')
            return (ratio * 100) if ratio is not None else None
        return None
    
    def _extract_sperm_efficiency(self, donation_data: Dict, metric: str) -> Optional[float]:
        """Extract efficiency value from sperm donation data (4 IVM-focused metrics)."""
        # Metrics matching _calculate_sperm_donor_efficiency
        if metric == 'fert_rate_ivm':
            return donation_data.get('fert_rate_ivm')
        elif metric == 'usable_per_ivm':
            # This is a ratio, convert to percentage for display
            ratio = donation_data.get('usable_per_ivm')
            return (ratio * 100) if ratio is not None else None
        elif metric == 'implantation_rate':
            return donation_data.get('implantation_rate')
        elif metric == 'fert_rate_in_vivo':
            return donation_data.get('fert_rate_in_vivo')
        return None
    
    def _extract_total_efficiency(self, total_data: Dict, metric: str, role: int) -> Optional[float]:
        """Extract efficiency value from total data (updated metrics)."""
        Role = self.parent_app.Role
        
        if role == Role.SPENDER.value:
            # EGG DONOR - 6 IVM-focused metrics
            if metric == 'ivm_rate':
                return total_data.get('ivm_rate')
            elif metric == 'fert_rate_after_ivm':
                return total_data.get('fert_rate_after_ivm')
            elif metric == 'usable_from_fert':
                return total_data.get('usable_from_fert')
            elif metric == 'implantation_rate':
                return total_data.get('implantation_rate')
            elif metric == 'usable_per_gv':
                # This is a ratio, convert to percentage for display
                ratio = total_data.get('usable_per_gv')
                return (ratio * 100) if ratio is not None else None
            elif metric == 'fert_per_gv':
                # This is a ratio, convert to percentage for display
                ratio = total_data.get('fert_per_gv')
                return (ratio * 100) if ratio is not None else None
        elif role == Role.SAMENSP.value:
            # SPERM DONOR - 4 IVM-focused metrics
            if metric == 'fert_rate_ivm':
                return total_data.get('fert_rate_ivm')
            elif metric == 'usable_per_ivm':
                # This is a ratio, convert to percentage for display
                ratio = total_data.get('usable_per_ivm')
                return (ratio * 100) if ratio is not None else None
            elif metric == 'implantation_rate':
                return total_data.get('implantation_rate')
            elif metric == 'fert_rate_in_vivo':
                return total_data.get('fert_rate_in_vivo')
        elif role == Role.AMME.value:
            # SURROGATE - 2 metrics
            if metric == 'total_implantation':
                return total_data.get('total_implantation_pct')
            elif metric == 'transfer_success_rate':
                return total_data.get('transfer_success_rate')
        
        return None
    
    def _get_metric_display_name(self, metric: str) -> str:
        """Get display name for metric using translation lookups (matches dropdown menu text)."""
        # Use same translation keys as settings dialog dropdowns
        metric_names = {
            # EGG DONOR - 6 IVM-focused metrics
            'ivm_rate': self.messages.get('flow_track.settings.metric.ivm_rate', 'IVM rate'),
            'fert_rate_after_ivm': self.messages.get('flow_track.settings.metric.fert_rate_after_ivm', 'Fertilization rate after IVM'),
            'usable_from_fert': self.messages.get('flow_track.settings.metric.usable_from_fert', 'Usable embryo yield from fertilized'),
            'implantation_rate': self.messages.get('flow_track.settings.metric.implantation_rate', 'Implantation rate'),
            'usable_per_gv': self.messages.get('flow_track.settings.metric.usable_per_gv', 'Overall usable yield per GV'),
            'fert_per_gv': self.messages.get('flow_track.settings.metric.fert_per_gv', 'Fertilized per GV'),
            # SPERM DONOR - 4 IVM-focused metrics
            'fert_rate_ivm': self.messages.get('flow_track.settings.metric.fert_rate_ivm', 'Fertilization rate (IVM M2)'),
            'usable_per_ivm': self.messages.get('flow_track.settings.metric.usable_per_ivm', 'Usable embryo yield per IVM M2'),
            'fert_rate_in_vivo': self.messages.get('flow_track.settings.metric.fert_rate_in_vivo', 'Fertilization rate (in vivo M2)'),
            # SURROGATE - 2 metrics
            'total_implantation': self.messages.get('flow_track.settings.metric.total_implantation', 'Total Implantation rate'),
            'transfer_success_rate': self.messages.get('flow_track.settings.metric.transfer_success_rate', 'Transfer success rate')
        }
        return metric_names.get(metric, metric)
    
    def _on_pick(self, event):
        """Handle pick events (clicks and drag initiation)."""
        if not self._can("use"):
            return
        artist = event.artist
        
        # Double-click opens dialogs
        if event.mouseevent.dblclick:
            if artist in self.node_artists:
                animal_name = self.node_artists[artist]
                # Special handling for freezer node: open freezer embryo list instead of info
                if animal_name == FREEZER_NODE_NAME:
                    self._open_freezer_embryo_list()
                else:
                    self._show_animal_info(animal_name)
            elif artist in self.embryo_artists:
                transfer_id, embryo_id = self.embryo_artists[artist]
                self._edit_embryo(transfer_id, embryo_id)
        
        # Left-click: start drag mode, menu opens on release if no drag occurred
        elif event.mouseevent.button == 1:
            if artist in self.node_artists:
                animal_name = self.node_artists[artist]
                # Store click position for drag detection
                mouse_x, mouse_y = event.mouseevent.xdata, event.mouseevent.ydata
                self.click_start_pos = (mouse_x, mouse_y)
                
                # Start potential drag mode
                self.drag_active = True
                self.drag_artist = artist
                self.is_dragging = False
                
                # Store offset from mouse to node center
                node_x, node_y = artist.get_data()
                self.drag_offset = (node_x[0] - mouse_x, node_y[0] - mouse_y)
                
                # Store menu action to execute on release if no drag
                self.pending_menu_action = ('open_menu', animal_name)
                
            elif artist in self.embryo_artists:
                transfer_id, embryo_id = self.embryo_artists[artist]
                # Store click position for drag detection
                mouse_x, mouse_y = event.mouseevent.xdata, event.mouseevent.ydata
                self.click_start_pos = (mouse_x, mouse_y)
                
                # Start potential drag mode
                self.drag_active = True
                self.drag_artist = artist
                self.is_dragging = False
                
                # Store offset from mouse to embryo center
                embryo_x, embryo_y = artist.get_data()
                self.drag_offset = (embryo_x[0] - mouse_x, embryo_y[0] - mouse_y)
                
                # Store menu action to execute on release if no drag
                self.pending_menu_action = ('edit_embryo', transfer_id, embryo_id)
                
            else:
                # Other draggable elements can still be dragged
                self.drag_active = True
                self.drag_artist = artist
                self.is_dragging = False
                
                mouse_x, mouse_y = event.mouseevent.xdata, event.mouseevent.ydata
                if artist in self.embryo_artists:
                    self.click_start_pos = (mouse_x, mouse_y)
        
        # Right-click toggles embryo visibility (freezer and surrogates only)
        elif event.mouseevent.button == 3:  # Right mouse button
            # Clear any pending menu action
            self.pending_menu_action = None
            
            if artist in self.node_artists:
                animal_name = self.node_artists[artist]
                # Only allow embryo visibility toggle for freezer and surrogates
                has_surrogate_transfers = bool(self._get_transfers_for_surrogate(animal_name))
                if animal_name == FREEZER_NODE_NAME or has_surrogate_transfers:
                    # Toggle visibility for this node
                    current_state = self.embryo_visibility.get(animal_name, False)
                    self.embryo_visibility[animal_name] = not current_state
                    self._save_settings()
                    self._redraw_canvas()
    
    def _generate_suggestions(self, transfer_date, surrogate_name):
        """
        Generate donor suggestions for a transfer based on timing.
        
        Returns list of (egg_donor, sperm_donor, confidence) tuples, sorted by confidence.
        """
        tolerance_days = self.settings.get('suggestion_tolerance_days', 7)
        suggestions = []
        
        # Find egg surgeries within tolerance
        for surgery in self.egg_surgeries:
            egg_date = surgery['date']
            days_diff_egg = abs((transfer_date - egg_date).days)
            
            if days_diff_egg <= tolerance_days:
                # Find sperm donations within tolerance
                for sperm in self.sperm_donations:
                    sperm_date = sperm['date']
                    days_diff_sperm = abs((transfer_date - sperm_date).days)
                    
                    if days_diff_sperm <= tolerance_days:
                        # Calculate confidence score (0.0-1.0)
                        egg_score = max(0, 1.0 - (days_diff_egg / tolerance_days) ** 2)
                        sperm_score = max(0, 1.0 - (days_diff_sperm / tolerance_days) ** 2)
                        
                        # Weight egg timing more heavily (70% egg, 30% sperm)
                        confidence = 0.7 * egg_score + 0.3 * sperm_score
                        
                        suggestions.append((
                            surgery['animal_name'],
                            sperm['animal_name'],
                            confidence
                        ))
        
        # Sort by confidence (descending)
        suggestions.sort(key=lambda x: x[2], reverse=True)
        
        return suggestions
    
    def _format_ratio(self, ratio: Optional[float]) -> str:
        """Format ratio for display (e.g., 1.23 instead of percentage)."""
        if ratio is None:
            return 'N/A'
        return f"{ratio:.2f}"
    
    def _format_efficiency(self, pct: Optional[float]) -> str:
        """Format efficiency percentage with fallback for None."""
        if pct is None:
            return 'N/A'
        return f"{pct:.1f}%"
    
    def _calculate_surrogate_efficiency(self, animal_name: str) -> Dict[str, Any]:
        """
        Calculate implantation efficiency for surrogate (v3.0 schema).
        
        Returns dict with 'per_transfer' list and 'total' dict.
        """
        per_transfer = []
        total_embryos = 0
        total_implanted = 0
        total_embryos_in_vivo = 0
        total_embryos_in_vitro = 0
        total_implanted_in_vivo = 0
        total_implanted_in_vitro = 0
        total_transfers = 0
        successful_transfers = 0  # Transfers with at least one implantation
        
        # Iterate through all transfers (v3.0 schema)
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_id == FREEZER_TRANSFER_ID:
                continue  # Skip freezer
            
            if transfer_data.get('surrogate_name') != animal_name:
                continue
            
            # Count embryos and implanted (v3.0: all embryos, and by stage)
            embryos = transfer_data.get('embryos', [])
            embryo_count_all = len(embryos)
            embryo_count_in_vivo = sum(1 for e in embryos if e.get('stage') == STAGE_IN_VIVO_M2)
            embryo_count_in_vitro = sum(1 for e in embryos if e.get('stage') == STAGE_IN_VITRO_M2)
            
            implanted_count_all = sum(1 for e in embryos if e.get('implanted', False))
            implanted_count_in_vivo = sum(1 for e in embryos if e.get('implanted', False) and e.get('stage') == STAGE_IN_VIVO_M2)
            implanted_count_in_vitro = sum(1 for e in embryos if e.get('implanted', False) and e.get('stage') == STAGE_IN_VITRO_M2)
            
            if embryo_count_all == 0:
                continue
            
            implantation_pct_all = (implanted_count_all / embryo_count_all * 100) if embryo_count_all > 0 else None
            implantation_pct_in_vivo = (implanted_count_in_vivo / embryo_count_in_vivo * 100) if embryo_count_in_vivo > 0 else None
            implantation_pct_in_vitro = (implanted_count_in_vitro / embryo_count_in_vitro * 100) if embryo_count_in_vitro > 0 else None
            
            # Get transfer date
            transfer_date = transfer_data.get('transfer_date')
            
            per_transfer.append({
                'transfer_id': transfer_id,
                'date': transfer_date,
                'embryo_count_all': embryo_count_all,
                'embryo_count_in_vivo': embryo_count_in_vivo,
                'embryo_count_in_vitro': embryo_count_in_vitro,
                'implanted_count_all': implanted_count_all,
                'implanted_count_in_vivo': implanted_count_in_vivo,
                'implanted_count_in_vitro': implanted_count_in_vitro,
                'implantation_pct_all': implantation_pct_all,
                'implantation_pct_in_vivo': implantation_pct_in_vivo,
                'implantation_pct_in_vitro': implantation_pct_in_vitro
            })
            
            total_embryos += embryo_count_all
            total_implanted += implanted_count_all
            total_embryos_in_vivo += embryo_count_in_vivo
            total_embryos_in_vitro += embryo_count_in_vitro
            total_implanted_in_vivo += implanted_count_in_vivo
            total_implanted_in_vitro += implanted_count_in_vitro
            
            # Track transfer success (at least one implantation)
            total_transfers += 1
            if implanted_count_all > 0:
                successful_transfers += 1
        
        # Sort chronologically
        per_transfer.sort(key=lambda x: x['date'] if x['date'] else datetime.min)
        
        total_impl_pct = (total_implanted / total_embryos * 100) if total_embryos > 0 else None
        total_impl_pct_in_vivo = (total_implanted_in_vivo / total_embryos_in_vivo * 100) if total_embryos_in_vivo > 0 else None
        total_impl_pct_in_vitro = (total_implanted_in_vitro / total_embryos_in_vitro * 100) if total_embryos_in_vitro > 0 else None
        
        # Calculate transfer success rate (transfers with at least one implantation / total transfers)
        transfer_success_rate = (successful_transfers / total_transfers * 100) if total_transfers > 0 else None
        
        return {
            'per_transfer': per_transfer,
            'total': {
                'total_embryos': total_embryos,
                'total_embryos_in_vivo': total_embryos_in_vivo,
                'total_embryos_in_vitro': total_embryos_in_vitro,
                'total_implanted': total_implanted,
                'total_implanted_in_vivo': total_implanted_in_vivo,
                'total_implanted_in_vitro': total_implanted_in_vitro,
                'total_implantation_pct': total_impl_pct,
                'total_implantation_pct_in_vivo': total_impl_pct_in_vivo,
                'total_implantation_pct_in_vitro': total_impl_pct_in_vitro,
                'total_transfers': total_transfers,
                'successful_transfers': successful_transfers,
                'transfer_success_rate': transfer_success_rate
            }
        }
    
    def _calculate_sperm_donor_efficiency(self, animal_name: str) -> Dict[str, Any]:
        """
        Calculate efficiency metrics for sperm donor (4 metrics focused on IVM M2).
        
        Returns dict with 'per_donation' list and 'total' dict containing:
        1. Fertilization rate (IVM M2): Proportion of IVM M2 oocytes that fertilize
        2. Usable embryo yield per IVM M2: Transferable/freezable embryos per IVM M2 inseminated
        3. Implantation rate: Proportion of transferred that implant
        4. Fertilization rate (in vivo M2, optional): Proportion of in vivo M2 that fertilize
        """
        per_donation = []
        
        # Totals across all donations
        total_ivm_m2_inseminated = 0
        total_ivm_m2_fertilized = 0
        total_transferred_ivm = 0
        total_frozen_ivm = 0
        total_implanted_ivm = 0
        
        total_in_vivo_m2_inseminated = 0
        total_in_vivo_m2_fertilized = 0
        
        # Access sperm donor data
        if animal_name not in self.manual_data.get('sperm_donors', {}):
            return {'per_donation': [], 'total': {}}
        
        animal_manual = self.manual_data['sperm_donors'][animal_name]
        donations_data = animal_manual.get('donations', {})
        
        for donation_id, donation_data in donations_data.items():
            # Get user-entered data from spinboxes
            ivm_m2_inseminated = donation_data.get('applied_to_in_vitro_m2', 0)
            ivm_m2_fertilized = donation_data.get('fertilized_in_vitro_m2', 0)
            
            in_vivo_m2_inseminated = donation_data.get('applied_to_in_vivo_m2', 0)
            in_vivo_m2_fertilized = donation_data.get('fertilized_in_vivo_m2', 0)
            
            # Count transferred, implanted, and frozen IVM M2 embryos
            transferred_ivm = 0
            frozen_ivm = 0
            implanted_ivm = 0
            
            for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
                for embryo in transfer_data.get('embryos', []):
                    if (embryo.get('sperm_donor_name') == animal_name and
                        embryo.get('sperm_donation_id') == donation_id and
                        embryo.get('stage') == STAGE_IN_VITRO_M2):  # IVM M2 only
                        
                        if transfer_id == FREEZER_TRANSFER_ID:
                            frozen_ivm += 1
                        else:
                            transferred_ivm += 1
                            if embryo.get('implanted', False):
                                implanted_ivm += 1
            
            # Calculate 4 efficiency metrics for this donation
            # 1. Fertilization rate (IVM M2)
            fert_rate_ivm = (ivm_m2_fertilized / ivm_m2_inseminated * 100) if ivm_m2_inseminated > 0 else None
            
            # 2. Usable embryo yield per IVM M2 (ratio, not percentage)
            usable_per_ivm = (transferred_ivm + frozen_ivm) / ivm_m2_inseminated if ivm_m2_inseminated > 0 else None
            
            # 3. Implantation rate
            implantation_rate = (implanted_ivm / transferred_ivm * 100) if transferred_ivm > 0 else None
            
            # 4. Fertilization rate (in vivo M2, optional)
            fert_rate_in_vivo = (in_vivo_m2_fertilized / in_vivo_m2_inseminated * 100) if in_vivo_m2_inseminated > 0 else None
            
            # Parse date from donation_id
            try:
                donation_date = datetime.fromisoformat(donation_id)
            except (TypeError, ValueError):
                donation_date = None
            
            per_donation.append({
                'donation_id': donation_id,
                'date': donation_date,
                # IVM M2 data
                'ivm_m2_inseminated': ivm_m2_inseminated,
                'ivm_m2_fertilized': ivm_m2_fertilized,
                'transferred_ivm': transferred_ivm,
                'frozen_ivm': frozen_ivm,
                'implanted_ivm': implanted_ivm,
                # In vivo M2 data
                'in_vivo_m2_inseminated': in_vivo_m2_inseminated,
                'in_vivo_m2_fertilized': in_vivo_m2_fertilized,
                # 4 efficiency metrics
                'fert_rate_ivm': fert_rate_ivm,
                'usable_per_ivm': usable_per_ivm,
                'implantation_rate': implantation_rate,
                'fert_rate_in_vivo': fert_rate_in_vivo
            })
            
            # Accumulate totals
            total_ivm_m2_inseminated += ivm_m2_inseminated
            total_ivm_m2_fertilized += ivm_m2_fertilized
            total_transferred_ivm += transferred_ivm
            total_frozen_ivm += frozen_ivm
            total_implanted_ivm += implanted_ivm
            
            total_in_vivo_m2_inseminated += in_vivo_m2_inseminated
            total_in_vivo_m2_fertilized += in_vivo_m2_fertilized
        
        # Sort chronologically
        per_donation.sort(key=lambda x: x['date'] if x['date'] else datetime.min)
        
        # Calculate total 4 efficiency metrics
        # 1. Fertilization rate (IVM M2)
        total_fert_rate_ivm = (total_ivm_m2_fertilized / total_ivm_m2_inseminated * 100) if total_ivm_m2_inseminated > 0 else None
        
        # 2. Usable embryo yield per IVM M2
        total_usable_per_ivm = (total_transferred_ivm + total_frozen_ivm) / total_ivm_m2_inseminated if total_ivm_m2_inseminated > 0 else None
        
        # 3. Implantation rate
        total_implantation_rate = (total_implanted_ivm / total_transferred_ivm * 100) if total_transferred_ivm > 0 else None
        
        # 4. Fertilization rate (in vivo M2)
        total_fert_rate_in_vivo = (total_in_vivo_m2_fertilized / total_in_vivo_m2_inseminated * 100) if total_in_vivo_m2_inseminated > 0 else None
        
        return {
            'per_donation': per_donation,
            'total': {
                # IVM M2 totals
                'total_ivm_m2_inseminated': total_ivm_m2_inseminated,
                'total_ivm_m2_fertilized': total_ivm_m2_fertilized,
                'total_transferred_ivm': total_transferred_ivm,
                'total_frozen_ivm': total_frozen_ivm,
                'total_implanted_ivm': total_implanted_ivm,
                # In vivo M2 totals
                'total_in_vivo_m2_inseminated': total_in_vivo_m2_inseminated,
                'total_in_vivo_m2_fertilized': total_in_vivo_m2_fertilized,
                # 4 efficiency metrics
                'fert_rate_ivm': total_fert_rate_ivm,
                'usable_per_ivm': total_usable_per_ivm,
                'implantation_rate': total_implantation_rate,
                'fert_rate_in_vivo': total_fert_rate_in_vivo
            }
        }
    
    def _count_unlinked_embryos(self, sperm_donor_name: str) -> int:
        """Count embryos for sperm donor without sperm_donation_id (v3.0 schema)."""
        unlinked = 0
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            for embryo in transfer_data.get('embryos', []):
                if (embryo.get('sperm_donor_name') == sperm_donor_name and
                    embryo.get('sperm_donation_id') is None):
                    unlinked += 1
        
        return unlinked
    
    def _calculate_egg_donor_efficiency(self, animal_name: str) -> Dict[str, Any]:
        """
        Calculate efficiency metrics for egg donor (6 metrics focused on IVM pathway).
        
        Returns dict with 'per_surgery' list and 'total' dict containing:
        1. IVM rate: Proportion of retrieved GV oocytes that mature to IVM M2
        2. Fertilization rate after IVM: Proportion of IVM-derived M2 that fertilize
        3. Usable embryo yield from fertilized: Proportion of fertilized that result in transferable/freezable
        4. Implantation rate: Proportion of transferred that implant
        5. Overall usable yield per GV: Transferable/freezable embryos per GV retrieved
        6. Fertilized per GV: Fertilized IVM-derived oocytes per GV retrieved
        """
        per_surgery = []
        
        # Totals across all surgeries
        total_gv_isolated = 0
        total_ivm_m2 = 0  # m1_to_m2
        total_ivm_m2_fertilized = 0
        total_transferred_ivm = 0
        total_frozen_ivm = 0
        total_implanted_ivm = 0
        
        # Access egg donor data
        if animal_name not in self.manual_data.get('egg_donors', {}):
            return {'per_surgery': [], 'total': {}}
        
        animal_manual = self.manual_data['egg_donors'][animal_name]
        surgery_data = animal_manual.get('surgeries', {})
        
        for surgery_id, surg_data in surgery_data.items():
            # Get user-entered data from spinboxes
            gv_isolated = surg_data.get('gv_isolated', 0)
            ivm_m2 = surg_data.get('m1_to_m2', 0)  # GV→M2 (IVM M2)
            ivm_m2_fertilized = surg_data.get('in_vitro_m2_fertilized', 0)
            
            # Count transferred, implanted, and frozen from transfers_by_id (IVM-derived only)
            transferred_ivm = 0
            frozen_ivm = 0
            implanted_ivm = 0
            
            surgery_date_str = surgery_id
            for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
                for embryo in transfer_data.get('embryos', []):
                    if (embryo.get('egg_donor_name') == animal_name and
                        embryo.get('egg_donation_date') == surgery_date_str and
                        embryo.get('stage') == STAGE_IN_VITRO_M2):  # IVM-derived only
                        
                        if transfer_id == FREEZER_TRANSFER_ID:
                            frozen_ivm += 1
                        else:
                            transferred_ivm += 1
                            if embryo.get('implanted', False):
                                implanted_ivm += 1
            
            # Calculate 6 efficiency metrics for this surgery
            # 1. IVM rate
            ivm_rate = (ivm_m2 / gv_isolated * 100) if gv_isolated > 0 else None
            
            # 2. Fertilization rate after IVM
            fert_rate_after_ivm = (ivm_m2_fertilized / ivm_m2 * 100) if ivm_m2 > 0 else None
            
            # 3. Usable embryo yield from fertilized
            usable_from_fert = ((transferred_ivm + frozen_ivm) / ivm_m2_fertilized * 100) if ivm_m2_fertilized > 0 else None
            
            # 4. Implantation rate
            implantation_rate = (implanted_ivm / transferred_ivm * 100) if transferred_ivm > 0 else None
            
            # 5. Overall usable yield per GV (ratio, not percentage)
            usable_per_gv = (transferred_ivm + frozen_ivm) / gv_isolated if gv_isolated > 0 else None
            
            # 6. Fertilized per GV (ratio, not percentage)
            fert_per_gv = ivm_m2_fertilized / gv_isolated if gv_isolated > 0 else None
            
            # Parse date from surgery_id
            try:
                surgery_date = datetime.fromisoformat(surgery_id)
            except (TypeError, ValueError):
                surgery_date = None
            
            per_surgery.append({
                'surgery_id': surgery_id,
                'date': surgery_date,
                # Raw counts
                'gv_isolated': gv_isolated,
                'ivm_m2': ivm_m2,
                'ivm_m2_fertilized': ivm_m2_fertilized,
                'transferred_ivm': transferred_ivm,
                'frozen_ivm': frozen_ivm,
                'implanted_ivm': implanted_ivm,
                # 6 efficiency metrics
                'ivm_rate': ivm_rate,
                'fert_rate_after_ivm': fert_rate_after_ivm,
                'usable_from_fert': usable_from_fert,
                'implantation_rate': implantation_rate,
                'usable_per_gv': usable_per_gv,
                'fert_per_gv': fert_per_gv
            })
            
            # Accumulate totals
            total_gv_isolated += gv_isolated
            total_ivm_m2 += ivm_m2
            total_ivm_m2_fertilized += ivm_m2_fertilized
            total_transferred_ivm += transferred_ivm
            total_frozen_ivm += frozen_ivm
            total_implanted_ivm += implanted_ivm
        
        # Sort chronologically
        per_surgery.sort(key=lambda x: x['date'] if x['date'] else datetime.min)
        
        # Calculate TOTAL 6 efficiency metrics
        total_ivm_rate = (total_ivm_m2 / total_gv_isolated * 100) if total_gv_isolated > 0 else None
        total_fert_rate_after_ivm = (total_ivm_m2_fertilized / total_ivm_m2 * 100) if total_ivm_m2 > 0 else None
        total_usable_from_fert = ((total_transferred_ivm + total_frozen_ivm) / total_ivm_m2_fertilized * 100) if total_ivm_m2_fertilized > 0 else None
        total_implantation_rate = (total_implanted_ivm / total_transferred_ivm * 100) if total_transferred_ivm > 0 else None
        total_usable_per_gv = (total_transferred_ivm + total_frozen_ivm) / total_gv_isolated if total_gv_isolated > 0 else None
        total_fert_per_gv = total_ivm_m2_fertilized / total_gv_isolated if total_gv_isolated > 0 else None
        
        return {
            'per_surgery': per_surgery,
            'total': {
                # Raw counts
                'total_gv_isolated': total_gv_isolated,
                'total_ivm_m2': total_ivm_m2,
                'total_ivm_m2_fertilized': total_ivm_m2_fertilized,
                'total_transferred_ivm': total_transferred_ivm,
                'total_frozen_ivm': total_frozen_ivm,
                'total_implanted_ivm': total_implanted_ivm,
                # 6 total efficiency metrics
                'ivm_rate': total_ivm_rate,
                'fert_rate_after_ivm': total_fert_rate_after_ivm,
                'usable_from_fert': total_usable_from_fert,
                'implantation_rate': total_implantation_rate,
                'usable_per_gv': total_usable_per_gv,
                'fert_per_gv': total_fert_per_gv
            }
        }
    
    def _show_animal_info(self, animal_name):
        """Show role-specific information and data entry for clicked animal."""
        if not self._can('edit'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        Role = self.parent_app.Role
        
        animal_data = self.parent_app.animals.get(animal_name, {})
        role = _animal_role_value(animal_data) or 'Unknown'
        
        # Show role-specific dialog
        if role == Role.SPENDER.value:
            self._show_egg_donor_dialog(animal_name, animal_data)
        elif role == Role.SAMENSP.value:
            self._show_sperm_donor_dialog(animal_name, animal_data)
        elif role == Role.AMME.value:
            self._show_surrogate_dialog(animal_name, animal_data)
        else:
            # Generic info for other roles
            QtWidgets.QMessageBox.information(
                self.widget,
                self.messages.get("flow_track.animal_info.title", "Animal Information"),
                f"Animal: {animal_name}\nRole: {role}"
            )
    
    def _show_egg_donor_dialog(self, animal_name, animal_data):
        """Show egg donor statistics (Flow_Track 3.0 schema - section 5.1)."""
        QtWidgets = self.parent_app.QtWidgets
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(f"{animal_name} - {self.messages.get('flow_track.dialog.egg_donor.title', 'Egg Donor Statistics')}")
        dialog.setModal(True)
        dialog.resize(500, 500)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Title
        title = QtWidgets.QLabel(f"<b>{self._animal_export_name(animal_name)}</b> - {self.messages.get('flow_track.dialog.egg_donor.label', 'Egg Donor')}")
        layout.addWidget(title)
        layout.addSpacing(10)
        
        # === EFFICIENCY METRICS SECTION ===
        efficiency_data = self._calculate_egg_donor_efficiency(animal_name)
        
        if efficiency_data['per_surgery'] or efficiency_data['total'].get('total_gv_isolated', 0) > 0:
            metrics_group = QtWidgets.QGroupBox(self.messages.get('flow_track.dialog.efficiency_metrics', 'EFFICIENCY METRICS'))
            metrics_layout = QtWidgets.QVBoxLayout()
            
            # Per-surgery metrics
            if efficiency_data['per_surgery']:
                scroll = QtWidgets.QScrollArea()
                scroll.setWidgetResizable(True)
                scroll.setMaximumHeight(240)  # Tall enough to show one complete surgery
                
                scroll_widget = QtWidgets.QWidget()
                scroll_layout = QtWidgets.QVBoxLayout(scroll_widget)
                
                per_surgery_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.per_surgery', 'Per Surgery Efficiency:')}</b>")
                scroll_layout.addWidget(per_surgery_label)
                
                for surgery in efficiency_data['per_surgery']:
                    # Handle both datetime objects and ISO strings
                    date_obj = surgery['date']
                    if date_obj:
                        if isinstance(date_obj, str):
                            try:
                                date_str = datetime.fromisoformat(date_obj).strftime('%d.%m.%Y')
                            except (TypeError, ValueError):
                                date_str = date_obj
                        else:
                            date_str = date_obj.strftime('%d.%m.%Y')
                    else:
                        date_str = self.messages.get('flow_track.efficiency.unknown_date', 'Unknown')
                    
                    surg_text = f"{date_str}:\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.ivm_rate', 'IVM rate')}: {self._format_efficiency(surgery.get('ivm_rate'))}\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.fert_rate_after_ivm', 'Fert rate after IVM')}: {self._format_efficiency(surgery.get('fert_rate_after_ivm'))}\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.usable_from_fert', 'Usable yield from fert')}: {self._format_efficiency(surgery.get('usable_from_fert'))}\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.implantation_rate', 'Implantation rate')}: {self._format_efficiency(surgery.get('implantation_rate'))}\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.usable_per_gv', 'Usable per GV')}: {self._format_ratio(surgery.get('usable_per_gv'))}\n"
                    surg_text += f"  {self.messages.get('flow_track.dialog.egg_donor.fert_per_gv', 'Fert per GV')}: {self._format_ratio(surgery.get('fert_per_gv'))}"
                    
                    surg_label = QtWidgets.QLabel(surg_text)
                    surg_label.setStyleSheet("padding: 5px; background-color: #f0f0f0; border-radius: 3px;")
                    scroll_layout.addWidget(surg_label)
                
                scroll_layout.addStretch()
                scroll.setWidget(scroll_widget)
                metrics_layout.addWidget(scroll)
            
            # Total metrics (6 metrics)
            total = efficiency_data['total']
            
            # Get saved visibility states for metrics
            metric_visibility = self.settings.get('egg_donor_metric_visibility', {
                'ivm_rate': True,
                'fert_rate_after_ivm': True,
                'usable_from_fert': True,
                'implantation_rate': True,
                'usable_per_gv': True,
                'fert_per_gv': True
            })
            
            # Metric 1: IVM rate
            metric1_row = QtWidgets.QHBoxLayout()
            metric1_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('ivm_rate', True) else "›")
            metric1_toggle.setMaximumWidth(30)
            metric1_toggle.setFlat(True)
            metric1_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric1_row.addWidget(metric1_toggle)
            ivm_rate_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.ivm_rate', 'IVM rate')}:</b>")
            metric1_row.addWidget(ivm_rate_label)
            metric1_row.addStretch()
            metrics_layout.addLayout(metric1_row)
            
            ivm_rate_text = f"  {self.messages.get('flow_track.dialog.egg_donor.ivm_rate_desc', 'Proportion of retrieved GV oocytes that mature to IVM M2')}\n"
            ivm_rate_text += f"  {total.get('total_ivm_m2', 0)}/{total.get('total_gv_isolated', 0)} ({self._format_efficiency(total.get('ivm_rate'))})"
            ivm_rate_display = QtWidgets.QLabel(ivm_rate_text)
            ivm_rate_display.setStyleSheet("padding: 5px;")
            ivm_rate_display.setVisible(metric_visibility.get('ivm_rate', True))
            metrics_layout.addWidget(ivm_rate_display)
            
            # Metric 2: Fertilization rate after IVM
            metric2_row = QtWidgets.QHBoxLayout()
            metric2_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('fert_rate_after_ivm', True) else "›")
            metric2_toggle.setMaximumWidth(30)
            metric2_toggle.setFlat(True)
            metric2_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric2_row.addWidget(metric2_toggle)
            fert_rate_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.fert_rate_after_ivm', 'Fertilization rate after IVM')}:</b>")
            metric2_row.addWidget(fert_rate_label)
            metric2_row.addStretch()
            metrics_layout.addLayout(metric2_row)
            
            fert_rate_text = f"  {self.messages.get('flow_track.dialog.egg_donor.fert_rate_after_ivm_desc', 'Proportion of IVM-derived M2 oocytes that fertilize successfully')}\n"
            fert_rate_text += f"  {total.get('total_ivm_m2_fertilized', 0)}/{total.get('total_ivm_m2', 0)} ({self._format_efficiency(total.get('fert_rate_after_ivm'))})"
            fert_rate_display = QtWidgets.QLabel(fert_rate_text)
            fert_rate_display.setStyleSheet("padding: 5px;")
            fert_rate_display.setVisible(metric_visibility.get('fert_rate_after_ivm', True))
            metrics_layout.addWidget(fert_rate_display)
            
            # Metric 3: Usable embryo yield from fertilized
            metric3_row = QtWidgets.QHBoxLayout()
            metric3_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('usable_from_fert', True) else "›")
            metric3_toggle.setMaximumWidth(30)
            metric3_toggle.setFlat(True)
            metric3_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric3_row.addWidget(metric3_toggle)
            usable_yield_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.usable_from_fert', 'Usable embryo yield from fertilized')}:</b>")
            metric3_row.addWidget(usable_yield_label)
            metric3_row.addStretch()
            metrics_layout.addLayout(metric3_row)
            
            total_usable = total.get('total_transferred_ivm', 0) + total.get('total_frozen_ivm', 0)
            usable_yield_text = f"  {self.messages.get('flow_track.dialog.egg_donor.usable_from_fert_desc', 'Proportion of fertilized IVM-derived oocytes that result in transferable/freezable embryos')}\n"
            usable_yield_text += f"  {total_usable}/{total.get('total_ivm_m2_fertilized', 0)} ({self._format_efficiency(total.get('usable_from_fert'))})"
            usable_yield_display = QtWidgets.QLabel(usable_yield_text)
            usable_yield_display.setStyleSheet("padding: 5px;")
            usable_yield_display.setVisible(metric_visibility.get('usable_from_fert', True))
            metrics_layout.addWidget(usable_yield_display)
            
            # Metric 4: Implantation rate
            metric4_row = QtWidgets.QHBoxLayout()
            metric4_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('implantation_rate', True) else "›")
            metric4_toggle.setMaximumWidth(30)
            metric4_toggle.setFlat(True)
            metric4_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric4_row.addWidget(metric4_toggle)
            impl_rate_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.implantation_rate', 'Implantation rate')}:</b>")
            metric4_row.addWidget(impl_rate_label)
            metric4_row.addStretch()
            metrics_layout.addLayout(metric4_row)
            
            impl_rate_text = f"  {self.messages.get('flow_track.dialog.egg_donor.implantation_rate_desc', 'Proportion of transferred embryos that implant successfully')}\n"
            impl_rate_text += f"  {total.get('total_implanted_ivm', 0)}/{total.get('total_transferred_ivm', 0)} ({self._format_efficiency(total.get('implantation_rate'))})"
            impl_rate_display = QtWidgets.QLabel(impl_rate_text)
            impl_rate_display.setStyleSheet("padding: 5px;")
            impl_rate_display.setVisible(metric_visibility.get('implantation_rate', True))
            metrics_layout.addWidget(impl_rate_display)
            
            # Metric 5: Overall usable yield per GV
            metric5_row = QtWidgets.QHBoxLayout()
            metric5_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('usable_per_gv', True) else "›")
            metric5_toggle.setMaximumWidth(30)
            metric5_toggle.setFlat(True)
            metric5_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric5_row.addWidget(metric5_toggle)
            usable_per_gv_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.usable_per_gv', 'Overall usable yield per GV (main donor KPI)')}:</b>")
            metric5_row.addWidget(usable_per_gv_label)
            metric5_row.addStretch()
            metrics_layout.addLayout(metric5_row)
            
            usable_per_gv_text = f"  {self.messages.get('flow_track.dialog.egg_donor.usable_per_gv_desc', 'Number of transferable/freezable embryos per GV oocyte retrieved')}\n"
            usable_per_gv_text += f"  {total_usable}/{total.get('total_gv_isolated', 0)} ({self._format_ratio(total.get('usable_per_gv'))})"
            usable_per_gv_display = QtWidgets.QLabel(usable_per_gv_text)
            usable_per_gv_display.setStyleSheet("padding: 5px;")
            usable_per_gv_display.setVisible(metric_visibility.get('usable_per_gv', True))
            metrics_layout.addWidget(usable_per_gv_display)
            
            # Metric 6: Fertilized per GV
            metric6_row = QtWidgets.QHBoxLayout()
            metric6_toggle = QtWidgets.QPushButton("˅" if metric_visibility.get('fert_per_gv', True) else "›")
            metric6_toggle.setMaximumWidth(30)
            metric6_toggle.setFlat(True)
            metric6_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric6_row.addWidget(metric6_toggle)
            fert_per_gv_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.egg_donor.fert_per_gv', 'Fertilized per GV (early outcome KPI)')}:</b>")
            metric6_row.addWidget(fert_per_gv_label)
            metric6_row.addStretch()
            metrics_layout.addLayout(metric6_row)
            
            fert_per_gv_text = f"  {self.messages.get('flow_track.dialog.egg_donor.fert_per_gv_desc', 'Number of successfully fertilized IVM-derived oocytes per GV oocyte retrieved')}\n"
            fert_per_gv_text += f"  {total.get('total_ivm_m2_fertilized', 0)}/{total.get('total_gv_isolated', 0)} ({self._format_ratio(total.get('fert_per_gv'))})"
            fert_per_gv_display = QtWidgets.QLabel(fert_per_gv_text)
            fert_per_gv_display.setStyleSheet("padding: 5px;")
            fert_per_gv_display.setVisible(metric_visibility.get('fert_per_gv', True))
            metrics_layout.addWidget(fert_per_gv_display)
            
            # Connect toggle handlers for metrics
            def make_metric_toggle_handler(metric_key, btn, widget):
                def toggle():
                    is_visible = widget.isVisible()
                    widget.setVisible(not is_visible)
                    btn.setText("›" if is_visible else "˅")
                    # Save state
                    if 'egg_donor_metric_visibility' not in self.settings:
                        self.settings['egg_donor_metric_visibility'] = {}
                    self.settings['egg_donor_metric_visibility'][metric_key] = not is_visible
                    self._save_settings()
                return toggle
            
            metric1_toggle.clicked.connect(make_metric_toggle_handler('ivm_rate', metric1_toggle, ivm_rate_display))
            metric2_toggle.clicked.connect(make_metric_toggle_handler('fert_rate_after_ivm', metric2_toggle, fert_rate_display))
            metric3_toggle.clicked.connect(make_metric_toggle_handler('usable_from_fert', metric3_toggle, usable_yield_display))
            metric4_toggle.clicked.connect(make_metric_toggle_handler('implantation_rate', metric4_toggle, impl_rate_display))
            metric5_toggle.clicked.connect(make_metric_toggle_handler('usable_per_gv', metric5_toggle, usable_per_gv_display))
            metric6_toggle.clicked.connect(make_metric_toggle_handler('fert_per_gv', metric6_toggle, fert_per_gv_display))
            
            metrics_group.setLayout(metrics_layout)
            layout.addWidget(metrics_group)
            layout.addSpacing(10)
        
        # Get surgeries from ProgTrack
        surgeries = []
        for event in animal_data.get('events', []):
            if event.get('typ') == 'surgery':
                date_obj = event.get('datum')
                if hasattr(date_obj, 'date'):
                    surgery_id = date_obj.date().isoformat()
                elif hasattr(date_obj, 'isoformat'):
                    surgery_id = date_obj.isoformat()
                else:
                    surgery_id = str(date_obj)
                surgeries.append({
                    'date': date_obj,
                    'id': surgery_id
                })
        # Initialize v3.0 schema for this egg donor if needed
        if animal_name not in self.manual_data.get('egg_donors', {}):
            if 'egg_donors' not in self.manual_data:
                self.manual_data['egg_donors'] = {}
            self.manual_data['egg_donors'][animal_name] = {'surgeries': {}}
        
        # Surgery list with v3.0 data entry fields
        surgery_group = QtWidgets.QGroupBox(f"{self.messages.get('flow_track.dialog.egg_donor.surgeries', 'Surgeries')} ({len(surgeries)})")
        surgery_layout = QtWidgets.QVBoxLayout()
        
        surgery_widgets = []
        for surgery in surgeries:
            surg_date_key = surgery['id']
            surg_date = surgery['date']
            
            # Get existing data (enhanced schema - separate in vivo/in vitro M2)
            if surg_date_key not in self.manual_data['egg_donors'][animal_name]['surgeries']:
                self.manual_data['egg_donors'][animal_name]['surgeries'][surg_date_key] = {
                    # Left column: Tracking only - eggs isolated during surgery
                    'gv_isolated': 0,       # GV-stage eggs collected
                    'm1_isolated': 0,       # M1-stage eggs collected
                    'm2_isolated': 0,       # M2-stage eggs collected in vivo (naturally matured)
                    # Right column: Calculations - what happened to the eggs
                    'm1_to_m2': 0,          # M1 eggs matured to M2 in vitro (IVM)
                    'in_vivo_m2_fertilized': 0,     # m2_isolated eggs that were fertilized
                    'in_vitro_m2_fertilized': 0,    # m1_to_m2 eggs that were fertilized (ICSI/IVF)
                    'embryos_from_in_vivo_m2': 0,   # Embryos from in_vivo_m2_fertilized
                    'embryos_from_in_vitro_m2': 0,  # Embryos from in_vitro_m2_fertilized (IVM path)
                    'frozen_from_in_vivo_m2': 0,    # Frozen embryos from in vivo M2
                    'frozen_from_in_vitro_m2': 0,   # Frozen embryos from in vitro M2
                    'comments': ''                   # Free text comments
                }
            
            surg_data = self.manual_data['egg_donors'][animal_name]['surgeries'][surg_date_key]
            
            # Get saved surgery visibility state (default: expanded)
            if 'egg_donor_surgery_visibility' not in self.settings:
                self.settings['egg_donor_surgery_visibility'] = {}
            if animal_name not in self.settings['egg_donor_surgery_visibility']:
                self.settings['egg_donor_surgery_visibility'][animal_name] = {}
            surgery_visible = self.settings['egg_donor_surgery_visibility'][animal_name].get(surg_date_key, True)
            
            # Surgery frame with vertical layout for toggle + content
            surg_frame = QtWidgets.QFrame()
            surg_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
            surg_frame_layout = QtWidgets.QVBoxLayout()
            
            # Header row with toggle button and date
            header_row = QtWidgets.QHBoxLayout()
            toggle_btn = QtWidgets.QPushButton("˅" if surgery_visible else "›")  # Down arrow (expanded) or right (collapsed)
            toggle_btn.setMaximumWidth(30)
            toggle_btn.setFlat(True)
            toggle_btn.setStyleSheet("font-size: 16px; font-weight: bold;")
            header_row.addWidget(toggle_btn)
            
            date_label = QtWidgets.QLabel(f"<b>{surg_date.strftime('%d.%m.%Y') if hasattr(surg_date, 'strftime') else str(surg_date)}</b>")
            header_row.addWidget(date_label)
            header_row.addStretch()
            surg_frame_layout.addLayout(header_row)
            
            # Content widget (collapsible)
            content_widget = QtWidgets.QWidget()
            content_widget.setVisible(surgery_visible)  # Apply saved state
            surg_layout = QtWidgets.QHBoxLayout()  # Main horizontal layout
            
            # === LEFT COLUMN: Tracking Only ===
            left_layout = QtWidgets.QVBoxLayout()
            
            # GV isolated - inline with label
            gv_row = QtWidgets.QHBoxLayout()
            gv_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.gv_label', 'GV isolated:'))
            gv_row.addWidget(gv_label)
            gv_isolated_spin = QtWidgets.QSpinBox()
            gv_isolated_spin.setRange(0, 9999)
            gv_isolated_spin.setValue(surg_data.get('gv_isolated', 0))
            gv_isolated_spin.setMaximumWidth(70)
            gv_row.addWidget(gv_isolated_spin)
            gv_row.addStretch()
            left_layout.addLayout(gv_row)
            
            # M2 isolated - inline with label
            m2_row = QtWidgets.QHBoxLayout()
            m2_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.m2_label', 'M2 isolated:'))
            m2_row.addWidget(m2_label)
            m2_isolated_spin = QtWidgets.QSpinBox()
            m2_isolated_spin.setRange(0, 9999)
            m2_isolated_spin.setValue(surg_data.get('m2_isolated', 0))
            m2_isolated_spin.setMaximumWidth(70)
            m2_row.addWidget(m2_isolated_spin)
            m2_row.addStretch()
            left_layout.addLayout(m2_row)
            
            # Comments field
            comments_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.comments_label', 'Comments:'))
            left_layout.addWidget(comments_label)
            comments_edit = QtWidgets.QTextEdit()
            comments_edit.setText(surg_data.get('comments', ''))
            comments_edit.setPlaceholderText(self.messages.get('flow_track.dialog.comments_placeholder', 'Enter comments...'))
            comments_edit.setMaximumHeight(80)
            left_layout.addWidget(comments_edit)
            
            surg_layout.addLayout(left_layout)
            
            # Separator
            separator = QtWidgets.QFrame()
            separator.setFrameShape(QtWidgets.QFrame.Shape.VLine)
            separator.setFrameShadow(QtWidgets.QFrame.Shadow.Sunken)
            surg_layout.addWidget(separator)
            
            # === RIGHT COLUMN: Calculations ===
            right_layout = QtWidgets.QVBoxLayout()
            
            # GV→M2
            gv_to_m2_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.gv_to_m2_label', 'GV→M2:'))
            right_layout.addWidget(gv_to_m2_label)
            m1_to_m2_spin = QtWidgets.QSpinBox()
            m1_to_m2_spin.setRange(0, 999)
            m1_to_m2_spin.setValue(surg_data.get('m1_to_m2', 0))
            right_layout.addWidget(m1_to_m2_spin)
            
            # Fertilized row
            fert_row = QtWidgets.QHBoxLayout()
            
            # In vivo column
            in_vivo_fert_layout = QtWidgets.QVBoxLayout()
            in_vivo_fert_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.in_vivo_fert_label', 'in vivo M2 fert:'))
            in_vivo_fert_layout.addWidget(in_vivo_fert_label)
            in_vivo_fert_spin = QtWidgets.QSpinBox()
            in_vivo_fert_spin.setRange(0, 999)
            in_vivo_fert_spin.setValue(surg_data.get('in_vivo_m2_fertilized', 0))
            in_vivo_fert_layout.addWidget(in_vivo_fert_spin)
            fert_row.addLayout(in_vivo_fert_layout)
            
            # In vitro column
            in_vitro_fert_layout = QtWidgets.QVBoxLayout()
            in_vitro_fert_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.ivm_m2_fert_label', 'IVM M2 fert:'))
            in_vitro_fert_layout.addWidget(in_vitro_fert_label)
            in_vitro_fert_spin = QtWidgets.QSpinBox()
            in_vitro_fert_spin.setRange(0, 999)
            in_vitro_fert_spin.setValue(surg_data.get('in_vitro_m2_fertilized', 0))
            in_vitro_fert_layout.addWidget(in_vitro_fert_spin)
            fert_row.addLayout(in_vitro_fert_layout)
            
            right_layout.addLayout(fert_row)
            
            # Embryos row
            embryo_row = QtWidgets.QHBoxLayout()
            
            # In vivo embryos column
            in_vivo_embryo_layout = QtWidgets.QVBoxLayout()
            in_vivo_embryo_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.in_vivo_embryos_label', 'in vivo M2 Embryos:'))
            in_vivo_embryo_layout.addWidget(in_vivo_embryo_label)
            in_vivo_embryo_spin = QtWidgets.QSpinBox()
            in_vivo_embryo_spin.setRange(0, 999)
            in_vivo_embryo_spin.setValue(surg_data.get('embryos_from_in_vivo_m2', 0))
            in_vivo_embryo_layout.addWidget(in_vivo_embryo_spin)
            embryo_row.addLayout(in_vivo_embryo_layout)
            
            # In vitro embryos column
            in_vitro_embryo_layout = QtWidgets.QVBoxLayout()
            in_vitro_embryo_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.ivm_m2_embryos_label', 'IVM M2 Embryos:'))
            in_vitro_embryo_layout.addWidget(in_vitro_embryo_label)
            in_vitro_embryo_spin = QtWidgets.QSpinBox()
            in_vitro_embryo_spin.setRange(0, 999)
            in_vitro_embryo_spin.setValue(surg_data.get('embryos_from_in_vitro_m2', 0))
            in_vitro_embryo_layout.addWidget(in_vitro_embryo_spin)
            embryo_row.addLayout(in_vitro_embryo_layout)
            
            right_layout.addLayout(embryo_row)
            
            # Auto-derived read-only display fields - create labels in columns
            stats_row = QtWidgets.QHBoxLayout()
            
            # In vivo column
            in_vivo_stats_layout = QtWidgets.QVBoxLayout()
            in_vivo_transferred_label = QtWidgets.QLabel('')
            in_vivo_transferred_label.setStyleSheet('color: #666; font-style: italic;')
            in_vivo_stats_layout.addWidget(in_vivo_transferred_label)
            
            in_vivo_implanted_label = QtWidgets.QLabel('')
            in_vivo_implanted_label.setStyleSheet('color: #666; font-style: italic;')
            in_vivo_stats_layout.addWidget(in_vivo_implanted_label)
            
            in_vivo_frozen_label = QtWidgets.QLabel('')
            in_vivo_frozen_label.setStyleSheet('color: #666; font-style: italic;')
            in_vivo_stats_layout.addWidget(in_vivo_frozen_label)
            
            stats_row.addLayout(in_vivo_stats_layout)
            
            # In vitro column
            in_vitro_stats_layout = QtWidgets.QVBoxLayout()
            in_vitro_transferred_label = QtWidgets.QLabel('')
            in_vitro_transferred_label.setStyleSheet('color: #666; font-style: italic;')
            in_vitro_stats_layout.addWidget(in_vitro_transferred_label)
            
            in_vitro_implanted_label = QtWidgets.QLabel('')
            in_vitro_implanted_label.setStyleSheet('color: #666; font-style: italic;')
            in_vitro_stats_layout.addWidget(in_vitro_implanted_label)
            
            in_vitro_frozen_label = QtWidgets.QLabel('')
            in_vitro_frozen_label.setStyleSheet('color: #666; font-style: italic;')
            in_vitro_stats_layout.addWidget(in_vitro_frozen_label)
            
            stats_row.addLayout(in_vitro_stats_layout)
            right_layout.addLayout(stats_row)
            
            # Function to update transferred/implanted/frozen labels for this surgery
            def update_transfer_labels_for_surgery(captured_surg_key=surg_date_key, captured_animal=animal_name):
                transferred_in_vivo = 0
                transferred_in_vitro = 0
                implanted_in_vivo = 0
                implanted_in_vitro = 0
                frozen_in_vivo = 0
                frozen_in_vitro = 0
                
                logger.debug(f"Egg donor dialog: Looking for embryos with egg_donor_name={captured_animal}, egg_donation_date={captured_surg_key}")
                
                for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
                    for embryo in transfer_data.get('embryos', []):
                        egg_name = embryo.get('egg_donor_name')
                        egg_date = embryo.get('egg_donation_date')
                        
                        if (egg_name == captured_animal and egg_date == captured_surg_key):
                            stage = embryo.get('stage')
                            is_implanted = embryo.get('implanted', False)
                            
                            # Count frozen embryos separately
                            if transfer_id == FREEZER_TRANSFER_ID:
                                if stage == STAGE_IN_VIVO_M2:
                                    frozen_in_vivo += 1
                                elif stage == STAGE_IN_VITRO_M2:
                                    frozen_in_vitro += 1
                            else:
                                # Count transferred (non-frozen)
                                if stage == STAGE_IN_VIVO_M2:
                                    transferred_in_vivo += 1
                                    if is_implanted:
                                        implanted_in_vivo += 1
                                elif stage == STAGE_IN_VITRO_M2:
                                    transferred_in_vitro += 1
                                    if is_implanted:
                                        implanted_in_vitro += 1
                
                logger.debug(f"Egg donor results: transferred_in_vivo={transferred_in_vivo}, implanted_in_vivo={implanted_in_vivo}, frozen_in_vivo={frozen_in_vivo}, transferred_in_vitro={transferred_in_vitro}, implanted_in_vitro={implanted_in_vitro}, frozen_in_vitro={frozen_in_vitro}")
                
                # Display independent stats in columns (without prefixes)
                transferred_text = self.messages.get('flow_track.dialog.egg_donor.transferred', 'transferred')
                implanted_text = self.messages.get('flow_track.dialog.egg_donor.implanted', 'implanted')
                frozen_text = self.messages.get('flow_track.dialog.egg_donor.frozen', 'frozen')
                
                in_vivo_transferred_label.setText(f'{transferred_text}: {transferred_in_vivo}')
                in_vivo_implanted_label.setText(f'{implanted_text}: {implanted_in_vivo}')
                in_vivo_frozen_label.setText(f'{frozen_text}: {frozen_in_vivo}')
                
                in_vitro_transferred_label.setText(f'{transferred_text}: {transferred_in_vitro}')
                in_vitro_implanted_label.setText(f'{implanted_text}: {implanted_in_vitro}')
                in_vitro_frozen_label.setText(f'{frozen_text}: {frozen_in_vitro}')
                
                logger.debug(f"Egg donor labels updated for surgery {captured_surg_key}")
            
            # Call immediately to populate initial values
            update_transfer_labels_for_surgery()
            
            surg_layout.addLayout(right_layout)
            
            content_widget.setLayout(surg_layout)
            surg_frame_layout.addWidget(content_widget)
            
            # Toggle function with state persistence
            def make_toggle_handler(btn, widget, animal, surg_key):
                def toggle():
                    is_visible = widget.isVisible()
                    widget.setVisible(not is_visible)
                    btn.setText("›" if is_visible else "˅")  # Right arrow when collapsed, down when expanded
                    # Save state
                    if 'egg_donor_surgery_visibility' not in self.settings:
                        self.settings['egg_donor_surgery_visibility'] = {}
                    if animal not in self.settings['egg_donor_surgery_visibility']:
                        self.settings['egg_donor_surgery_visibility'][animal] = {}
                    self.settings['egg_donor_surgery_visibility'][animal][surg_key] = not is_visible
                    self._save_settings()
                return toggle
            
            toggle_btn.clicked.connect(make_toggle_handler(toggle_btn, content_widget, animal_name, surg_date_key))
            
            surg_frame.setLayout(surg_frame_layout)
            surgery_layout.addWidget(surg_frame)
            
            # Connect value changes
            def make_value_changed_handler(surgery_key, gv_spin, m2_spin, m1_m2_spin, 
                                          in_vivo_spin, in_vitro_spin, in_vivo_emb_spin, in_vitro_emb_spin,
                                          comments_widget, label_updater):
                def handler():
                    existing = self.manual_data['egg_donors'][animal_name]['surgeries'].get(surgery_key, {})
                    self.manual_data['egg_donors'][animal_name]['surgeries'][surgery_key] = {
                        'gv_isolated': gv_spin.value(),
                        'm1_isolated': 0,
                        'm2_isolated': m2_spin.value(),
                        'm1_to_m2': m1_m2_spin.value(),
                        'in_vivo_m2_fertilized': in_vivo_spin.value(),
                        'in_vitro_m2_fertilized': in_vitro_spin.value(),
                        'embryos_from_in_vivo_m2': in_vivo_emb_spin.value(),
                        'embryos_from_in_vitro_m2': in_vitro_emb_spin.value(),
                        'comments': comments_widget.toPlainText(),
                        'frozen_from_in_vivo_m2': existing.get('frozen_from_in_vivo_m2', 0),
                        'frozen_from_in_vitro_m2': existing.get('frozen_from_in_vitro_m2', 0)
                    }
                    label_updater()
                return handler
            
            value_changed = make_value_changed_handler(surg_date_key, gv_isolated_spin,
                                                      m2_isolated_spin, m1_to_m2_spin, in_vivo_fert_spin,
                                                      in_vitro_fert_spin, in_vivo_embryo_spin, in_vitro_embryo_spin,
                                                      comments_edit, update_transfer_labels_for_surgery)
            gv_isolated_spin.valueChanged.connect(value_changed)
            m2_isolated_spin.valueChanged.connect(value_changed)
            m1_to_m2_spin.valueChanged.connect(value_changed)
            in_vivo_fert_spin.valueChanged.connect(value_changed)
            in_vitro_fert_spin.valueChanged.connect(value_changed)
            in_vivo_embryo_spin.valueChanged.connect(value_changed)
            in_vitro_embryo_spin.valueChanged.connect(value_changed)
            comments_edit.textChanged.connect(value_changed)
            
            surgery_widgets.append({
                'key': surg_date_key,
                'gv_isolated': gv_isolated_spin,
                'm2_isolated': m2_isolated_spin,
                'm1_to_m2': m1_to_m2_spin,
                'in_vivo_fert': in_vivo_fert_spin,
                'in_vitro_fert': in_vitro_fert_spin,
                'in_vivo_embryos': in_vivo_embryo_spin,
                'in_vitro_embryos': in_vitro_embryo_spin,
                'comments': comments_edit
            })
        
        if not surgeries:
            surgery_layout.addWidget(QtWidgets.QLabel(self.messages.get('flow_track.dialog.egg_donor.no_surgeries', 'No surgeries recorded')))
        
        surgery_group.setLayout(surgery_layout)
        layout.addWidget(surgery_group)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        save_btn = QtWidgets.QPushButton(self.messages.get("button.save", "Save"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog and save data
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            # Confirm before saving
            msg_template = self.messages.get("flow_track.confirm_save.message", 
                                            "Save changes to {name}'s {type} data?")
            message = msg_template.format(name=animal_name, type="egg donor")
            
            reply = QtWidgets.QMessageBox.question(
                self.widget,
                self.messages.get("flow_track.confirm_save.title", "Confirm Save"),
                message,
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.Yes
            )
            
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                for widget_set in surgery_widgets:
                    surg_key = widget_set['key']
                    # Preserve frozen counts from schema (not editable in UI)
                    existing = self.manual_data['egg_donors'][animal_name]['surgeries'].get(surg_key, {})
                    self.manual_data['egg_donors'][animal_name]['surgeries'][surg_key] = {
                        'gv_isolated': widget_set['gv_isolated'].value(),
                        'm1_isolated': 0,
                        'm2_isolated': widget_set['m2_isolated'].value(),
                        'm1_to_m2': widget_set['m1_to_m2'].value(),
                        'in_vivo_m2_fertilized': widget_set['in_vivo_fert'].value(),
                        'in_vitro_m2_fertilized': widget_set['in_vitro_fert'].value(),
                        'embryos_from_in_vivo_m2': widget_set['in_vivo_embryos'].value(),
                        'embryos_from_in_vitro_m2': widget_set['in_vitro_embryos'].value(),
                        'comments': widget_set['comments'].toPlainText(),
                        'frozen_from_in_vivo_m2': existing.get('frozen_from_in_vivo_m2', 0),
                        'frozen_from_in_vitro_m2': existing.get('frozen_from_in_vitro_m2', 0)
                    }
                
                self._save_flow_track_data()
                self._redraw_canvas()
                
                logger.info(f"Updated egg donor data (v3.0) for {animal_name}")
    
    def _show_sperm_donor_dialog(self, animal_name, animal_data):
        """Show sperm donor statistics (Flow_Track 3.0 schema - section 5.2)."""
        QtWidgets = self.parent_app.QtWidgets
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(f"{animal_name} - {self.messages.get('flow_track.dialog.sperm_donor.title', 'Sperm Donor Statistics')}")
        dialog.setModal(True)
        dialog.resize(550, 500)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Title
        title = QtWidgets.QLabel(f"<b>{self._animal_export_name(animal_name)}</b> - {self.messages.get('flow_track.dialog.sperm_donor.label', 'Sperm Donor')}")
        layout.addWidget(title)
        layout.addSpacing(10)
        
        # === EFFICIENCY METRICS SECTION ===
        efficiency_data = self._calculate_sperm_donor_efficiency(animal_name)
        
        if efficiency_data['per_donation'] or efficiency_data['total'].get('total_ivm_m2_inseminated', 0) > 0:
            metrics_group = QtWidgets.QGroupBox(self.messages.get('flow_track.dialog.efficiency_metrics', 'EFFICIENCY METRICS'))
            metrics_layout = QtWidgets.QVBoxLayout()
            
            # Per-donation metrics
            if efficiency_data['per_donation']:
                scroll = QtWidgets.QScrollArea()
                scroll.setWidgetResizable(True)
                scroll.setMaximumHeight(150)
                
                scroll_widget = QtWidgets.QWidget()
                scroll_layout = QtWidgets.QVBoxLayout(scroll_widget)
                
                per_donation_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.per_donation', 'Per Donation Efficiency:')}</b>")
                scroll_layout.addWidget(per_donation_label)
                
                for donation in efficiency_data['per_donation']:
                    date_obj = donation['date']
                    if date_obj:
                        if isinstance(date_obj, str):
                            try:
                                date_str = datetime.fromisoformat(date_obj).strftime('%d.%m.%Y')
                            except (TypeError, ValueError):
                                date_str = date_obj
                        else:
                            date_str = date_obj.strftime('%d.%m.%Y')
                    else:
                        date_str = self.messages.get('flow_track.efficiency.unknown_date', 'Unknown')
                    
                    donation_text = f"{date_str}:\n"
                    donation_text += f"  {self.messages.get('flow_track.dialog.sperm_donor.ivm_m2_inseminated', 'IVM M2 inseminated')}: {donation.get('ivm_m2_inseminated', 0)}\n"
                    donation_text += f"  {self.messages.get('flow_track.dialog.sperm_donor.fert_rate_ivm', 'Fertilization rate (IVM M2)')}: {self._format_efficiency(donation.get('fert_rate_ivm'))}\n"
                    donation_text += f"  {self.messages.get('flow_track.dialog.sperm_donor.usable_per_ivm', 'Usable per IVM M2')}: {self._format_ratio(donation.get('usable_per_ivm'))}\n"
                    donation_text += f"  {self.messages.get('flow_track.dialog.sperm_donor.implantation_rate', 'Implantation rate')}: {self._format_efficiency(donation.get('implantation_rate'))}"
                    
                    donation_label = QtWidgets.QLabel(donation_text)
                    donation_label.setStyleSheet("padding: 5px; background-color: #f0f0f0; border-radius: 3px;")
                    scroll_layout.addWidget(donation_label)
                
                scroll_layout.addStretch()
                scroll.setWidget(scroll_widget)
                metrics_layout.addWidget(scroll)
            
            # Total metrics (4 metrics)
            total = efficiency_data['total']
            
            # Get saved visibility states for sperm donor metrics
            sperm_metric_visibility = self.settings.get('sperm_donor_metric_visibility', {
                'fert_rate_ivm': True,
                'usable_per_ivm': True,
                'implantation_rate': True,
                'fert_rate_in_vivo': True
            })
            
            # Metric 1: Fertilization rate (IVM M2)
            metric1_row = QtWidgets.QHBoxLayout()
            metric1_toggle = QtWidgets.QPushButton("˅" if sperm_metric_visibility.get('fert_rate_ivm', True) else "›")
            metric1_toggle.setMaximumWidth(30)
            metric1_toggle.setFlat(True)
            metric1_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric1_row.addWidget(metric1_toggle)
            fert_ivm_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.sperm_donor.fert_rate_ivm', 'Fertilization rate (IVM M2)')}:</b>")
            metric1_row.addWidget(fert_ivm_label)
            metric1_row.addStretch()
            metrics_layout.addLayout(metric1_row)
            
            fert_ivm_text = f"  {self.messages.get('flow_track.dialog.sperm_donor.fert_rate_ivm_desc', 'Proportion of IVM-derived M2 oocytes that fertilize successfully with this sperm')}\n"
            fert_ivm_text += f"  {total.get('total_ivm_m2_fertilized', 0)}/{total.get('total_ivm_m2_inseminated', 0)} ({self._format_efficiency(total.get('fert_rate_ivm'))})"
            fert_ivm_display = QtWidgets.QLabel(fert_ivm_text)
            fert_ivm_display.setStyleSheet("padding: 5px;")
            fert_ivm_display.setVisible(sperm_metric_visibility.get('fert_rate_ivm', True))
            metrics_layout.addWidget(fert_ivm_display)
            
            # Metric 2: Usable embryo yield per IVM M2
            metric2_row = QtWidgets.QHBoxLayout()
            metric2_toggle = QtWidgets.QPushButton("˅" if sperm_metric_visibility.get('usable_per_ivm', True) else "›")
            metric2_toggle.setMaximumWidth(30)
            metric2_toggle.setFlat(True)
            metric2_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric2_row.addWidget(metric2_toggle)
            usable_ivm_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.sperm_donor.usable_per_ivm', 'Usable embryo yield per IVM M2')}:</b>")
            metric2_row.addWidget(usable_ivm_label)
            metric2_row.addStretch()
            metrics_layout.addLayout(metric2_row)
            
            total_usable_ivm = total.get('total_transferred_ivm', 0) + total.get('total_frozen_ivm', 0)
            usable_ivm_text = f"  {self.messages.get('flow_track.dialog.sperm_donor.usable_per_ivm_desc', 'Number of transferable or freezable embryos obtained per IVM-derived M2 oocyte inseminated')}\n"
            usable_ivm_text += f"  {total_usable_ivm}/{total.get('total_ivm_m2_inseminated', 0)} ({self._format_ratio(total.get('usable_per_ivm'))})"
            usable_ivm_display = QtWidgets.QLabel(usable_ivm_text)
            usable_ivm_display.setStyleSheet("padding: 5px;")
            usable_ivm_display.setVisible(sperm_metric_visibility.get('usable_per_ivm', True))
            metrics_layout.addWidget(usable_ivm_display)
            
            # Metric 3: Implantation rate
            metric3_row = QtWidgets.QHBoxLayout()
            metric3_toggle = QtWidgets.QPushButton("˅" if sperm_metric_visibility.get('implantation_rate', True) else "›")
            metric3_toggle.setMaximumWidth(30)
            metric3_toggle.setFlat(True)
            metric3_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric3_row.addWidget(metric3_toggle)
            impl_rate_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.sperm_donor.implantation_rate', 'Implantation rate')}:</b>")
            metric3_row.addWidget(impl_rate_label)
            metric3_row.addStretch()
            metrics_layout.addLayout(metric3_row)
            
            impl_rate_text = f"  {self.messages.get('flow_track.dialog.sperm_donor.implantation_rate_desc', 'Proportion of transferred IVM M2 embryos that implant successfully')}\n"
            impl_rate_text += f"  {total.get('total_implanted_ivm', 0)}/{total.get('total_transferred_ivm', 0)} ({self._format_efficiency(total.get('implantation_rate'))})"
            impl_rate_display = QtWidgets.QLabel(impl_rate_text)
            impl_rate_display.setStyleSheet("padding: 5px;")
            impl_rate_display.setVisible(sperm_metric_visibility.get('implantation_rate', True))
            metrics_layout.addWidget(impl_rate_display)
            
            # Metric 4: Fertilization rate (in vivo M2, optional)
            metric4_row = QtWidgets.QHBoxLayout()
            metric4_toggle = QtWidgets.QPushButton("˅" if sperm_metric_visibility.get('fert_rate_in_vivo', True) else "›")
            metric4_toggle.setMaximumWidth(30)
            metric4_toggle.setFlat(True)
            metric4_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric4_row.addWidget(metric4_toggle)
            fert_vivo_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.sperm_donor.fert_rate_in_vivo', 'Fertilization rate (in vivo M2)')} ({self.messages.get('flow_track.optional', 'optional')}):</b>")
            metric4_row.addWidget(fert_vivo_label)
            metric4_row.addStretch()
            metrics_layout.addLayout(metric4_row)
            
            fert_vivo_text = f"  {self.messages.get('flow_track.dialog.sperm_donor.fert_rate_in_vivo_desc', 'Proportion of in vivo-matured M2 oocytes that fertilize successfully')}\n"
            fert_vivo_text += f"  {total.get('total_in_vivo_m2_fertilized', 0)}/{total.get('total_in_vivo_m2_inseminated', 0)} ({self._format_efficiency(total.get('fert_rate_in_vivo'))})"
            fert_vivo_display = QtWidgets.QLabel(fert_vivo_text)
            fert_vivo_display.setStyleSheet("padding: 5px;")
            fert_vivo_display.setVisible(sperm_metric_visibility.get('fert_rate_in_vivo', True))
            metrics_layout.addWidget(fert_vivo_display)
            
            # Connect toggle handlers for sperm donor metrics
            def make_sperm_metric_toggle_handler(metric_key, btn, widget):
                def toggle():
                    is_visible = widget.isVisible()
                    widget.setVisible(not is_visible)
                    btn.setText("›" if is_visible else "˅")
                    # Save state
                    if 'sperm_donor_metric_visibility' not in self.settings:
                        self.settings['sperm_donor_metric_visibility'] = {}
                    self.settings['sperm_donor_metric_visibility'][metric_key] = not is_visible
                    self._save_settings()
                return toggle
            
            metric1_toggle.clicked.connect(make_sperm_metric_toggle_handler('fert_rate_ivm', metric1_toggle, fert_ivm_display))
            metric2_toggle.clicked.connect(make_sperm_metric_toggle_handler('usable_per_ivm', metric2_toggle, usable_ivm_display))
            metric3_toggle.clicked.connect(make_sperm_metric_toggle_handler('implantation_rate', metric3_toggle, impl_rate_display))
            metric4_toggle.clicked.connect(make_sperm_metric_toggle_handler('fert_rate_in_vivo', metric4_toggle, fert_vivo_display))
            
            metrics_group.setLayout(metrics_layout)
            layout.addWidget(metrics_group)
            layout.addSpacing(10)
        
        # === SPERM DONATION DATA ENTRY (v3.0 schema) ===
        sperm_list = animal_data.get('sperm', [])
        
        # Initialize v3.0 schema for this sperm donor if needed
        if animal_name not in self.manual_data.get('sperm_donors', {}):
            if 'sperm_donors' not in self.manual_data:
                self.manual_data['sperm_donors'] = {}
            self.manual_data['sperm_donors'][animal_name] = {'donations': {}}
        
        sperm_group = QtWidgets.QGroupBox(f"{self.messages.get('flow_track.dialog.sperm_donor.donations', 'SPERM DONATION DATA ENTRY')} ({len(sperm_list)})")
        sperm_layout = QtWidgets.QVBoxLayout()
        
        sperm_widgets = []
        if sperm_list:
            for sperm_entry in sperm_list:
                date_obj = sperm_entry.get('datum')
                if hasattr(date_obj, 'date'):
                    donation_id = date_obj.date().isoformat()
                elif hasattr(date_obj, 'isoformat'):
                    donation_id = date_obj.isoformat()
                else:
                    donation_id = str(date_obj)
                
                # Get existing data (enhanced schema - separate in vivo/in vitro M2)
                if donation_id not in self.manual_data['sperm_donors'][animal_name]['donations']:
                    self.manual_data['sperm_donors'][animal_name]['donations'][donation_id] = {
                        'applied_to_in_vivo_m2': 0,
                        'fertilized_in_vivo_m2': 0,
                        'embryos_from_in_vivo_m2': 0,
                        'applied_to_in_vitro_m2': 0,
                        'fertilized_in_vitro_m2': 0,
                        'embryos_from_in_vitro_m2': 0
                    }
                
                donation_data = self.manual_data['sperm_donors'][animal_name]['donations'][donation_id]
                
                # Get saved donation visibility state (default: expanded)
                if 'sperm_donor_donation_visibility' not in self.settings:
                    self.settings['sperm_donor_donation_visibility'] = {}
                if animal_name not in self.settings['sperm_donor_donation_visibility']:
                    self.settings['sperm_donor_donation_visibility'][animal_name] = {}
                donation_visible = self.settings['sperm_donor_donation_visibility'][animal_name].get(donation_id, True)
                
                # Create frame for each donation with vertical layout for toggle + content
                sperm_frame = QtWidgets.QFrame()
                sperm_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
                sperm_frame_layout = QtWidgets.QVBoxLayout()
                
                # Header row with toggle button and date
                header_row = QtWidgets.QHBoxLayout()
                toggle_btn = QtWidgets.QPushButton("˅" if donation_visible else "›")
                toggle_btn.setMaximumWidth(30)
                toggle_btn.setFlat(True)
                toggle_btn.setStyleSheet("font-size: 16px; font-weight: bold;")
                header_row.addWidget(toggle_btn)
                
                date_label = QtWidgets.QLabel(f"<b>{date_obj.strftime('%d.%m.%Y') if hasattr(date_obj, 'strftime') else str(date_obj)}</b>")
                header_row.addWidget(date_label)
                header_row.addStretch()
                sperm_frame_layout.addLayout(header_row)
                
                # Content widget (collapsible)
                content_widget = QtWidgets.QWidget()
                content_widget.setVisible(donation_visible)
                sperm_inner_layout = QtWidgets.QHBoxLayout()  # Main horizontal layout
                
                # === LEFT COLUMN: In vivo M2 ===
                left_layout = QtWidgets.QVBoxLayout()
                left_header = QtWidgets.QLabel(f"<u>{self.messages.get('flow_track.stage.in_vivo_m2', 'in vivo M2')}</u>")
                left_layout.addWidget(left_header)
                
                # Inseminated - inline with label
                in_vivo_insem_row = QtWidgets.QHBoxLayout()
                in_vivo_insem_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.inseminated_label', 'Inseminated:'))
                in_vivo_insem_row.addWidget(in_vivo_insem_label)
                in_vivo_applied_spin = QtWidgets.QSpinBox()
                in_vivo_applied_spin.setRange(0, 999)
                in_vivo_applied_spin.setValue(donation_data.get('applied_to_in_vivo_m2', 0))
                in_vivo_applied_spin.setMaximumWidth(70)
                in_vivo_insem_row.addWidget(in_vivo_applied_spin)
                in_vivo_insem_row.addStretch()
                left_layout.addLayout(in_vivo_insem_row)
                
                # Fertilized - inline with label
                in_vivo_fert_row = QtWidgets.QHBoxLayout()
                in_vivo_fert_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.fertilized_label', 'Fertilized:'))
                in_vivo_fert_row.addWidget(in_vivo_fert_label)
                in_vivo_fert_spin = QtWidgets.QSpinBox()
                in_vivo_fert_spin.setRange(0, 999)
                in_vivo_fert_spin.setValue(donation_data.get('fertilized_in_vivo_m2', 0))
                in_vivo_fert_spin.setMaximumWidth(70)
                in_vivo_fert_row.addWidget(in_vivo_fert_spin)
                in_vivo_fert_row.addStretch()
                left_layout.addLayout(in_vivo_fert_row)
                
                # Embryos - inline with label
                in_vivo_embryo_row = QtWidgets.QHBoxLayout()
                in_vivo_embryo_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.embryos_label', 'Embryos:'))
                in_vivo_embryo_row.addWidget(in_vivo_embryo_label)
                in_vivo_embryo_spin = QtWidgets.QSpinBox()
                in_vivo_embryo_spin.setRange(0, 999)
                in_vivo_embryo_spin.setValue(donation_data.get('embryos_from_in_vivo_m2', 0))
                in_vivo_embryo_spin.setMaximumWidth(70)
                in_vivo_embryo_row.addWidget(in_vivo_embryo_spin)
                in_vivo_embryo_row.addStretch()
                left_layout.addLayout(in_vivo_embryo_row)
                
                # Auto-derived transferred/implanted/frozen counts for in vivo - create labels in column
                in_vivo_transferred_label = QtWidgets.QLabel('')
                in_vivo_transferred_label.setStyleSheet('color: #666; font-style: italic;')
                left_layout.addWidget(in_vivo_transferred_label)
                
                in_vivo_implanted_label = QtWidgets.QLabel('')
                in_vivo_implanted_label.setStyleSheet('color: #666; font-style: italic;')
                left_layout.addWidget(in_vivo_implanted_label)
                
                in_vivo_frozen_label = QtWidgets.QLabel('')
                in_vivo_frozen_label.setStyleSheet('color: #666; font-style: italic;')
                left_layout.addWidget(in_vivo_frozen_label)
                
                sperm_inner_layout.addLayout(left_layout)
                
                # Separator
                separator = QtWidgets.QFrame()
                separator.setFrameShape(QtWidgets.QFrame.Shape.VLine)
                separator.setFrameShadow(QtWidgets.QFrame.Shadow.Sunken)
                sperm_inner_layout.addWidget(separator)
                
                # === RIGHT COLUMN: IVM M2 ===
                right_layout = QtWidgets.QVBoxLayout()
                right_header = QtWidgets.QLabel(f"<u>{self.messages.get('flow_track.stage.in_vitro_m2', 'IVM M2')}</u>")
                right_layout.addWidget(right_header)
                
                # Inseminated - inline with label
                in_vitro_insem_row = QtWidgets.QHBoxLayout()
                in_vitro_insem_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.inseminated_label', 'Inseminated:'))
                in_vitro_insem_row.addWidget(in_vitro_insem_label)
                in_vitro_applied_spin = QtWidgets.QSpinBox()
                in_vitro_applied_spin.setRange(0, 999)
                in_vitro_applied_spin.setValue(donation_data.get('applied_to_in_vitro_m2', 0))
                in_vitro_applied_spin.setMaximumWidth(70)
                in_vitro_insem_row.addWidget(in_vitro_applied_spin)
                in_vitro_insem_row.addStretch()
                right_layout.addLayout(in_vitro_insem_row)
                
                # Fertilized - inline with label
                in_vitro_fert_row = QtWidgets.QHBoxLayout()
                in_vitro_fert_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.fertilized_label', 'Fertilized:'))
                in_vitro_fert_row.addWidget(in_vitro_fert_label)
                in_vitro_fert_spin = QtWidgets.QSpinBox()
                in_vitro_fert_spin.setRange(0, 999)
                in_vitro_fert_spin.setValue(donation_data.get('fertilized_in_vitro_m2', 0))
                in_vitro_fert_spin.setMaximumWidth(70)
                in_vitro_fert_row.addWidget(in_vitro_fert_spin)
                in_vitro_fert_row.addStretch()
                right_layout.addLayout(in_vitro_fert_row)
                
                # Embryos - inline with label
                in_vitro_embryo_row = QtWidgets.QHBoxLayout()
                in_vitro_embryo_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.embryos_label', 'Embryos:'))
                in_vitro_embryo_row.addWidget(in_vitro_embryo_label)
                in_vitro_embryo_spin = QtWidgets.QSpinBox()
                in_vitro_embryo_spin.setRange(0, 999)
                in_vitro_embryo_spin.setValue(donation_data.get('embryos_from_in_vitro_m2', 0))
                in_vitro_embryo_spin.setMaximumWidth(70)
                in_vitro_embryo_row.addWidget(in_vitro_embryo_spin)
                in_vitro_embryo_row.addStretch()
                right_layout.addLayout(in_vitro_embryo_row)
                
                # Auto-derived transferred/implanted/frozen counts for in vitro - create labels in column
                in_vitro_transferred_label = QtWidgets.QLabel('')
                in_vitro_transferred_label.setStyleSheet('color: #666; font-style: italic;')
                right_layout.addWidget(in_vitro_transferred_label)
                
                in_vitro_implanted_label = QtWidgets.QLabel('')
                in_vitro_implanted_label.setStyleSheet('color: #666; font-style: italic;')
                right_layout.addWidget(in_vitro_implanted_label)
                
                in_vitro_frozen_label = QtWidgets.QLabel('')
                in_vitro_frozen_label.setStyleSheet('color: #666; font-style: italic;')
                right_layout.addWidget(in_vitro_frozen_label)
                
                sperm_inner_layout.addLayout(right_layout)
                
                content_widget.setLayout(sperm_inner_layout)
                sperm_frame_layout.addWidget(content_widget)
                
                # Toggle function with state persistence
                def make_donation_toggle_handler(btn, widget, animal, don_id):
                    def toggle():
                        is_visible = widget.isVisible()
                        widget.setVisible(not is_visible)
                        btn.setText("›" if is_visible else "˅")
                        # Save state
                        if 'sperm_donor_donation_visibility' not in self.settings:
                            self.settings['sperm_donor_donation_visibility'] = {}
                        if animal not in self.settings['sperm_donor_donation_visibility']:
                            self.settings['sperm_donor_donation_visibility'][animal] = {}
                        self.settings['sperm_donor_donation_visibility'][animal][don_id] = not is_visible
                        self._save_settings()
                    return toggle
                
                toggle_btn.clicked.connect(make_donation_toggle_handler(toggle_btn, content_widget, animal_name, donation_id))
                
                sperm_frame.setLayout(sperm_frame_layout)
                
                # Function to update transferred/implanted/frozen labels for this donation
                def update_transfer_labels_for_donation(captured_donation_id=donation_id, captured_animal=animal_name):
                    transferred_in_vivo = 0
                    implanted_in_vivo = 0
                    frozen_in_vivo = 0
                    transferred_in_vitro = 0
                    implanted_in_vitro = 0
                    frozen_in_vitro = 0
                    
                    logger.debug(f"Sperm donor dialog: Looking for embryos with sperm_donor_name={captured_animal}, sperm_donation_id={captured_donation_id}")
                    
                    for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
                        for embryo in transfer_data.get('embryos', []):
                            sperm_name = embryo.get('sperm_donor_name')
                            sperm_id = embryo.get('sperm_donation_id')
                            
                            logger.debug(f"  Checking embryo: sperm_donor_name={sperm_name}, sperm_donation_id={sperm_id}")
                            
                            if (sperm_name == captured_animal and sperm_id == captured_donation_id):
                                stage = embryo.get('stage')
                                is_implanted = embryo.get('implanted', False)
                                
                                logger.debug(f"    MATCH! stage={stage}, implanted={is_implanted}")
                                
                                # Count frozen separately
                                if transfer_id == FREEZER_TRANSFER_ID:
                                    if stage == STAGE_IN_VIVO_M2:
                                        frozen_in_vivo += 1
                                    elif stage == STAGE_IN_VITRO_M2:
                                        frozen_in_vitro += 1
                                else:
                                    # Count transferred (non-frozen)
                                    if stage == STAGE_IN_VIVO_M2:
                                        transferred_in_vivo += 1
                                        if is_implanted:
                                            implanted_in_vivo += 1
                                    elif stage == STAGE_IN_VITRO_M2:
                                        transferred_in_vitro += 1
                                        if is_implanted:
                                            implanted_in_vitro += 1
                    
                    logger.debug(f"Sperm donor results: transferred_in_vivo={transferred_in_vivo}, implanted_in_vivo={implanted_in_vivo}, frozen_in_vivo={frozen_in_vivo}, transferred_in_vitro={transferred_in_vitro}, implanted_in_vitro={implanted_in_vitro}, frozen_in_vitro={frozen_in_vitro}")
                    
                    # Display independent stats in columns (without prefixes) - matching egg donor design
                    transferred_text = self.messages.get('flow_track.dialog.sperm_donor.transferred', 'transferred')
                    implanted_text = self.messages.get('flow_track.dialog.sperm_donor.implanted', 'implanted')
                    frozen_text = self.messages.get('flow_track.dialog.sperm_donor.frozen', 'frozen')
                    
                    in_vivo_transferred_label.setText(f'{transferred_text}: {transferred_in_vivo}')
                    in_vivo_implanted_label.setText(f'{implanted_text}: {implanted_in_vivo}')
                    in_vivo_frozen_label.setText(f'{frozen_text}: {frozen_in_vivo}')
                    
                    in_vitro_transferred_label.setText(f'{transferred_text}: {transferred_in_vitro}')
                    in_vitro_implanted_label.setText(f'{implanted_text}: {implanted_in_vitro}')
                    in_vitro_frozen_label.setText(f'{frozen_text}: {frozen_in_vitro}')
                    
                    logger.debug(f"Sperm donor labels updated for donation {captured_donation_id}")
                
                # Call immediately to populate initial values
                update_transfer_labels_for_donation()
                
                # Connect spinbox value changes to update labels
                in_vivo_applied_spin.valueChanged.connect(update_transfer_labels_for_donation)
                in_vivo_fert_spin.valueChanged.connect(update_transfer_labels_for_donation)
                in_vivo_embryo_spin.valueChanged.connect(update_transfer_labels_for_donation)
                in_vitro_applied_spin.valueChanged.connect(update_transfer_labels_for_donation)
                in_vitro_fert_spin.valueChanged.connect(update_transfer_labels_for_donation)
                in_vitro_embryo_spin.valueChanged.connect(update_transfer_labels_for_donation)
                
                sperm_inner_layout.addLayout(right_layout)
                
                sperm_frame.setLayout(sperm_inner_layout)
                
                # Characteristics below (read-only)
                motility = sperm_entry.get('motility', 'N/A')
                progressive = sperm_entry.get('progressive', 'N/A')
                count = sperm_entry.get('count', 'N/A')
                
                char_frame = QtWidgets.QFrame()
                char_layout = QtWidgets.QHBoxLayout()
                char_label = QtWidgets.QLabel(
                    f"{self.messages.get('flow_track.dialog.sperm_donor.characteristics', 'Characteristics')}: "
                    f"{self.messages.get('flow_track.dialog.sperm_donor.motility', 'Motility')} {motility}, "
                    f"{self.messages.get('flow_track.dialog.sperm_donor.progressive', 'Progressive')} {progressive}, "
                    f"{self.messages.get('flow_track.dialog.sperm_donor.count', 'Count')} {count}"
                )
                char_label.setStyleSheet("font-size: 9pt; color: gray;")
                char_layout.addWidget(char_label)
                char_frame.setLayout(char_layout)
                
                # Add both frames to main layout
                donation_container = QtWidgets.QVBoxLayout()
                donation_container.addWidget(sperm_frame)
                donation_container.addWidget(char_frame)
                
                container_widget = QtWidgets.QWidget()
                container_widget.setLayout(donation_container)
                sperm_layout.addWidget(container_widget)
                
                sperm_widgets.append({
                    'id': donation_id,
                    'in_vivo_applied': in_vivo_applied_spin,
                    'in_vivo_fert': in_vivo_fert_spin,
                    'in_vivo_embryos': in_vivo_embryo_spin,
                    'in_vitro_applied': in_vitro_applied_spin,
                    'in_vitro_fert': in_vitro_fert_spin,
                    'in_vitro_embryos': in_vitro_embryo_spin
                })
        else:
            sperm_layout.addWidget(QtWidgets.QLabel(self.messages.get('flow_track.dialog.sperm_donor.no_donations', 'No sperm donations recorded')))
        
        sperm_group.setLayout(sperm_layout)
        layout.addWidget(sperm_group)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        save_btn = QtWidgets.QPushButton(self.messages.get("button.save", "Save"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog and save data
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            msg_template = self.messages.get("flow_track.confirm_save.message", 
                                            "Save changes to {name}'s {type} data?")
            message = msg_template.format(name=animal_name, type="sperm donor")
            
            reply = QtWidgets.QMessageBox.question(
                self.widget,
                self.messages.get("flow_track.confirm_save.title", "Confirm Save"),
                message,
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.Yes
            )
            
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                for widget_set in sperm_widgets:
                    donation_id = widget_set['id']
                    self.manual_data['sperm_donors'][animal_name]['donations'][donation_id] = {
                        'applied_to_in_vivo_m2': widget_set['in_vivo_applied'].value(),
                        'fertilized_in_vivo_m2': widget_set['in_vivo_fert'].value(),
                        'embryos_from_in_vivo_m2': widget_set['in_vivo_embryos'].value(),
                        'applied_to_in_vitro_m2': widget_set['in_vitro_applied'].value(),
                        'fertilized_in_vitro_m2': widget_set['in_vitro_fert'].value(),
                        'embryos_from_in_vitro_m2': widget_set['in_vitro_embryos'].value()
                    }
                
                if self.settings.get('auto_save_enabled', True):
                    self._save_flow_track_data()
                
                # Redraw canvas to update lifebars with new efficiency metrics
                self._redraw_canvas()
                
                logger.info(f"Updated sperm donor data (v3.0) for {animal_name}")
    
    def _show_surrogate_dialog(self, animal_name, animal_data):
        """Show surrogate statistics and allow embryo count entry."""
        QtWidgets = self.parent_app.QtWidgets
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(f"{animal_name} - {self.messages.get('flow_track.dialog.surrogate.title', 'Surrogate Statistics')}")
        dialog.setModal(True)
        dialog.resize(400, 425)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Title
        title = QtWidgets.QLabel(f"<b>{self._animal_export_name(animal_name)}</b> - {self.messages.get('flow_track.dialog.surrogate.label', 'Surrogate')}")
        layout.addWidget(title)
        layout.addSpacing(10)
        
        # === EFFICIENCY METRICS SECTION ===
        efficiency_data = self._calculate_surrogate_efficiency(animal_name)
        
        if efficiency_data['per_transfer'] or efficiency_data['total'].get('total_embryos', 0) > 0:
            metrics_group = QtWidgets.QGroupBox(self.messages.get('flow_track.dialog.efficiency_metrics', 'EFFICIENCY METRICS'))
            metrics_layout = QtWidgets.QVBoxLayout()
            
            # Per-transfer metrics
            if efficiency_data['per_transfer']:
                scroll = QtWidgets.QScrollArea()
                scroll.setWidgetResizable(True)
                scroll.setMaximumHeight(150)
                
                scroll_widget = QtWidgets.QWidget()
                scroll_layout = QtWidgets.QVBoxLayout(scroll_widget)
                
                per_transfer_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.per_transfer', 'Per Transfer Efficiency:')}</b>")
                scroll_layout.addWidget(per_transfer_label)
                
                for transfer in efficiency_data['per_transfer']:
                    # Handle both datetime objects and ISO strings
                    date_obj = transfer['date']
                    if date_obj:
                        if isinstance(date_obj, str):
                            try:
                                date_str = datetime.fromisoformat(date_obj).strftime('%d.%m.%Y')
                            except (TypeError, ValueError):
                                date_str = date_obj
                        else:
                            date_str = date_obj.strftime('%d.%m.%Y')
                    else:
                        date_str = self.messages.get('flow_track.efficiency.unknown_date', 'Unknown')
                    
                    trans_text = f"{date_str}:\n"
                    transferred_label = self.messages.get('flow_track.efficiency.transferred', 'Transferred')
                    implanted_label = self.messages.get('flow_track.efficiency.implanted', 'Implanted')
                    embryos_word = self.messages.get('flow_track.efficiency.embryos', 'embryos')
                    trans_text += f"  {transferred_label}: {transfer.get('embryo_count_all', 0)} {embryos_word}\n"
                    trans_text += f"  {implanted_label}: {transfer.get('implanted_count_all', 0)} ({self._format_efficiency(transfer.get('implantation_pct_all'))})"
                    
                    trans_label = QtWidgets.QLabel(trans_text)
                    trans_label.setStyleSheet("padding: 5px; background-color: #f0f0f0; border-radius: 3px;")
                    scroll_layout.addWidget(trans_label)
                
                scroll_layout.addStretch()
                scroll.setWidget(scroll_widget)
                metrics_layout.addWidget(scroll)
            
            # Total metrics
            total = efficiency_data['total']
            
            # Get saved visibility states for surrogate metrics
            surrogate_metric_visibility = self.settings.get('surrogate_metric_visibility', {
                'total_implantation_rate': True,
                'transfer_success_rate': True
            })
            
            # Metric 1: Total Implantation rate
            metric1_row = QtWidgets.QHBoxLayout()
            metric1_toggle = QtWidgets.QPushButton("˅" if surrogate_metric_visibility.get('total_implantation_rate', True) else "›")
            metric1_toggle.setMaximumWidth(30)
            metric1_toggle.setFlat(True)
            metric1_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric1_row.addWidget(metric1_toggle)
            total_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.total_implantation_rate', 'Total Implantation rate:')}</b>")
            metric1_row.addWidget(total_label)
            metric1_row.addStretch()
            metrics_layout.addLayout(metric1_row)
            
            total_implanted = total.get('total_implanted', 0)
            total_embryos = total.get('total_embryos', 0)
            total_implantation_pct = total.get('total_implantation_pct')
            
            total_text = f"  {self.messages.get('flow_track.dialog.surrogate.proportion_transferred', 'Proportion of transferred embryos which resulted in implantation')}\n"
            total_text += f"  {total_implanted}/{total_embryos} ({self._format_efficiency(total_implantation_pct)})"
            
            total_display = QtWidgets.QLabel(total_text)
            total_display.setStyleSheet("padding: 5px;")
            total_display.setVisible(surrogate_metric_visibility.get('total_implantation_rate', True))
            metrics_layout.addWidget(total_display)
            
            # Metric 2: Transfer success rate
            metric2_row = QtWidgets.QHBoxLayout()
            metric2_toggle = QtWidgets.QPushButton("˅" if surrogate_metric_visibility.get('transfer_success_rate', True) else "›")
            metric2_toggle.setMaximumWidth(30)
            metric2_toggle.setFlat(True)
            metric2_toggle.setStyleSheet("font-size: 14px; font-weight: bold;")
            metric2_row.addWidget(metric2_toggle)
            transfer_success_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.dialog.transfer_success_rate', 'Transfer success rate:')}</b>")
            metric2_row.addWidget(transfer_success_label)
            metric2_row.addStretch()
            metrics_layout.addLayout(metric2_row)
            
            successful_transfers = total.get('successful_transfers', 0)
            total_transfers = total.get('total_transfers', 0)
            transfer_success_rate = total.get('transfer_success_rate')
            
            success_text = f"  {self.messages.get('flow_track.dialog.surrogate.proportion_transfers', 'Proportion of transfers which resulted in implantation')}\n"
            success_text += f"  {successful_transfers}/{total_transfers} ({self._format_efficiency(transfer_success_rate)})"
            
            success_display = QtWidgets.QLabel(success_text)
            success_display.setStyleSheet("padding: 5px;")
            success_display.setVisible(surrogate_metric_visibility.get('transfer_success_rate', True))
            metrics_layout.addWidget(success_display)
            
            # Connect toggle handlers for surrogate metrics
            def make_surrogate_metric_toggle_handler(metric_key, btn, widget):
                def toggle():
                    is_visible = widget.isVisible()
                    widget.setVisible(not is_visible)
                    btn.setText("›" if is_visible else "˅")
                    # Save state
                    if 'surrogate_metric_visibility' not in self.settings:
                        self.settings['surrogate_metric_visibility'] = {}
                    self.settings['surrogate_metric_visibility'][metric_key] = not is_visible
                    self._save_settings()
                return toggle
            
            metric1_toggle.clicked.connect(make_surrogate_metric_toggle_handler('total_implantation_rate', metric1_toggle, total_display))
            metric2_toggle.clicked.connect(make_surrogate_metric_toggle_handler('transfer_success_rate', metric2_toggle, success_display))
            
            metrics_group.setLayout(metrics_layout)
            layout.addWidget(metrics_group)
            layout.addSpacing(10)
        
        # Get transfers - use same ID format as _populate_events_from_progtrack
        transfers = []
        for event in animal_data.get('events', []):
            if event.get('typ') == 'embryo_transfer':
                date_str = event.get('datum').isoformat() if hasattr(event.get('datum'), 'isoformat') else str(event.get('datum'))
                transfers.append({
                    'date': event.get('datum'),
                    'id': f"transfer_{animal_name}_{date_str}"
                })
        # Get or create manual data
        if animal_name not in self.manual_data:
            self.manual_data[animal_name] = {'transfers': {}}
        
        # Transfer list with embryo count entry
        transfer_group = QtWidgets.QGroupBox(f"{self.messages.get('flow_track.dialog.surrogate.transfers', 'Embryo Transfers')} ({len(transfers)})")
        transfer_layout = QtWidgets.QVBoxLayout()
        
        transfer_widgets = []
        for transfer in transfers:
            trans_id = transfer['id']
            trans_date = transfer['date']
            
            # Get existing data
            if trans_id not in self.manual_data[animal_name].get('transfers', {}):
                if 'transfers' not in self.manual_data[animal_name]:
                    self.manual_data[animal_name]['transfers'] = {}
                self.manual_data[animal_name]['transfers'][trans_id] = {
                    'embryo_count': 1
                }
            
            trans_data = self.manual_data[animal_name]['transfers'][trans_id]
            
            # Get saved transfer visibility state (default: expanded)
            if 'surrogate_transfer_visibility' not in self.settings:
                self.settings['surrogate_transfer_visibility'] = {}
            if animal_name not in self.settings['surrogate_transfer_visibility']:
                self.settings['surrogate_transfer_visibility'][animal_name] = {}
            transfer_visible = self.settings['surrogate_transfer_visibility'][animal_name].get(trans_id, True)
            
            # Transfer frame with vertical layout for toggle + content
            trans_frame = QtWidgets.QFrame()
            trans_frame.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
            trans_frame_layout = QtWidgets.QVBoxLayout()
            
            # Header row with toggle button and date
            header_row = QtWidgets.QHBoxLayout()
            toggle_btn = QtWidgets.QPushButton("˅" if transfer_visible else "›")
            toggle_btn.setMaximumWidth(30)
            toggle_btn.setFlat(True)
            toggle_btn.setStyleSheet("font-size: 16px; font-weight: bold;")
            header_row.addWidget(toggle_btn)
            
            date_str = trans_date.strftime('%d.%m.%Y') if hasattr(trans_date, 'strftime') else str(trans_date)
            date_label = QtWidgets.QLabel(f"<b>{date_str}</b>")
            header_row.addWidget(date_label)
            header_row.addStretch()
            trans_frame_layout.addLayout(header_row)
            
            # Content widget (collapsible)
            content_widget = QtWidgets.QWidget()
            content_widget.setVisible(transfer_visible)
            trans_row = QtWidgets.QHBoxLayout()
            
            # Embryos label
            embryos_label = QtWidgets.QLabel(self.messages.get('flow_track.dialog.surrogate.embryos_prefix', 'Embryos:'))
            trans_row.addWidget(embryos_label)
            
            # Embryo count (0 = no embryos transferred)
            embryo_spin = QtWidgets.QSpinBox()
            embryo_spin.setRange(0, 20)
            embryo_spin.setValue(trans_data.get('embryo_count', 1))
            trans_row.addWidget(embryo_spin)
            trans_row.addStretch()
            
            content_widget.setLayout(trans_row)
            trans_frame_layout.addWidget(content_widget)
            
            # Toggle function with state persistence
            def make_transfer_toggle_handler(btn, widget, animal, transfer_id):
                def toggle():
                    is_visible = widget.isVisible()
                    widget.setVisible(not is_visible)
                    btn.setText("›" if is_visible else "˅")
                    # Save state
                    if 'surrogate_transfer_visibility' not in self.settings:
                        self.settings['surrogate_transfer_visibility'] = {}
                    if animal not in self.settings['surrogate_transfer_visibility']:
                        self.settings['surrogate_transfer_visibility'][animal] = {}
                    self.settings['surrogate_transfer_visibility'][animal][transfer_id] = not is_visible
                    self._save_settings()
                return toggle
            
            toggle_btn.clicked.connect(make_transfer_toggle_handler(toggle_btn, content_widget, animal_name, trans_id))
            
            trans_frame.setLayout(trans_frame_layout)
            transfer_layout.addWidget(trans_frame)
            
            transfer_widgets.append({
                'id': trans_id,
                'embryo_count': embryo_spin
            })
        
        if not transfers:
            transfer_layout.addWidget(QtWidgets.QLabel(self.messages.get('flow_track.dialog.surrogate.no_transfers', 'No transfers recorded')))
        
        transfer_group.setLayout(transfer_layout)
        layout.addWidget(transfer_group)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        save_btn = QtWidgets.QPushButton(self.messages.get("button.save", "Save"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        save_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog and save data
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            # Confirm before saving
            msg_template = self.messages.get("flow_track.confirm_save.message", 
                                            "Save changes to {name}'s {type} data?")
            message = msg_template.format(name=animal_name, type="surrogate")
            
            reply = QtWidgets.QMessageBox.question(
                self.widget,
                self.messages.get("flow_track.confirm_save.title", "Confirm Save"),
                message,
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.Yes
            )
            
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                for widget_set in transfer_widgets:
                    trans_id = widget_set['id']
                    if 'transfers' not in self.manual_data[animal_name]:
                        self.manual_data[animal_name]['transfers'] = {}
                    self.manual_data[animal_name]['transfers'][trans_id] = {
                        'embryo_count': widget_set['embryo_count'].value()
                    }
                
                if self.settings.get('auto_save_enabled', True):
                    self._save_flow_track_data()
                
                # Refresh visualization to show correct number of circles
                self._redraw_canvas()
                
                logger.info(f"Updated surrogate data for {animal_name}")
    
    def _edit_embryo(self, transfer_id, embryo_id):
        """Edit embryo via dialog (Flow_Track 3.0 schema - section 4.2)."""
        if not self._can('edit'):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        # Get transfer and embryo data from v3.0 schema
        transfer_data = self.manual_data.get('transfers_by_id', {}).get(transfer_id, {})
        embryos = transfer_data.get('embryos', [])
        embryo = next((e for e in embryos if e['embryo_id'] == embryo_id), None)
        
        if not embryo:
            return
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(self.messages.get("flow_track.edit_embryo.title", "Edit Embryo"))
        dialog.setModal(True)
        layout = QtWidgets.QVBoxLayout()
        
        # Embryo ID edit field
        embryo_id_label = QtWidgets.QLabel(self.messages.get('flow_track.edit_embryo.embryo_label', 'Embryo:'))
        layout.addWidget(embryo_id_label)
        
        embryo_id_edit = QtWidgets.QLineEdit(embryo_id)
        layout.addWidget(embryo_id_edit)
        layout.addSpacing(10)
        
        Role = self.parent_app.Role
        
        # === EGG DONOR GROUP ===
        egg_group = QtWidgets.QGroupBox(self.messages.get("flow_track.edit_embryo.egg_donor_group", "Egg Donor"))
        egg_group_layout = QtWidgets.QVBoxLayout()
        
        # Egg donor name selection
        egg_name_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.donor_name", "Name:"))
        egg_group_layout.addWidget(egg_name_label)
        
        egg_combo = QtWidgets.QComboBox()
        egg_combo.addItem(self.messages.get('flow_track.edit_embryo.none', '(None)'), None)
        
        for name, data in self.parent_app.animals.items():
            if _animal_role_value(data) == Role.SPENDER.value:
                egg_combo.addItem(name, name)
        
        # Set current selection
        current_egg = embryo.get('egg_donor_name')
        if current_egg:
            idx = egg_combo.findData(current_egg)
            if idx >= 0:
                egg_combo.setCurrentIndex(idx)
        
        egg_group_layout.addWidget(egg_combo)
        
        # Egg donation date selection
        egg_date_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.donation_date", "Donation Date:"))
        egg_group_layout.addWidget(egg_date_label)
        
        egg_date_combo = QtWidgets.QComboBox()
        egg_date_combo.addItem(
            self.messages.get('flow_track.dialog.no_donation', '-- No donation selected --'),
            None
        )
        
        # Function to update egg donation dates based on selected egg donor
        def update_egg_date_combo():
            egg_date_combo.clear()
            egg_date_combo.addItem(
                self.messages.get('flow_track.dialog.no_donation', '-- No donation selected --'),
                None
            )
            
            selected_egg = egg_combo.currentData()
            if selected_egg:
                # Get egg donor surgeries
                animal_data = self.parent_app.animals.get(selected_egg, {})
                surgery_dates = []
                
                # Collect all surgery dates
                for event in animal_data.get('events', []):
                    if event.get('typ') == 'surgery':
                        surgery_dates.append(event.get('datum'))
                
                # Sort by date, most recent first
                surgery_dates = sorted(set(surgery_dates), reverse=True)
                
                for date_obj in surgery_dates:
                    if date_obj:
                        date_str = date_obj.strftime('%d.%m.%Y') if hasattr(date_obj, 'strftime') else str(date_obj)
                        if hasattr(date_obj, 'date'):
                            donation_id = date_obj.date().isoformat()
                        elif hasattr(date_obj, 'isoformat'):
                            donation_id = date_obj.isoformat()
                        else:
                            donation_id = str(date_obj)
                        egg_date_combo.addItem(date_str, donation_id)
        
        # Connect egg donor change to update date list
        egg_combo.currentIndexChanged.connect(update_egg_date_combo)
        
        # Initial population
        update_egg_date_combo()
        
        # Set current egg donation date if exists
        current_egg_date = embryo.get('egg_donation_date')
        if current_egg_date:
            idx = egg_date_combo.findData(current_egg_date)
            if idx >= 0:
                egg_date_combo.setCurrentIndex(idx)
        
        egg_group_layout.addWidget(egg_date_combo)
        egg_group.setLayout(egg_group_layout)
        layout.addWidget(egg_group)
        layout.addSpacing(10)
        
        # === SPERM DONOR GROUP ===
        sperm_group = QtWidgets.QGroupBox(self.messages.get("flow_track.edit_embryo.sperm_donor_group", "Sperm Donor"))
        sperm_group_layout = QtWidgets.QVBoxLayout()
        
        # Sperm donor name selection
        sperm_name_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.donor_name", "Name:"))
        sperm_group_layout.addWidget(sperm_name_label)
        
        sperm_combo = QtWidgets.QComboBox()
        sperm_combo.addItem(self.messages.get('flow_track.edit_embryo.none', '(None)'), None)
        
        for name, data in self.parent_app.animals.items():
            if _animal_role_value(data) == Role.SAMENSP.value:
                sperm_combo.addItem(name, name)
        
        # Set current selection
        current_sperm = embryo.get('sperm_donor_name')
        if current_sperm:
            idx = sperm_combo.findData(current_sperm)
            if idx >= 0:
                sperm_combo.setCurrentIndex(idx)
        
        sperm_group_layout.addWidget(sperm_combo)
        
        # Sperm donation date selection
        sperm_date_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.donation_date", "Donation Date:"))
        sperm_group_layout.addWidget(sperm_date_label)
        
        donation_combo = QtWidgets.QComboBox()
        donation_combo.addItem(
            self.messages.get('flow_track.dialog.no_donation', '-- No donation selected --'),
            None
        )
        
        # Function to update sperm donation dates based on selected sperm donor
        def update_donation_combo():
            donation_combo.clear()
            donation_combo.addItem(
                self.messages.get('flow_track.dialog.no_donation', '-- No donation selected --'),
                None
            )
            
            selected_sperm = sperm_combo.currentData()
            if selected_sperm:
                # Get sperm donations for selected donor
                animal_data = self.parent_app.animals.get(selected_sperm, {})
                sperm_list = animal_data.get('sperm', [])
                
                # Sort by date, most recent first
                sperm_list_sorted = sorted(sperm_list, 
                                          key=lambda x: x.get('datum') if x.get('datum') else datetime.min,
                                          reverse=True)
                
                for sperm_entry in sperm_list_sorted:
                    date_obj = sperm_entry.get('datum')
                    if date_obj:
                        date_str = date_obj.strftime('%d.%m.%Y') if hasattr(date_obj, 'strftime') else str(date_obj)
                        if hasattr(date_obj, 'date'):
                            donation_id = date_obj.date().isoformat()
                        elif hasattr(date_obj, 'isoformat'):
                            donation_id = date_obj.isoformat()
                        else:
                            donation_id = str(date_obj)
                        donation_combo.addItem(date_str, donation_id)
        
        # Connect sperm donor change to update donation list
        sperm_combo.currentIndexChanged.connect(update_donation_combo)
        
        # Initial population
        update_donation_combo()
        
        # Set current selection if exists
        current_donation = embryo.get('sperm_donation_id')
        if current_donation:
            idx = donation_combo.findData(current_donation)
            if idx >= 0:
                donation_combo.setCurrentIndex(idx)
        
        sperm_group_layout.addWidget(donation_combo)
        sperm_group.setLayout(sperm_group_layout)
        layout.addWidget(sperm_group)
        layout.addSpacing(10)
        
        # === STAGE SELECTION (v3.0: in_vivo_m2 / in_vitro_m2 only) ===
        stage_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.origin", "Embryo origin:"))
        layout.addWidget(stage_label)
        
        stage_combo = QtWidgets.QComboBox()
        stage_combo.addItem(self.messages.get('flow_track.stage.in_vivo_m2', 'in vivo M2'), STAGE_IN_VIVO_M2)
        stage_combo.addItem(self.messages.get('flow_track.stage.in_vitro_m2', 'IVM M2'), STAGE_IN_VITRO_M2)
        
        # Set current stage (default to in vitro M2)
        current_stage = embryo.get('stage', STAGE_IN_VITRO_M2)
        if current_stage == STAGE_IN_VITRO_M2:
            stage_combo.setCurrentIndex(1)
        else:
            stage_combo.setCurrentIndex(0)
        
        layout.addWidget(stage_combo)
        layout.addSpacing(10)
        
        # === IMPLANTED/CRYOPRESERVED CHECKBOXES ===
        implanted_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.edit_embryo.implanted", "Implanted")
        )
        implanted_check.setChecked(embryo.get('implanted', embryo.get('pregnant', False)))
        
        # Check if this is a freezer embryo and block implanted checkbox
        if transfer_id == FREEZER_TRANSFER_ID:
            implanted_check.setEnabled(False)
            implanted_check.setChecked(False)
            implanted_check.setToolTip(self.messages.get("flow_track.edit_embryo.implanted_blocked_freezer", 
                                                   "Implanted status cannot be set for embryos in the freezer"))
        
        layout.addWidget(implanted_check)
        
        cryo_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.edit_embryo.cryopreserved", "Cryopreserved")
        )
        cryo_check.setChecked(embryo.get('cryopreserved', False))
        layout.addWidget(cryo_check)
        layout.addSpacing(10)
        
        # === FREEZE DATE (v3.0: required if cryopreserved=True, hidden otherwise) ===
        freeze_date_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.freeze_date", "Freeze Date (required for frozen embryos):"))
        
        freeze_date_edit = QtWidgets.QDateEdit()
        freeze_date_edit.setCalendarPopup(True)
        freeze_date_edit.setSpecialValueText(self.messages.get('flow_track.edit_embryo.not_frozen', '(Not frozen)'))
        
        current_freeze_date = embryo.get('freeze_date')
        if current_freeze_date:
            try:
                if isinstance(current_freeze_date, str):
                    freeze_date_obj = datetime.fromisoformat(current_freeze_date)
                else:
                    freeze_date_obj = current_freeze_date
                qt_date = self.parent_app.QtCore.QDate(freeze_date_obj.year, freeze_date_obj.month, freeze_date_obj.day)
                freeze_date_edit.setDate(qt_date)
            except (AttributeError, TypeError, ValueError):
                freeze_date_edit.setDate(self.parent_app.QtCore.QDate.currentDate())
        else:
            freeze_date_edit.setDate(self.parent_app.QtCore.QDate.currentDate())
            freeze_date_edit.clear()  # Show special value text
        
        layout.addWidget(freeze_date_label)
        layout.addWidget(freeze_date_edit)
        
        # Dynamic visibility based on cryopreserved checkbox
        def update_freeze_date_visibility():
            is_frozen = cryo_check.isChecked()
            freeze_date_label.setVisible(is_frozen)
            freeze_date_edit.setVisible(is_frozen)
        
        # Connect checkbox to update visibility
        cryo_check.stateChanged.connect(update_freeze_date_visibility)
        
        # Set initial visibility
        update_freeze_date_visibility()
        layout.addSpacing(10)
        
        # === DELETE BUTTON ===
        delete_btn = QtWidgets.QPushButton(self.messages.get("button.delete", "Delete Embryo"))
        
        # Delete button handler
        def on_delete():
            reply = QtWidgets.QMessageBox.question(
                dialog,
                self.messages.get("flow_track.confirm_delete.title", "Confirm Delete"),
                self.messages.get("flow_track.confirm_delete.embryo_message", 
                                 f"Are you sure you want to delete embryo '{embryo_id}'? This will also update all related donor and surrogate counts."),
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.No
            )
            
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                self._delete_embryo(transfer_id, embryo_id)
                dialog.reject()  # Close dialog after deletion
        
        delete_btn.clicked.connect(on_delete)
        layout.addWidget(delete_btn)
        layout.addSpacing(10)
        
        # === COMMENT FIELD (v3.0: free text) ===
        comment_label = QtWidgets.QLabel(self.messages.get("flow_track.edit_embryo.comment", "Comment:"))
        layout.addWidget(comment_label)
        
        comment_edit = QtWidgets.QTextEdit()
        comment_edit.setMaximumHeight(80)
        comment_edit.setPlainText(embryo.get('comment', ''))
        layout.addWidget(comment_edit)
        
        layout.addSpacing(20)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        ok_btn = QtWidgets.QPushButton(self.messages.get("button.ok", "OK"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            # Check if donor data is being changed
            new_egg = egg_combo.currentData()
            new_sperm = sperm_combo.currentData()
            
            # Check if donors were previously set and are being changed
            donors_changed = False
            if current_egg is not None and new_egg != current_egg:
                donors_changed = True
            if current_sperm is not None and new_sperm != current_sperm:
                donors_changed = True
            
            # Confirm if changing existing donor assignments
            if donors_changed:
                reply = QtWidgets.QMessageBox.question(
                    self.widget,
                    self.messages.get("flow_track.confirm_change.title", "Confirm Change"),
                    self.messages.get("flow_track.confirm_change.message", 
                                     "Do you want to change this data?"),
                    QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                    QtWidgets.QMessageBox.StandardButton.No
                )
                
                if reply != QtWidgets.QMessageBox.StandardButton.Yes:
                    logger.info(f"User cancelled donor change for embryo {embryo_id}")
                    return
            
            # Get new embryo ID from edit field
            new_embryo_id = embryo_id_edit.text().strip()
            
            # Validate embryo ID
            if not new_embryo_id:
                QtWidgets.QMessageBox.warning(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.embryo_id_empty", "Embryo ID cannot be empty")
                )
                return
            
            # Check uniqueness if ID changed
            if new_embryo_id != embryo_id:
                if not self._is_embryo_id_unique(new_embryo_id, exclude_transfer_id=transfer_id, exclude_embryo_id=embryo_id):
                    QtWidgets.QMessageBox.warning(
                        self.widget,
                        self.messages.get("error.title", "Error"),
                        self.messages.get("flow_track.error.embryo_id_duplicate", "Embryo ID must be unique")
                    )
                    return
            
            # Get freeze date if set
            freeze_date_value = None
            if freeze_date_edit.date().isValid() and not freeze_date_edit.text() == freeze_date_edit.specialValueText():
                qt_date = freeze_date_edit.date()
                freeze_date_value = f"{qt_date.year():04d}-{qt_date.month():02d}-{qt_date.day():02d}"
            
            # Validate: cryopreserved=True requires freeze_date (v3.0 hard requirement)
            is_cryopreserved = cryo_check.isChecked()
            if is_cryopreserved and not freeze_date_value:
                QtWidgets.QMessageBox.warning(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.freeze_date_required", "Freeze date is required for frozen embryos")
                )
                return
            
            # Validate: freezer embryos cannot be implanted
            if transfer_id == FREEZER_TRANSFER_ID and implanted_check.isChecked():
                QtWidgets.QMessageBox.warning(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.error.freezer_embryo_implanted", 
                                     "Embryos in the freezer cannot be marked as implanted")
                )
                return
            
            # Update embryo data (v3.0 schema)
            embryo['embryo_id'] = new_embryo_id
            embryo['egg_donor_name'] = new_egg
            embryo['egg_donation_date'] = egg_date_combo.currentData()
            embryo['sperm_donor_name'] = new_sperm
            embryo['sperm_donation_id'] = donation_combo.currentData()
            embryo['stage'] = stage_combo.currentData()  # v3.0: in_vivo_m2 or in_vitro_m2
            embryo['implanted'] = implanted_check.isChecked()
            embryo['cryopreserved'] = is_cryopreserved
            embryo['freeze_date'] = freeze_date_value  # v3.0: required if cryopreserved
            embryo['comment'] = comment_edit.toPlainText().strip()  # v3.0: comment field
            
            # Remove legacy fields if they exist
            if 'pregnant' in embryo:
                del embryo['pregnant']
            if 'maturation_stage' in embryo:
                del embryo['maturation_stage']
            
            self._save_flow_track_data()
            self._redraw_canvas()
            logger.info(f"Updated embryo (v3.0) {embryo_id} -> {new_embryo_id}")
    
    def _delete_embryo(self, transfer_id, embryo_id):
        """Delete embryo and update all related donor/surrogate counts."""
        # Get transfer and embryo data
        transfer_data = self.manual_data.get('transfers_by_id', {}).get(transfer_id, {})
        embryos = transfer_data.get('embryos', [])
        embryo = next((e for e in embryos if e['embryo_id'] == embryo_id), None)
        
        if not embryo:
            logger.warning(f"Embryo {embryo_id} not found in transfer {transfer_id}")
            return
        
        # Get embryo metadata before deletion
        egg_donor_name = embryo.get('egg_donor_name')
        egg_donation_date = embryo.get('egg_donation_date')
        sperm_donor_name = embryo.get('sperm_donor_name')
        sperm_donation_id = embryo.get('sperm_donation_id')
        stage = embryo.get('stage')
        
        # Remove embryo from transfer
        embryos.remove(embryo)
        
        # Update egg donor counts if applicable
        if egg_donor_name and egg_donation_date:
            egg_donor_data = self.manual_data.get('egg_donors', {}).get(egg_donor_name, {})
            surgery_data = egg_donor_data.get('surgeries', {}).get(egg_donation_date, {})
            
            if surgery_data:
                if stage == STAGE_IN_VIVO_M2:
                    current = surgery_data.get('embryos_from_in_vivo_m2', 0)
                    surgery_data['embryos_from_in_vivo_m2'] = max(0, current - 1)
                elif stage == STAGE_IN_VITRO_M2:
                    current = surgery_data.get('embryos_from_in_vitro_m2', 0)
                    surgery_data['embryos_from_in_vitro_m2'] = max(0, current - 1)
                logger.info(f"Updated egg donor {egg_donor_name} counts after embryo deletion")
        
        # Update sperm donor counts if applicable
        if sperm_donor_name and sperm_donation_id:
            sperm_donor_data = self.manual_data.get('sperm_donors', {}).get(sperm_donor_name, {})
            donation_data = sperm_donor_data.get('donations', {}).get(sperm_donation_id, {})
            
            if donation_data:
                if stage == STAGE_IN_VIVO_M2:
                    current = donation_data.get('embryos_from_in_vivo_m2', 0)
                    donation_data['embryos_from_in_vivo_m2'] = max(0, current - 1)
                elif stage == STAGE_IN_VITRO_M2:
                    current = donation_data.get('embryos_from_in_vitro_m2', 0)
                    donation_data['embryos_from_in_vitro_m2'] = max(0, current - 1)
                logger.info(f"Updated sperm donor {sperm_donor_name} counts after embryo deletion")
        
        # Save and refresh
        self._save_flow_track_data()
        self._redraw_canvas()
        
        logger.info(f"Deleted embryo {embryo_id} from transfer {transfer_id}")
        
        # Show confirmation message
        QtWidgets = self.parent_app.QtWidgets
        QtWidgets.QMessageBox.information(
            self.widget,
            self.messages.get("flow_track.delete_success.title", "Embryo Deleted"),
            self.messages.get("flow_track.delete_success.message", 
                             f"Embryo '{embryo_id}' has been deleted and all related counts have been updated.")
        )
    
    def _on_resize(self, event):
        """Handle canvas resize to maintain aspect ratio without gutters."""
        if self.current_xlim is not None and self.current_ylim is not None:
            # Re-apply aspect fill to current view
            self.current_xlim, self.current_ylim = self._apply_aspect_fill(
                self.current_xlim, self.current_ylim
            )
            self.ax.set_xlim(self.current_xlim)
            self.ax.set_ylim(self.current_ylim)
            self.canvas.draw_idle()
    
    def _on_scroll(self, event):
        """Handle mouse wheel zoom."""
        if event.inaxes != self.ax:
            return
        
        # Get current axis limits
        cur_xlim = self.ax.get_xlim()
        cur_ylim = self.ax.get_ylim()
        
        # Get mouse position in data coordinates
        xdata, ydata = event.xdata, event.ydata
        
        # Zoom factor
        zoom_factor = 1.2 if event.button == 'down' else 0.8
        
        # Calculate new limits centered on mouse position
        new_xlim = (xdata - (xdata - cur_xlim[0]) * zoom_factor,
                    xdata + (cur_xlim[1] - xdata) * zoom_factor)
        new_ylim = (ydata - (ydata - cur_ylim[0]) * zoom_factor,
                    ydata + (cur_ylim[1] - ydata) * zoom_factor)
        
        # Apply aspect fill to prevent gutters
        new_xlim, new_ylim = self._apply_aspect_fill(new_xlim, new_ylim)
        
        self.ax.set_xlim(new_xlim)
        self.ax.set_ylim(new_ylim)
        
        # Store state
        self.current_xlim = new_xlim
        self.current_ylim = new_ylim
        
        # Use draw_idle for better performance
        self.canvas.draw_idle()
    
    def _on_mouse_press(self, event):
        """Handle mouse button press for panning and dragging."""
        # Middle button (2) for panning
        if event.button == 2 and event.inaxes == self.ax:
            self.pan_active = True
            self.pan_start = (event.xdata, event.ydata)
    
    def _on_mouse_release(self, event):
        """Handle mouse button release."""
        if event.button == 2:
            self.pan_active = False
            self.pan_start = None
        elif event.button == 1:
            # Check if this was a click or a drag
            if self.drag_active and self.drag_artist is not None:
                if self.is_dragging:
                    # Was a drag - apply snap-to-grid if enabled
                    if self.drag_artist in self.node_artists:
                        node_id = self.node_artists[self.drag_artist]
                        if ('animal', node_id) in self.temp_positions:
                            current_x, current_y = self.temp_positions[('animal', node_id)]
                            snapped_x, snapped_y = self._snap_to_grid(current_x, current_y)
                            self.temp_positions[('animal', node_id)] = (snapped_x, snapped_y)
                    elif self.drag_artist in self.embryo_artists:
                        transfer_id, embryo_id = self.embryo_artists[self.drag_artist]
                        if ('embryo', embryo_id) in self.temp_positions:
                            current_x, current_y = self.temp_positions[('embryo', embryo_id)]
                            snapped_x, snapped_y = self._snap_to_grid(current_x, current_y)
                            self.temp_positions[('embryo', embryo_id)] = (snapped_x, snapped_y)
                    
                    self._redraw_canvas()
            
            # Execute pending menu action if no drag occurred
            if not self.is_dragging and self.pending_menu_action:
                action = self.pending_menu_action[0]
                if action == 'open_menu':
                    _, animal_name = self.pending_menu_action
                    # Open appropriate menu for this node
                    if animal_name == FREEZER_NODE_NAME:
                        self._open_freezer_embryo_list()
                    else:
                        self._show_animal_info(animal_name)
                elif action == 'edit_embryo':
                    _, transfer_id, embryo_id = self.pending_menu_action
                    self._edit_embryo(transfer_id, embryo_id)
                self.pending_menu_action = None
            else:
                # Clear pending action if drag occurred
                self.pending_menu_action = None
            
            # Reset drag state
            self.drag_active = False
            self.drag_artist = None
            self.drag_offset = (0, 0)
            self.click_start_pos = None
            self.is_dragging = False
    
    def _on_mouse_move(self, event):
        """Handle mouse movement for panning and dragging."""
        # Handle panning
        if self.pan_active and self.pan_start is not None:
            if event.inaxes != self.ax or event.xdata is None or event.ydata is None:
                return
            
            # Calculate pan offset
            dx = event.xdata - self.pan_start[0]
            dy = event.ydata - self.pan_start[1]
            
            # Get current limits
            cur_xlim = self.ax.get_xlim()
            cur_ylim = self.ax.get_ylim()
            
            # Apply pan
            new_xlim = (cur_xlim[0] - dx, cur_xlim[1] - dx)
            new_ylim = (cur_ylim[0] - dy, cur_ylim[1] - dy)
            
            # Apply aspect fill to prevent gutters
            new_xlim, new_ylim = self._apply_aspect_fill(new_xlim, new_ylim)
            
            self.ax.set_xlim(new_xlim)
            self.ax.set_ylim(new_ylim)
            
            # Store state
            self.current_xlim = new_xlim
            self.current_ylim = new_ylim
            
            # Use draw_idle for better performance
            self.canvas.draw_idle()
        
        # Handle dragging
        elif self.drag_active and self.drag_artist is not None:
            if event.xdata is None or event.ydata is None:
                return
            
            # Check if we've exceeded drag threshold
            if not self.is_dragging and self.click_start_pos is not None:
                dx = event.xdata - self.click_start_pos[0]
                dy = event.ydata - self.click_start_pos[1]
                distance = (dx**2 + dy**2)**0.5
                
                if distance > self.drag_threshold:
                    self.is_dragging = True
            
            # Only update position if we're actually dragging
            if self.is_dragging:
                # Update artist position
                new_x = event.xdata + self.drag_offset[0]
                new_y = event.ydata + self.drag_offset[1]
                
                self.drag_artist.set_data([new_x], [new_y])
                
                # Store temporary position
                if self.drag_artist in self.node_artists:
                    node_id = self.node_artists[self.drag_artist]
                    self.temp_positions[('animal', node_id)] = (new_x, new_y)
                elif self.drag_artist in self.embryo_artists:
                    transfer_id, embryo_id = self.embryo_artists[self.drag_artist]
                    self.temp_positions[('embryo', embryo_id)] = (new_x, new_y)
                
                # Just redraw the canvas without clearing (fast update during drag)
                self.canvas.draw_idle()
    
    def _show_hover_annotation(self, text, x, y):
        """Show hover annotation using matplotlib like ProgTrack weight plot."""
        # Remove old annotation
        if self._hover_annotation:
            self._hover_annotation.remove()
            self._hover_annotation = None
        
        # Convert data coordinates to figure coordinates
        disp_x, disp_y = self.ax.transData.transform((x, y))
        fig_x, fig_y = self.figure.transFigure.inverted().transform((disp_x, disp_y))
        
        # Create annotation exactly like ProgTrack weight plot tooltips
        self._hover_annotation = self.figure.text(
            fig_x, fig_y,
            text,
            transform=self.figure.transFigure,
            bbox=dict(boxstyle='round,pad=0.3', fc='yellow', alpha=0.9),
            zorder=20000,
            clip_on=False
        )
        
        self.canvas.draw_idle()
    
    def _on_hover(self, event):
        """Handle hover for tooltips with statistics."""
        if not event.inaxes or event.xdata is None or event.ydata is None:
            # Remove annotation when not over axes
            if self._hover_annotation:
                self._hover_annotation.remove()
                self._hover_annotation = None
                self.canvas.draw_idle()
            return
        
        # Access Role enum
        Role = self.parent_app.Role
        
        # Check lifebar hover first (NEW in v2.0)
        for animal_name, lifebars in self.lifebar_data.items():
            for lifebar in lifebars:
                # Check if annotation contains the hover point
                if 'annotation' in lifebar:
                    ann = lifebar['annotation']
                    contains, _ = ann.contains(event)
                    if contains:
                        # Build tooltip text
                        event_date = lifebar['event_date']
                        if event_date:
                            if isinstance(event_date, str):
                                try:
                                    event_date = datetime.fromisoformat(event_date)
                                    date_str = event_date.strftime('%d.%m.%Y')
                                except (TypeError, ValueError):
                                    date_str = event_date
                            else:
                                date_str = event_date.strftime('%d.%m.%Y')
                        else:
                            date_str = 'Total'
                        
                        if lifebar['efficiency_pct'] is not None:
                            eff_str = f"{lifebar['efficiency_pct']:.1f}%"
                        else:
                            eff_str = self.messages.get('flow_track.no_data', 'No data')
                        
                        tooltip = (f"{animal_name}\n"
                                  f"{date_str}\n"
                                  f"{lifebar['metric_name']}: {eff_str}")
                        
                        self._show_hover_annotation(tooltip, event.xdata, event.ydata)
                        return
                elif 'patches' in lifebar:  # Backward compatibility with old patch-based lifebars
                    border_patch, fill_patch = lifebar['patches']
                    contains_border, _ = border_patch.contains(event)
                    contains_fill, _ = fill_patch.contains(event)
                    
                    if contains_border or contains_fill:
                        date_str = (lifebar['event_date'].strftime('%d.%m.%Y') 
                                   if lifebar['event_date'] else 'Total')
                        
                        if lifebar['efficiency_pct'] is not None:
                            eff_str = f"{lifebar['efficiency_pct']:.1f}%"
                        else:
                            eff_str = self.messages.get('flow_track.no_data', 'No data')
                        
                        tooltip = (f"{animal_name}\n"
                                  f"{date_str}\n"
                                  f"{lifebar['metric_name']}: {eff_str}")
                        
                        self._show_hover_annotation(tooltip, event.xdata, event.ydata)
                        return
        
        # Check registered artists for hover tooltips
        for artist, animal_name in self.node_artists.items():
            if artist.contains(event)[0]:
                animal_data = self.parent_app.animals.get(animal_name, {})
                role = _animal_role_value(animal_data)
                
                # Get node position
                node_x, node_y = artist.get_data()
                
                # Build tooltip with statistics
                tooltip_lines = [animal_name]
                
                # Add role-specific statistics
                if role == Role.SPENDER.value:
                    # Count unique surgery days from unified + legacy fields
                    surgery_date_keys = set()
                    for event in animal_data.get('events', []):
                        if event.get('typ') != 'surgery':
                            continue
                        date_key = self._to_date_key(event.get('datum'))
                        if date_key:
                            surgery_date_keys.add(date_key)
                    surgeries_count = len(surgery_date_keys)
                    max_surgeries = animal_data.get('max_op')
                    if max_surgeries in (None, ''):
                        max_surgeries = '?'

                    if surgeries_count > 0 or max_surgeries != '?':
                        surgeries_text = self.messages.get('flow_track.tooltip.surgeries', 'surgeries')
                        tooltip_lines.append(f"{surgeries_count}/{max_surgeries} {surgeries_text}")

                elif role == Role.SAMENSP.value:
                    # Count unique sperm donation days
                    sperm_date_keys = set()
                    for sperm_entry in animal_data.get('sperm', []):
                        date_key = self._to_date_key(sperm_entry.get('datum'))
                        if date_key:
                            sperm_date_keys.add(date_key)

                    sperm_count = len(sperm_date_keys)
                    max_sperm = animal_data.get('max_spermaproben')
                    if max_sperm in (None, ''):
                        max_sperm = '?'

                    if sperm_count > 0 or max_sperm != '?':
                        sperm_samples_text = self.messages.get('flow_track.tooltip.sperm_samples', 'sperm samples')
                        tooltip_lines.append(f"{sperm_count}/{max_sperm} {sperm_samples_text}")

                elif role == Role.AMME.value:
                    # Count unique transfer days from unified + legacy fields
                    transfer_date_keys = set()
                    for event in animal_data.get('events', []):
                        if event.get('typ') != 'embryo_transfer':
                            continue
                        date_key = self._to_date_key(event.get('datum'))
                        if date_key:
                            transfer_date_keys.add(date_key)
                    transfers_count = len(transfer_date_keys)
                    max_transfers = animal_data.get('max_embryo')
                    if max_transfers in (None, ''):
                        max_transfers = '?'

                    if transfers_count > 0 or max_transfers != '?':
                        transfers_text = self.messages.get('flow_track.tooltip.transfers', 'transfers')
                        tooltip_lines.append(f"{transfers_count}/{max_transfers} {transfers_text}")
                
                tooltip = "\n".join(tooltip_lines)
                self._show_hover_annotation(tooltip, node_x[0], node_y[0])
                return
        
        # Check embryo artists
        for artist, (transfer_id, embryo_id) in self.embryo_artists.items():
            if artist.contains(event)[0]:
                transfer_data = self.manual_data.get('transfers_by_id', {}).get(transfer_id, {})
                embryo = next((e for e in transfer_data.get('embryos', []) 
                              if e['embryo_id'] == embryo_id), None)
                
                if embryo:
                    # Get embryo position
                    emb_x, emb_y = artist.get_data()
                    
                    egg = self._animal_export_name(embryo.get('egg_donor_name', '?'))
                    sperm = self._animal_export_name(embryo.get('sperm_donor_name', '?'))
                    implanted = "Yes" if embryo.get('implanted', embryo.get('pregnant', False)) else "No"
                    cryo = "Yes" if embryo.get('cryopreserved', False) else "No"
                    
                    # Build localized tooltip
                    embryo_label = self.messages.get('flow_track.tooltip.embryo', 'Embryo')
                    egg_label = self.messages.get('flow_track.tooltip.egg', 'Egg:')
                    sperm_label = self.messages.get('flow_track.tooltip.sperm', 'Sperm:')
                    implanted_label = self.messages.get('flow_track.tooltip.implanted', 'Implanted:')
                    cryo_label = self.messages.get('flow_track.tooltip.cryo', 'Cryo:')
                    
                    tooltip = f"{embryo_label}\n{egg_label} {egg}\n{sperm_label} {sperm}\n{implanted_label} {implanted}\n{cryo_label} {cryo}"
                    self._show_hover_annotation(tooltip, emb_x[0], emb_y[0])
                return
        
        # No hover match - remove annotation
        if self._hover_annotation:
            self._hover_annotation.remove()
            self._hover_annotation = None
            self.canvas.draw_idle()
    
    def _on_undo(self):
        """Handle undo action."""
        if not self._can('edit'):
            self._deny()
            return
        action = self.undo_manager.undo()
        if action:
            action.undo(self)
            self._update_undo_redo_buttons()
            self._redraw_canvas()
            if self.settings.get('auto_save_enabled', True):
                self._save_flow_track_data()
    
    def _on_redo(self):
        """Handle redo action."""
        if not self._can('edit'):
            self._deny()
            return
        action = self.undo_manager.redo()
        if action:
            action.execute(self)
            self._update_undo_redo_buttons()
            self._redraw_canvas()
            if self.settings.get('auto_save_enabled', True):
                self._save_flow_track_data()
    
    def _update_undo_redo_buttons(self):
        """Update undo/redo button states."""
        self.undo_btn.setEnabled(self.undo_manager.can_undo())
        self.redo_btn.setEnabled(self.undo_manager.can_redo())
    
    def _open_settings_dialog(self):
        """Open settings dialog with all configuration options."""
        if not self._can("use"):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(self.messages.get("flow_track.settings.title", "Flow Track Settings"))
        dialog.setModal(True)
        dialog.resize(500, 600)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Scrollable area for all settings
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QtWidgets.QWidget()
        scroll_layout = QtWidgets.QVBoxLayout(scroll_widget)
        
        # Display options group
        display_group = QtWidgets.QGroupBox(
            self.messages.get("flow_track.settings.display", "DISPLAY OPTIONS")
        )
        display_layout = QtWidgets.QVBoxLayout()
        
        # Show animal names checkbox
        show_names_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.settings.show_names", "Show animal names")
        )
        show_names_check.setChecked(self.settings.get('show_animal_names', True))
        display_layout.addWidget(show_names_check)
        
        # Show lifebars checkbox (NEW in v2.0)
        show_lifebars_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.settings.show_lifebars", "Show efficiency lifebars")
        )
        show_lifebars_check.setChecked(self.settings.get('show_lifebars', True))
        display_layout.addWidget(show_lifebars_check)

        # Graph scope toggle
        render_full_graph_check = QtWidgets.QCheckBox(
            self.messages.get(
                "flow_track.settings.render_full_graph",
                "Render full graph (ignore sidebar selection)"
            )
        )
        render_full_graph_check.setChecked(self.settings.get('render_full_graph', True))
        display_layout.addWidget(render_full_graph_check)
        
        # Grid options (NEW in v2.0)
        grid_layout = QtWidgets.QHBoxLayout()
        
        show_grid_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.settings.show_grid", "Show grid")
        )
        show_grid_check.setChecked(self.settings.get('show_grid', False))
        grid_layout.addWidget(show_grid_check)
        
        snap_to_grid_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.settings.snap_to_grid", "Snap to grid")
        )
        snap_to_grid_check.setChecked(self.settings.get('snap_to_grid', False))
        grid_layout.addWidget(snap_to_grid_check)
        
        grid_layout.addStretch()
        display_layout.addLayout(grid_layout)
        
        # Timeline checkbox (NEW)
        timeline_check = QtWidgets.QCheckBox(
            self.messages.get("flow_track.settings.show_timeline", "Show Transfer Timeline")
        )
        timeline_check.setChecked(self.settings.get('show_timeline', False))
        timeline_check.stateChanged.connect(self._toggle_timeline_visibility)
        display_layout.addWidget(timeline_check)
        
        display_group.setLayout(display_layout)
        scroll_layout.addWidget(display_group)
        
        # Lifebar metrics group (NEW in v2.0)
        metrics_group = QtWidgets.QGroupBox(
            self.messages.get("flow_track.settings.lifebar_metrics", "LIFEBAR EFFICIENCY METRICS")
        )
        metrics_layout = QtWidgets.QVBoxLayout()
        
        # === EGG DONORS ===
        egg_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.settings.egg_donors', 'Egg Donors:')}</b>")
        metrics_layout.addWidget(egg_label)
        
        # Metric type dropdown - NEW 11 METRICS
        egg_metric_label = QtWidgets.QLabel(self.messages.get("flow_track.settings.metric.label", "Metric:"))
        metrics_layout.addWidget(egg_metric_label)
        
        egg_metric_combo = QtWidgets.QComboBox()
        # 6 IVM-focused metrics matching _calculate_egg_donor_efficiency
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.ivm_rate', 'IVM rate'), "ivm_rate")
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.fert_rate_after_ivm', 'Fertilization rate after IVM'), "fert_rate_after_ivm")
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.usable_from_fert', 'Usable embryo yield from fertilized'), "usable_from_fert")
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.implantation_rate', 'Implantation rate'), "implantation_rate")
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.usable_per_gv', 'Overall usable yield per GV'), "usable_per_gv")
        egg_metric_combo.addItem(self.messages.get('flow_track.settings.metric.fert_per_gv', 'Fertilized per GV'), "fert_per_gv")
        
        # Parse current setting (default to usable yield per GV)
        current_egg_metric = self.settings.get('egg_donor_metric', 'usable_per_gv')
        metric_index = 4  # default to usable yield per GV
        for i in range(egg_metric_combo.count()):
            if egg_metric_combo.itemData(i) == current_egg_metric:
                metric_index = i
                break
        egg_metric_combo.setCurrentIndex(metric_index)
        
        metrics_layout.addWidget(egg_metric_combo)
        
        # Aggregation radiobuttons for egg donors
        egg_agg_layout = QtWidgets.QHBoxLayout()
        egg_agg_group = QtWidgets.QButtonGroup(dialog)
        
        egg_per_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.per_surgery", "Per Surgery"))
        egg_total_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.total", "Total"))
        
        # Parse current aggregation setting
        current_egg_agg = self.settings.get('egg_donor_aggregation', 'total')
        if current_egg_agg == 'per_surgery':
            egg_per_rb.setChecked(True)
        else:
            egg_total_rb.setChecked(True)
        
        egg_agg_group.addButton(egg_per_rb, 0)
        egg_agg_group.addButton(egg_total_rb, 1)
        egg_agg_layout.addWidget(egg_per_rb)
        egg_agg_layout.addWidget(egg_total_rb)
        egg_agg_layout.addStretch()
        metrics_layout.addLayout(egg_agg_layout)
        
        metrics_layout.addSpacing(15)
        
        # === SPERM DONORS ===
        sperm_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.settings.sperm_donors', 'Sperm Donors:')}</b>")
        metrics_layout.addWidget(sperm_label)
        
        # Metric type dropdown - NEW 12 METRICS
        sperm_metric_label = QtWidgets.QLabel(self.messages.get("flow_track.settings.metric.label", "Metric:"))
        metrics_layout.addWidget(sperm_metric_label)
        
        sperm_metric_combo = QtWidgets.QComboBox()
        # 4 IVM-focused metrics matching _calculate_sperm_donor_efficiency
        sperm_metric_combo.addItem(self.messages.get('flow_track.settings.metric.fert_rate_ivm', 'Fertilization rate (IVM M2)'), "fert_rate_ivm")
        sperm_metric_combo.addItem(self.messages.get('flow_track.settings.metric.usable_per_ivm', 'Usable embryo yield per IVM M2'), "usable_per_ivm")
        sperm_metric_combo.addItem(self.messages.get('flow_track.settings.metric.implantation_rate', 'Implantation rate'), "implantation_rate")
        sperm_metric_combo.addItem(self.messages.get('flow_track.settings.metric.fert_rate_in_vivo', 'Fertilization rate (in vivo M2)'), "fert_rate_in_vivo")
        
        # Parse current setting (default to usable yield per IVM M2)
        current_sperm_metric = self.settings.get('sperm_donor_metric', 'usable_per_ivm')
        metric_index = 1  # default to usable yield per IVM M2
        for i in range(sperm_metric_combo.count()):
            if sperm_metric_combo.itemData(i) == current_sperm_metric:
                metric_index = i
                break
        sperm_metric_combo.setCurrentIndex(metric_index)
        
        metrics_layout.addWidget(sperm_metric_combo)
        
        # Aggregation radiobuttons for sperm donors
        sperm_agg_layout = QtWidgets.QHBoxLayout()
        sperm_agg_group = QtWidgets.QButtonGroup(dialog)
        
        sperm_per_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.per_donation", "Per Donation"))
        sperm_total_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.total", "Total"))
        
        # Parse current aggregation setting
        current_sperm_agg = self.settings.get('sperm_donor_aggregation', 'total')
        if current_sperm_agg == 'per_donation':
            sperm_per_rb.setChecked(True)
        else:
            sperm_total_rb.setChecked(True)
        
        sperm_agg_group.addButton(sperm_per_rb, 0)
        sperm_agg_group.addButton(sperm_total_rb, 1)
        sperm_agg_layout.addWidget(sperm_per_rb)
        sperm_agg_layout.addWidget(sperm_total_rb)
        sperm_agg_layout.addStretch()
        metrics_layout.addLayout(sperm_agg_layout)
        
        metrics_layout.addSpacing(15)
        
        # === SURROGATES ===
        surr_label = QtWidgets.QLabel(f"<b>{self.messages.get('flow_track.settings.surrogates', 'Surrogates:')}</b>")
        metrics_layout.addWidget(surr_label)
        
        # Metric type dropdown - 2 metrics matching _calculate_surrogate_efficiency
        surr_metric_label = QtWidgets.QLabel(self.messages.get("flow_track.settings.metric.label", "Metric:"))
        metrics_layout.addWidget(surr_metric_label)
        
        surr_metric_combo = QtWidgets.QComboBox()
        surr_metric_combo.addItem(self.messages.get('flow_track.settings.metric.total_implantation', 'Total Implantation rate'), "total_implantation")
        surr_metric_combo.addItem(self.messages.get('flow_track.settings.metric.transfer_success_rate', 'Transfer success rate'), "transfer_success_rate")
        
        # Parse current setting (default to total implantation)
        current_surr_metric_key = self.settings.get('surrogate_metric', 'total_implantation')
        metric_index = 0  # default to total implantation
        for i in range(surr_metric_combo.count()):
            if surr_metric_combo.itemData(i) == current_surr_metric_key:
                metric_index = i
                break
        surr_metric_combo.setCurrentIndex(metric_index)
        
        metrics_layout.addWidget(surr_metric_combo)
        
        # Aggregation radiobuttons
        surr_agg_layout = QtWidgets.QHBoxLayout()
        surr_agg_group = QtWidgets.QButtonGroup(dialog)
        
        surr_per_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.per_transfer", "Per Transfer"))
        surr_total_rb = QtWidgets.QRadioButton(self.messages.get("flow_track.settings.aggregation.total", "Total"))
        
        current_surr_metric = self.settings.get('surrogate_metric', 'total_implantation')
        if current_surr_metric.startswith('per_'):
            surr_per_rb.setChecked(True)
        else:
            surr_total_rb.setChecked(True)
        
        surr_agg_group.addButton(surr_per_rb, 0)
        surr_agg_group.addButton(surr_total_rb, 1)
        surr_agg_layout.addWidget(surr_per_rb)
        surr_agg_layout.addWidget(surr_total_rb)
        surr_agg_layout.addStretch()
        metrics_layout.addLayout(surr_agg_layout)
        
        metrics_group.setLayout(metrics_layout)
        scroll_layout.addWidget(metrics_group)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        layout.addSpacing(10)
        
        # Export button
        export_btn = QtWidgets.QPushButton(
            self.messages.get("flow_track.settings.export", "Export Data...")
        )
        export_btn.clicked.connect(lambda: self._export_dialog())
        layout.addWidget(export_btn)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        ok_btn = QtWidgets.QPushButton(self.messages.get("button.ok", "OK"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            # Update settings
            self.settings['show_animal_names'] = show_names_check.isChecked()
            self.settings['show_lifebars'] = show_lifebars_check.isChecked()
            self.settings['render_full_graph'] = render_full_graph_check.isChecked()
            self.settings['show_grid'] = show_grid_check.isChecked()
            self.settings['snap_to_grid'] = snap_to_grid_check.isChecked()
            
            # Save egg donor metric and aggregation
            self.settings['egg_donor_metric'] = egg_metric_combo.currentData()
            self.settings['egg_donor_aggregation'] = 'per_surgery' if egg_per_rb.isChecked() else 'total'
            
            # Save sperm donor metric and aggregation
            self.settings['sperm_donor_metric'] = sperm_metric_combo.currentData()
            self.settings['sperm_donor_aggregation'] = 'per_donation' if sperm_per_rb.isChecked() else 'total'
            
            # Save surrogate metric and aggregation
            self.settings['surrogate_metric'] = surr_metric_combo.currentData()
            self.settings['surrogate_aggregation'] = 'per_transfer' if surr_per_rb.isChecked() else 'total'
            
            # Save settings and redraw to apply changes
            self._save_settings()
            self._redraw_canvas()
            logger.info("Settings updated")
    
    def _export_dialog(self):
        """Show export dialog for flow track data."""
        if not self._can("use"):
            self._deny()
            return
        QtWidgets = self.parent_app.QtWidgets
        
        # Create dialog
        dialog = QtWidgets.QDialog(self.widget)
        dialog.setWindowTitle(self.messages.get("flow_track.export.title", "Export Flow Track Data"))
        dialog.setModal(True)
        
        layout = QtWidgets.QVBoxLayout()
        
        # Format selection
        format_group = QtWidgets.QGroupBox(
            self.messages.get("flow_track.export.format", "Export Format")
        )
        format_layout = QtWidgets.QVBoxLayout()
        
        rb_json = QtWidgets.QRadioButton(
            self.messages.get("flow_track.export.json", "JSON Backup")
        )
        rb_json.setChecked(True)
        format_layout.addWidget(rb_json)
        
        rb_excel = QtWidgets.QRadioButton(
            self.messages.get("flow_track.export.excel", "Excel (.xlsx)")
        )
        format_layout.addWidget(rb_excel)
        
        rb_pdf = QtWidgets.QRadioButton(
            self.messages.get("flow_track.export.pdf", "PDF Report")
        )
        format_layout.addWidget(rb_pdf)
        
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        export_btn = QtWidgets.QPushButton(self.messages.get("button.export", "Export"))
        cancel_btn = QtWidgets.QPushButton(self.messages.get("button.cancel", "Cancel"))
        
        export_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        button_layout.addWidget(export_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self._set_scrollable_dialog_layout(dialog, layout)
        
        # Show dialog and export
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            if rb_json.isChecked():
                self._export_json()
            elif rb_excel.isChecked():
                self._export_excel()
            elif rb_pdf.isChecked():
                self._export_pdf()
    
    def _export_json(self):
        """Export flow track data as JSON backup."""
        QtWidgets = self.parent_app.QtWidgets
        
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self.widget,
            self.messages.get("flow_track.export.save_json", "Save JSON Backup"),
            str(default_save_path(f"flow_track_backup_{datetime.now().strftime('%Y%m%d')}.json")),
            self.messages.get("flow_track.export.json_filter", "JSON files (*.json)")
        )
        
        if filename:
            try:
                data = {
                    'schema_version': '3.0',
                    'export_date': datetime.now().isoformat(),
                    'animal_identity': {
                        animal_name: {
                            'ipid': animal_name,
                            'name': animal_base_name(animal_name, record),
                        }
                        for animal_name, record in sorted((self.parent_app.animals or {}).items())
                    },
                    'transfers': self.manual_data.get('transfers_by_id', {}),
                    'manual_data': self.manual_data
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, default=str, ensure_ascii=False)
                
                QtWidgets.QMessageBox.information(
                    self.widget,
                    self.messages.get("flow_track.export.success.title", "Export Successful"),
                    self.messages.get("flow_track.export.success.message", "Data exported to:\n{filename}").format(filename=filename)
                )
            except Exception as e:
                logger.error(f"Export failed: {e}")
                QtWidgets.QMessageBox.critical(
                    self.widget,
                    self.messages.get("error.title", "Error"),
                    self.messages.get("flow_track.export.error.json", "Export failed:\n{error}").format(error=e)
            )
    
    def _animal_export_name(self, animal_name):
        """Return the human-readable animal name for exports while keeping IPID elsewhere."""
        records = dict(self.parent_app.animals or {})
        archived = getattr(self.parent_app, 'archived', {}) or {}
        if isinstance(archived, dict):
            records.update(archived)
        record = records.get(animal_name, {})
        return animal_base_name(animal_name, record)

    def _safe_excel_sheet_token(self, value, max_length=28):
        text = str(value or '').strip() or 'Animal'
        for char in '\\/:*?[]|':
            text = text.replace(char, '_')
        text = text[:max_length].strip()
        return text or 'Animal'

    def _create_animal_sheet(self, wb, prefix, animal_name):
        name_token = self._safe_excel_sheet_token(self._animal_export_name(animal_name), 22)
        suffix_source = ''.join(ch for ch in str(animal_name) if ch.isalnum())
        suffix = (suffix_source[-4:] or 'IPID')[:4]
        base_title = f"{prefix}_{name_token}_{suffix}"
        title = self._safe_excel_sheet_token(base_title, 31)
        counter = 2
        while title in wb.sheetnames:
            counter_suffix = f"_{counter}"
            title = f"{self._safe_excel_sheet_token(base_title, 31 - len(counter_suffix))}{counter_suffix}"
            counter += 1
        return wb.create_sheet(title)

    def _export_excel(self):
        """Export comprehensive flow track data to Excel with formulas."""
        QtWidgets = self.parent_app.QtWidgets
        
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self.widget,
            self.messages.get("flow_track.export.save_excel", "Save Excel File"),
            str(default_save_path(f"flow_track_{datetime.now().strftime('%Y%m%d')}.xlsx")),
            self.messages.get("flow_track.export.excel_filter", "Excel files (*.xlsx)")
        )
        
        if not filename:
            return
            
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            Role = self.parent_app.Role
            
            # === 1. ALL ANIMALS OVERVIEW ===
            self._create_all_animals_overview(wb)
            
            # === 2. TYPE-SPECIFIC OVERVIEWS ===
            self._create_egg_donors_overview(wb)
            self._create_sperm_donors_overview(wb)
            self._create_surrogates_overview(wb)
            
            # === 3. INDIVIDUAL ANIMAL SHEETS ===
            # Egg Donors
            for animal_name, animal_data in sorted(self.parent_app.animals.items()):
                if _animal_role_value(animal_data) == Role.SPENDER.value:
                    self._create_egg_donor_sheet(wb, animal_name)
            
            # Sperm Donors
            for animal_name, animal_data in sorted(self.parent_app.animals.items()):
                if _animal_role_value(animal_data) == Role.SAMENSP.value:
                    self._create_sperm_donor_sheet(wb, animal_name)
            
            # Surrogates
            for animal_name, animal_data in sorted(self.parent_app.animals.items()):
                if _animal_role_value(animal_data) == Role.AMME.value:
                    self._create_surrogate_sheet(wb, animal_name)
            
            # === 4. ALL EMBRYOS SHEET ===
            self._create_all_embryos_sheet(wb)
            
            # === 5. FREEZER SHEET ===
            self._create_freezer_sheet(wb)
            
            # Save workbook
            wb.save(filename)
            
            QtWidgets.QMessageBox.information(
                self.widget,
                self.messages.get("flow_track.export.success.title", "Export Successful"),
                self.messages.get("flow_track.export.success.message", "Data exported to:\n{filename}").format(filename=filename)
            )
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}", exc_info=True)
            QtWidgets.QMessageBox.critical(
                self.widget,
                self.messages.get("error.title", "Error"),
                self.messages.get("flow_track.export.error.excel", "Excel export failed:\n{error}").format(error=str(e))
            )
    
    def _flow_tracked_animals(self) -> set[str]:
        """Return animal IDs directly referenced by Flow Track transfer data."""
        tracked: set[str] = set()
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            surrogate_name = transfer_data.get('surrogate_name')
            if surrogate_name:
                tracked.add(surrogate_name)
            for embryo in transfer_data.get('embryos', []):
                for key in ('egg_donor_name', 'sperm_donor_name'):
                    animal_name = embryo.get(key)
                    if animal_name:
                        tracked.add(animal_name)
        return tracked

    def _export_animal_records(self) -> Dict[str, Any]:
        records = dict(getattr(self.parent_app, 'animals', {}) or {})
        archived = getattr(self.parent_app, 'archived', {}) or {}
        if isinstance(archived, dict):
            records.update(archived)
        return records

    def _create_all_animals_overview(self, wb):
        """Create overview sheet with animals referenced by Flow Track data."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("All Animals Overview", 0)
        Role = self.parent_app.Role
        
        headers = ['IPID', 'Animal', 'Role', 'Embryos created', 'Transferred', 'Implanted', 'Cryopreserved']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        records = self._export_animal_records()
        role_order = {
            Role.SPENDER.value: 0,
            Role.SAMENSP.value: 1,
            Role.AMME.value: 2,
        }
        flow_animals = [
            animal_name
            for animal_name in self._flow_tracked_animals()
            if animal_name in records
        ]
        flow_animals.sort(
            key=lambda name: (
                role_order.get(_animal_role_value(records.get(name, {})), 99),
                self._animal_export_name(name).casefold(),
                name.casefold(),
            )
        )

        for animal_name in flow_animals:
            animal_data = records.get(animal_name, {})
            role = _animal_role_value(animal_data)
            
            if role == Role.SPENDER.value:
                role_name = "Egg Donor"
            elif role == Role.SAMENSP.value:
                role_name = "Sperm Donor"
            elif role == Role.AMME.value:
                role_name = "Surrogate"
            else:
                role_name = "Other"
            
            embryos_created = 0
            transferred = 0
            implanted = 0
            cryopreserved = 0
            
            for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
                for embryo in transfer_data.get('embryos', []):
                    if embryo.get('egg_donor_name') == animal_name or embryo.get('sperm_donor_name') == animal_name:
                        embryos_created += 1
                    
                    if transfer_data.get('surrogate_name') == animal_name and transfer_id != FREEZER_TRANSFER_ID:
                        transferred += 1
                        if embryo.get('implanted', False):
                            implanted += 1
                    
                    if transfer_id == FREEZER_TRANSFER_ID:
                        if embryo.get('egg_donor_name') == animal_name or embryo.get('sperm_donor_name') == animal_name:
                            cryopreserved += 1
            
            ws.append([
                animal_name,
                self._animal_export_name(animal_name),
                role_name,
                embryos_created,
                transferred,
                implanted,
                cryopreserved,
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_egg_donors_overview(self, wb):
        """Create overview sheet for all egg donors."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("Egg Donors Overview")
        Role = self.parent_app.Role
        
        headers = ['IPID', 'Egg Donor', 'Surgeries', 'Total GV', 'Total IVM M2', 'IVM M2 Fert',
                   'Transferred', 'Frozen', 'Implanted', 'IVM %', 'Fert %', 'Impl %']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF1493", end_color="FF1493", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for animal_name in sorted(self.parent_app.animals.keys()):
            if _animal_role_value(self.parent_app.animals[animal_name]) != Role.SPENDER.value:
                continue
            
            efficiency = self._calculate_egg_donor_efficiency(animal_name)
            total = efficiency['total']
            
            num_surgeries = len(efficiency['per_surgery'])
            ws.append([animal_name, self._animal_export_name(animal_name), num_surgeries,
                       total.get('total_gv_isolated', 0),
                       total.get('total_ivm_m2', 0),
                       total.get('total_ivm_m2_fertilized', 0),
                       total.get('total_transferred_ivm', 0),
                       total.get('total_frozen_ivm', 0),
                       total.get('total_implanted_ivm', 0),
                       '', '', ''])
            
            ws[f'J{row}'] = f'=IF(D{row}>0, E{row}/D{row}*100, 0)'
            ws[f'K{row}'] = f'=IF(E{row}>0, F{row}/E{row}*100, 0)'
            ws[f'L{row}'] = f'=IF(G{row}>0, I{row}/G{row}*100, 0)'
            ws[f'J{row}'].number_format = '0.0'
            ws[f'K{row}'].number_format = '0.0'
            ws[f'L{row}'].number_format = '0.0'
            row += 1
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_sperm_donors_overview(self, wb):
        """Create overview sheet for all sperm donors."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("Sperm Donors Overview")
        Role = self.parent_app.Role
        
        headers = ['IPID', 'Sperm Donor', 'Donations', 'IVM M2 Insem', 'IVM M2 Fert',
                   'Transferred', 'Frozen', 'Implanted', 'Fert %', 'Impl %']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for animal_name in sorted(self.parent_app.animals.keys()):
            if _animal_role_value(self.parent_app.animals[animal_name]) != Role.SAMENSP.value:
                continue
            
            efficiency = self._calculate_sperm_donor_efficiency(animal_name)
            total = efficiency['total']
            
            num_donations = len(efficiency['per_donation'])
            ws.append([animal_name, self._animal_export_name(animal_name), num_donations,
                       total.get('total_ivm_m2_inseminated', 0),
                       total.get('total_ivm_m2_fertilized', 0),
                       total.get('total_transferred_ivm', 0),
                       total.get('total_frozen_ivm', 0),
                       total.get('total_implanted_ivm', 0),
                       '', ''])
            
            ws[f'I{row}'] = f'=IF(D{row}>0, E{row}/D{row}*100, 0)'
            ws[f'J{row}'] = f'=IF(F{row}>0, H{row}/F{row}*100, 0)'
            ws[f'I{row}'].number_format = '0.0'
            ws[f'J{row}'].number_format = '0.0'
            row += 1
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_surrogates_overview(self, wb):
        """Create overview sheet for all surrogates."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("Surrogates Overview")
        Role = self.parent_app.Role
        
        headers = ['IPID', 'Surrogate', 'Transfers', 'Embryos Received', 'Implanted',
                   'Implantation %', 'Success Rate %']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for animal_name in sorted(self.parent_app.animals.keys()):
            if _animal_role_value(self.parent_app.animals[animal_name]) != Role.AMME.value:
                continue
            
            efficiency = self._calculate_surrogate_efficiency(animal_name)
            total = efficiency['total']
            
            ws.append([animal_name, self._animal_export_name(animal_name),
                       total.get('total_transfers', 0),
                       total.get('total_embryos', 0),
                       total.get('total_implanted', 0),
                       '', ''])
            
            ws[f'F{row}'] = f'=IF(D{row}>0, E{row}/D{row}*100, 0)'
            ws[f'G{row}'] = f'=IF(C{row}>0, {total.get("successful_transfers", 0)}/C{row}*100, 0)'
            ws[f'F{row}'].number_format = '0.0'
            ws[f'G{row}'].number_format = '0.0'
            row += 1
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_egg_donor_sheet(self, wb, animal_name):
        """Create individual egg donor sheet with per-surgery data + total row with formulas."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        efficiency = self._calculate_egg_donor_efficiency(animal_name)
        if not efficiency['per_surgery'] and efficiency['total'].get('total_gv_isolated', 0) == 0:
            return
        
        ws = self._create_animal_sheet(wb, "ED", animal_name)
        ws.append(['IPID', animal_name])
        ws.append(['Animal', self._animal_export_name(animal_name)])
        ws.append([])
        
        headers = ['Date', 'GV Isolated', 'IVM M2', 'IVM M2 Fert', 'Transferred', 'Frozen', 
                   'Implanted', 'IVM %', 'Fert %', 'Impl %']
        ws.append(headers)
        
        for cell in ws[4]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF1493", end_color="FF1493", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 5
        for surgery in efficiency['per_surgery']:
            date_obj = surgery['date']
            if date_obj and hasattr(date_obj, 'strftime'):
                date_str = date_obj.strftime('%d.%m.%Y')
            else:
                date_str = str(date_obj) if date_obj else 'Unknown'
            
            ws.append([date_str,
                       surgery.get('gv_isolated', 0),
                       surgery.get('ivm_m2', 0),
                       surgery.get('ivm_m2_fertilized', 0),
                       surgery.get('transferred_ivm', 0),
                       surgery.get('frozen_ivm', 0),
                       surgery.get('implanted_ivm', 0),
                       '', '', ''])
            
            ws[f'H{row}'] = f'=IF(B{row}>0, C{row}/B{row}*100, 0)'
            ws[f'I{row}'] = f'=IF(C{row}>0, D{row}/C{row}*100, 0)'
            ws[f'J{row}'] = f'=IF(E{row}>0, G{row}/E{row}*100, 0)'
            ws[f'H{row}'].number_format = '0.0'
            ws[f'I{row}'].number_format = '0.0'
            ws[f'J{row}'].number_format = '0.0'
            row += 1
        
        total_row = row
        ws.append(['TOTAL', '', '', '', '', '', '', '', '', ''])
        ws[f'A{total_row}'].font = Font(bold=True, size=12)
        ws[f'B{total_row}'] = f'=SUM(B5:B{total_row-1})'
        ws[f'C{total_row}'] = f'=SUM(C5:C{total_row-1})'
        ws[f'D{total_row}'] = f'=SUM(D5:D{total_row-1})'
        ws[f'E{total_row}'] = f'=SUM(E5:E{total_row-1})'
        ws[f'F{total_row}'] = f'=SUM(F5:F{total_row-1})'
        ws[f'G{total_row}'] = f'=SUM(G5:G{total_row-1})'
        ws[f'H{total_row}'] = f'=IF(B{total_row}>0, C{total_row}/B{total_row}*100, 0)'
        ws[f'I{total_row}'] = f'=IF(C{total_row}>0, D{total_row}/C{total_row}*100, 0)'
        ws[f'J{total_row}'] = f'=IF(E{total_row}>0, G{total_row}/E{total_row}*100, 0)'
        ws[f'H{total_row}'].number_format = '0.0'
        ws[f'I{total_row}'].number_format = '0.0'
        ws[f'J{total_row}'].number_format = '0.0'
        
        ws[f'A{total_row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{total_row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            ws[f'{col}{total_row}'].font = Font(bold=True)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        embryo_start_row = total_row + 3
        ws[f'A{embryo_start_row}'] = 'Associated Embryos:'
        ws[f'A{embryo_start_row}'].font = Font(bold=True, size=11)
        embryo_start_row += 1
        
        embryo_headers = ['Embryo ID', 'Sperm Donor IPID', 'Sperm Donor', 'Surrogate IPID', 'Surrogate', 'Transfer Date', 'Stage', 'Implanted', 'Frozen']
        ws.append(embryo_headers)
        for cell in ws[embryo_start_row]:
            cell.font = Font(bold=True)
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            for embryo in transfer_data.get('embryos', []):
                if embryo.get('egg_donor_name') == animal_name:
                    ws.append([
                        embryo.get('embryo_id', ''),
                        embryo.get('sperm_donor_name', ''),
                        self._animal_export_name(embryo.get('sperm_donor_name', '')),
                        transfer_data.get('surrogate_name', ''),
                        self._animal_export_name(transfer_data.get('surrogate_name', '')),
                        transfer_data.get('transfer_date', ''),
                        embryo.get('stage', ''),
                        'Yes' if embryo.get('implanted', False) else 'No',
                        'Yes' if transfer_id == FREEZER_TRANSFER_ID else 'No'
                    ])
    
    def _create_sperm_donor_sheet(self, wb, animal_name):
        """Create individual sperm donor sheet with per-donation data + total row with formulas."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        efficiency = self._calculate_sperm_donor_efficiency(animal_name)
        if not efficiency['per_donation'] and efficiency['total'].get('total_ivm_m2_inseminated', 0) == 0:
            return
        
        ws = self._create_animal_sheet(wb, "SD", animal_name)
        ws.append(['IPID', animal_name])
        ws.append(['Animal', self._animal_export_name(animal_name)])
        ws.append([])
        
        headers = ['Date', 'IVM M2 Insem', 'IVM M2 Fert', 'Transferred', 'Frozen', 
                   'Implanted', 'Fert %', 'Impl %']
        ws.append(headers)
        
        for cell in ws[4]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 5
        for donation in efficiency['per_donation']:
            date_obj = donation['date']
            if date_obj and hasattr(date_obj, 'strftime'):
                date_str = date_obj.strftime('%d.%m.%Y')
            else:
                date_str = str(date_obj) if date_obj else 'Unknown'
            
            ws.append([date_str,
                       donation.get('ivm_m2_inseminated', 0),
                       donation.get('ivm_m2_fertilized', 0),
                       donation.get('transferred_ivm', 0),
                       donation.get('frozen_ivm', 0),
                       donation.get('implanted_ivm', 0),
                       '', ''])
            
            ws[f'G{row}'] = f'=IF(B{row}>0, C{row}/B{row}*100, 0)'
            ws[f'H{row}'] = f'=IF(D{row}>0, F{row}/D{row}*100, 0)'
            ws[f'G{row}'].number_format = '0.0'
            ws[f'H{row}'].number_format = '0.0'
            row += 1
        
        total_row = row
        ws.append(['TOTAL', '', '', '', '', '', '', ''])
        ws[f'A{total_row}'].font = Font(bold=True, size=12)
        ws[f'B{total_row}'] = f'=SUM(B5:B{total_row-1})'
        ws[f'C{total_row}'] = f'=SUM(C5:C{total_row-1})'
        ws[f'D{total_row}'] = f'=SUM(D5:D{total_row-1})'
        ws[f'E{total_row}'] = f'=SUM(E5:E{total_row-1})'
        ws[f'F{total_row}'] = f'=SUM(F5:F{total_row-1})'
        ws[f'G{total_row}'] = f'=IF(B{total_row}>0, C{total_row}/B{total_row}*100, 0)'
        ws[f'H{total_row}'] = f'=IF(D{total_row}>0, F{total_row}/D{total_row}*100, 0)'
        ws[f'G{total_row}'].number_format = '0.0'
        ws[f'H{total_row}'].number_format = '0.0'
        
        ws[f'A{total_row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{total_row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            ws[f'{col}{total_row}'].font = Font(bold=True)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        embryo_start_row = total_row + 3
        ws[f'A{embryo_start_row}'] = 'Associated Embryos:'
        ws[f'A{embryo_start_row}'].font = Font(bold=True, size=11)
        embryo_start_row += 1
        
        embryo_headers = ['Embryo ID', 'Egg Donor IPID', 'Egg Donor', 'Surrogate IPID', 'Surrogate', 'Transfer Date', 'Stage', 'Implanted', 'Frozen']
        ws.append(embryo_headers)
        for cell in ws[embryo_start_row]:
            cell.font = Font(bold=True)
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            for embryo in transfer_data.get('embryos', []):
                if embryo.get('sperm_donor_name') == animal_name:
                    ws.append([
                        embryo.get('embryo_id', ''),
                        embryo.get('egg_donor_name', ''),
                        self._animal_export_name(embryo.get('egg_donor_name', '')),
                        transfer_data.get('surrogate_name', ''),
                        self._animal_export_name(transfer_data.get('surrogate_name', '')),
                        transfer_data.get('transfer_date', ''),
                        embryo.get('stage', ''),
                        'Yes' if embryo.get('implanted', False) else 'No',
                        'Yes' if transfer_id == FREEZER_TRANSFER_ID else 'No'
                    ])
    
    def _create_surrogate_sheet(self, wb, animal_name):
        """Create individual surrogate sheet with per-transfer data + total row with formulas."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        efficiency = self._calculate_surrogate_efficiency(animal_name)
        if not efficiency['per_transfer'] and efficiency['total'].get('total_embryos', 0) == 0:
            return
        
        ws = self._create_animal_sheet(wb, "SU", animal_name)
        ws.append(['IPID', animal_name])
        ws.append(['Animal', self._animal_export_name(animal_name)])
        ws.append([])
        
        headers = ['Date', 'Embryos Received', 'In Vivo', 'In Vitro', 'Implanted', 'Implant %']
        ws.append(headers)
        
        for cell in ws[4]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 5
        for transfer in efficiency['per_transfer']:
            date_obj = transfer['date']
            if date_obj and hasattr(date_obj, 'strftime'):
                date_str = date_obj.strftime('%d.%m.%Y')
            else:
                date_str = str(date_obj) if date_obj else 'Unknown'
            
            ws.append([date_str,
                       transfer.get('embryo_count_all', 0),
                       transfer.get('embryo_count_in_vivo', 0),
                       transfer.get('embryo_count_in_vitro', 0),
                       transfer.get('implanted_count_all', 0),
                       ''])
            
            ws[f'F{row}'] = f'=IF(B{row}>0, E{row}/B{row}*100, 0)'
            ws[f'F{row}'].number_format = '0.0'
            row += 1
        
        total_row = row
        ws.append(['TOTAL', '', '', '', '', ''])
        ws[f'A{total_row}'].font = Font(bold=True, size=12)
        ws[f'B{total_row}'] = f'=SUM(B5:B{total_row-1})'
        ws[f'C{total_row}'] = f'=SUM(C5:C{total_row-1})'
        ws[f'D{total_row}'] = f'=SUM(D5:D{total_row-1})'
        ws[f'E{total_row}'] = f'=SUM(E5:E{total_row-1})'
        ws[f'F{total_row}'] = f'=IF(B{total_row}>0, E{total_row}/B{total_row}*100, 0)'
        ws[f'F{total_row}'].number_format = '0.0'
        
        ws[f'A{total_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{total_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws[f'{col}{total_row}'].font = Font(bold=True)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        embryo_start_row = total_row + 3
        ws[f'A{embryo_start_row}'] = 'Received Embryos:'
        ws[f'A{embryo_start_row}'].font = Font(bold=True, size=11)
        embryo_start_row += 1
        
        embryo_headers = ['Embryo ID', 'Egg Donor IPID', 'Egg Donor', 'Sperm Donor IPID', 'Sperm Donor', 'Transfer Date', 'Stage', 'Implanted']
        ws.append(embryo_headers)
        for cell in ws[embryo_start_row]:
            cell.font = Font(bold=True)
        
        for transfer_id, transfer_data in self.manual_data.get('transfers_by_id', {}).items():
            if transfer_data.get('surrogate_name') == animal_name and transfer_id != FREEZER_TRANSFER_ID:
                for embryo in transfer_data.get('embryos', []):
                    ws.append([
                        embryo.get('embryo_id', ''),
                        embryo.get('egg_donor_name', ''),
                        self._animal_export_name(embryo.get('egg_donor_name', '')),
                        embryo.get('sperm_donor_name', ''),
                        self._animal_export_name(embryo.get('sperm_donor_name', '')),
                        transfer_data.get('transfer_date', ''),
                        embryo.get('stage', ''),
                        'Yes' if embryo.get('implanted', False) else 'No'
                    ])
    
    def _create_all_embryos_sheet(self, wb):
        """Create sheet with all embryos from all transfers."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("All Embryos")
        
        headers = ['Embryo ID', 'Egg Donor IPID', 'Egg Donor', 'Sperm Donor IPID', 'Sperm Donor',
                   'Surrogate IPID', 'Surrogate', 'Transfer Date', 'Stage', 'Implanted', 'Frozen', 'Comment']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for transfer_id, transfer_data in sorted(self.manual_data.get('transfers_by_id', {}).items()):
            for embryo in transfer_data.get('embryos', []):
                ws.append([
                    embryo.get('embryo_id', ''),
                    embryo.get('egg_donor_name', ''),
                    self._animal_export_name(embryo.get('egg_donor_name', '')),
                    embryo.get('sperm_donor_name', ''),
                    self._animal_export_name(embryo.get('sperm_donor_name', '')),
                    transfer_data.get('surrogate_name', '') if transfer_id != FREEZER_TRANSFER_ID else '',
                    self._animal_export_name(transfer_data.get('surrogate_name', '')) if transfer_id != FREEZER_TRANSFER_ID else '',
                    transfer_data.get('transfer_date', '') if transfer_id != FREEZER_TRANSFER_ID else '',
                    embryo.get('stage', ''),
                    'Yes' if embryo.get('implanted', False) else 'No',
                    'Yes' if transfer_id == FREEZER_TRANSFER_ID else 'No',
                    embryo.get('comment', '')
                ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_freezer_sheet(self, wb):
        """Create sheet with all frozen embryos."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("Freezer")
        
        headers = ['Embryo ID', 'Egg Donor IPID', 'Egg Donor', 'Sperm Donor IPID', 'Sperm Donor', 'Freeze Date', 'Stage', 'Comment']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="000000")
            cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        freezer_transfer = self.manual_data.get('transfers_by_id', {}).get(FREEZER_TRANSFER_ID, {})
        embryos = freezer_transfer.get('embryos', [])
        
        for embryo in embryos:
            ws.append([
                embryo.get('embryo_id', ''),
                embryo.get('egg_donor_name', ''),
                self._animal_export_name(embryo.get('egg_donor_name', '')),
                embryo.get('sperm_donor_name', ''),
                self._animal_export_name(embryo.get('sperm_donor_name', '')),
                embryo.get('freeze_date', ''),
                embryo.get('stage', ''),
                embryo.get('comment', '')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_pdf(self):
        """Export flow track visualization as PDF (including timeline if visible)."""
        QtWidgets = self.parent_app.QtWidgets
        
        # Check if timeline is visible
        include_timeline = self.settings.get('show_timeline', False)
        
        if include_timeline:
            # Create combined figure with both charts
            from matplotlib.backends.backend_pdf import PdfPages
            
            filename, _ = QtWidgets.QFileDialog.getSaveFileName(
                self.widget,
                self.messages.get("flow_track.export.save_pdf", "Save PDF Report"),
                str(default_save_path(f"flow_track_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")),
                self.messages.get("flow_track.export.pdf_filter", "PDF files (*.pdf)")
            )
            
            if filename:
                try:
                    with PdfPages(filename) as pdf:
                        # Export main Flow Track chart
                        pdf.savefig(self.figure, bbox_inches='tight')
                        
                        # Export timeline chart on separate page
                        pdf.savefig(self.timeline_figure, bbox_inches='tight')
                    from Plugins.core.institution_branding import brand_generated_pdf
                    brand_generated_pdf(self.parent_app, filename)
                    
                    QtWidgets.QMessageBox.information(
                        self.widget,
                        self.messages.get("flow_track.export.success.title", "Export Successful"),
                        self.messages.get("flow_track.export.pdf.success", 
                                         "Flow Track and timeline exported successfully to PDF")
                    )
                except Exception as e:
                    QtWidgets.QMessageBox.critical(
                        self.widget,
                        self.messages.get("flow_track.export.error.title", "Export Failed"),
                        self.messages.get("flow_track.export.pdf.error", 
                                         f"PDF export failed:\n{e}")
                    )
        else:
            # Existing PDF export logic for main chart only
            filename, _ = QtWidgets.QFileDialog.getSaveFileName(
                self.widget,
                self.messages.get("flow_track.export.save_pdf", "Save PDF Report"),
                str(default_save_path(f"flow_track_{datetime.now().strftime('%Y%m%d')}.pdf")),
                self.messages.get("flow_track.export.pdf_filter", "PDF files (*.pdf)")
            )
            
            if filename:
                try:
                    # Save current figure as PDF
                    self.figure.savefig(filename, format='pdf', bbox_inches='tight')
                    from Plugins.core.institution_branding import brand_generated_pdf
                    brand_generated_pdf(self.parent_app, filename)
                    
                    QtWidgets.QMessageBox.information(
                        self.widget,
                        self.messages.get("flow_track.export.success.title", "Export Successful"),
                        self.messages.get("flow_track.export.success.message", "Data exported to:\n{filename}").format(filename=filename)
                    )
                except Exception as e:
                    logger.error(f"PDF export failed: {e}")
                    QtWidgets.QMessageBox.critical(
                        self.widget,
                        self.messages.get("error.title", "Error"),
                        self.messages.get("flow_track.export.error.pdf", "PDF export failed:\n{error}").format(error=e)
                    )
    
    def _fit_automatically(self):
        """Reset all node positions to automatic layout and fit view to content."""
        # Clear all temporary positions (nodes return to algorithmic placement)
        self.temp_positions.clear()
        
        # Reset zoom/pan to auto-calculated bounds
        self.current_xlim = None
        self.current_ylim = None
        
        # Redraw with fresh automatic layout
        self._redraw_canvas()
    
    def _smart_refresh(self):
        """Refresh data from ProgTrack."""
        self._populate_events_from_progtrack()
        self._redraw_canvas()
        
        QtWidgets = self.parent_app.QtWidgets
        QtWidgets.QMessageBox.information(
            self.widget,
            self.messages.get("flow_track.refresh.title", "Refresh Complete"),
            self.messages.get("flow_track.refresh.message", "Data refreshed from ProgTrack")
        )


class UndoManager:
    """Simple undo/redo manager."""
    
    MAX_UNDO_STEPS = 10
    
    def __init__(self):
        self.undo_stack = []
        self.redo_stack = []
    
    def push(self, action):
        """Add action to undo stack."""
        self.undo_stack.append(action)
        if len(self.undo_stack) > self.MAX_UNDO_STEPS:
            self.undo_stack.pop(0)
        self.redo_stack.clear()
    
    def undo(self):
        """Undo last action."""
        if not self.undo_stack:
            return None
        action = self.undo_stack.pop()
        self.redo_stack.append(action)
        return action
    
    def redo(self):
        """Redo last undone action."""
        if not self.redo_stack:
            return None
        action = self.redo_stack.pop()
        self.undo_stack.append(action)
        return action
    
    def can_undo(self):
        """Check if undo is available."""
        return len(self.undo_stack) > 0
    
    def can_redo(self):
        """Check if redo is available."""
        return len(self.redo_stack) > 0
