# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright © 2026 Dimitri L. Lindenwald and Deutsches Primatenzentrum GmbH
# Part of: ProgTrack 0.1.0 RC
# Required ProgTrack version: see plugin manifest.
# Required Launcher version: 0.1.0 RC or newer.
# Module: Medi Track medical-history widget and plugin logic.

from __future__ import annotations

import json
import logging
import os
import shutil
import tempfile
import uuid
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from PyQt6.QtCore import Qt, QDate, QSize, QStandardPaths, QTimer
from PyQt6.QtGui import QIcon, QPixmap
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QApplication,
)
from Plugins.core.animal_identity import animal_base_name
from Plugins.core.animal_status import status_summary_with_death_priority
from Plugins.core.lifecycle_events import ever_in_experiment
from Plugins.core.platform_helpers import open_local_path
from Plugins.core.ui_icons import apply_icon

logger = logging.getLogger(__name__)

PLUGIN_DIR = Path(__file__).parent
ICONS_DIR = Path(__file__).parent.parent.parent / "icons"
CONDITIONS_SUBDIR = "lang"
MEDI_DOCS_SUBDIR = "medi_track"

_IMAGE_EXTS  = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.svg', '.webp', '.ico', '.heic'}
_TEXT_EXTS   = {'.doc', '.docx', '.txt', '.odt', '.rtf'}
_TABLE_EXTS  = {'.csv', '.xls', '.xlsx', '.ods'}
_PDF_EXTS    = {'.pdf'}


class UnifiedExportDialog(QDialog):
    """Shared export options dialog for Reports and Medi Track.

    ``tab_current`` deliberately renders only an identity label.  The
    ``file_multi`` mode adds prefix search and multi-selection.  Keeping those
    contexts explicit prevents a tab-local export from accidentally exposing
    other animals while still giving the File menu its batch workflow.
    """

    MODE_TAB_CURRENT = "tab_current"
    MODE_FILE_MULTI = "file_multi"

    def __init__(
        self,
        parent: Optional[QWidget],
        *,
        title: str,
        candidates: Iterable[Tuple[str, ...]],
        mode: str,
        current_animal: Optional[str] = None,
        formats: Iterable[str] = ("PDF", "XLSX"),
        messages: Optional[Dict[str, Any]] = None,
        show_signatures: bool = False,
        show_documents: bool = False,
        show_date_range: bool = False,
        initial_from: Optional[date] = None,
        initial_to: Optional[date] = None,
        project_options: Optional[Iterable[str]] = None,
    ) -> None:
        super().__init__(parent)
        if mode not in (self.MODE_TAB_CURRENT, self.MODE_FILE_MULTI):
            raise ValueError(f"Unsupported export dialog mode: {mode}")

        self.messages = messages or {}
        self.mode = mode
        self._candidates = [
            (
                str(row[0]),
                str(row[1]),
                str(row[2]) if len(row) > 2 else "",
                str(row[3]) if len(row) > 3 else "",
                str(row[4]) if len(row) > 4 else "",
            )
            for row in candidates
        ]
        self._candidate_ids = {row[0] for row in self._candidates}
        self._current_animal = (
            str(current_animal)
            if current_animal is not None and str(current_animal) in self._candidate_ids
            else None
        )
        self._animal_list: Optional[QListWidget] = None
        self._search_edit: Optional[QLineEdit] = None

        self.setObjectName("unifiedExportDialog")
        self.setProperty("exportContext", mode)
        self.setWindowModality(Qt.WindowModality.WindowModal)
        self.setWindowTitle(title)
        self.setStyleSheet(
            "QDialog#unifiedExportDialog QGroupBox { font-weight: bold; margin-top: 8px; }"
            "QDialog#unifiedExportDialog QGroupBox::title { subcontrol-origin: margin; left: 8px; }"
            "QDialog#unifiedExportDialog QListWidget { min-height: 180px; }"
        )

        self._root_layout = QVBoxLayout(self)
        self._root_layout.setContentsMargins(12, 12, 12, 12)
        self._root_layout.setSpacing(8)

        animal_group = QGroupBox(
            _msg(
                self.messages,
                "export.animals" if mode == self.MODE_FILE_MULTI else "export.animal",
                "Animals" if mode == self.MODE_FILE_MULTI else "Animal",
            )
        )
        animal_layout = QVBoxLayout(animal_group)

        if mode == self.MODE_TAB_CURRENT:
            current_label = next(
                (row[1] for row in self._candidates if row[0] == self._current_animal),
                "",
            )
            self.current_animal_label = QLabel(current_label)
            self.current_animal_label.setObjectName("exportCurrentAnimal")
            self.current_animal_label.setWordWrap(True)
            animal_layout.addWidget(self.current_animal_label)
        else:
            filter_row = QHBoxLayout()
            self.species_filter = QComboBox()
            self.species_filter.setObjectName("exportSpeciesFilter")
            self.species_filter.addItem(
                _msg(self.messages, "filter.all_species", "All species"), ""
            )
            for species in sorted(
                {row[3] for row in self._candidates if row[3]},
                key=str.casefold,
            ):
                self.species_filter.addItem(species, species)
            filter_row.addWidget(self.species_filter)

            self.project_filter = QComboBox()
            self.project_filter.setObjectName("exportProjectFilter")
            self.project_filter.addItem(
                _msg(self.messages, "filter.all_projects", "All projects"), ""
            )
            allowed_projects = (
                {str(value) for value in project_options if str(value).strip()}
                if project_options is not None
                else {row[4] for row in self._candidates if row[4]}
            )
            for project in sorted(allowed_projects, key=str.casefold):
                self.project_filter.addItem(project, project)
            filter_row.addWidget(self.project_filter)
            animal_layout.addLayout(filter_row)

            self._search_edit = QLineEdit()
            self._search_edit.setObjectName("exportAnimalSearch")
            self._search_edit.setPlaceholderText(
                _msg(
                    self.messages,
                    "medi_track.export.search",
                    "Search animals by name prefix",
                )
            )
            animal_layout.addWidget(self._search_edit)

            self.select_all_checkbox = QCheckBox(
                _msg(self.messages, "checkbox.select_all", "Select all visible")
            )
            self.select_all_checkbox.setObjectName("exportSelectAll")
            animal_layout.addWidget(self.select_all_checkbox)

            self._animal_list = QListWidget()
            self._animal_list.setObjectName("exportAnimalList")
            self._animal_list.setSelectionMode(
                QAbstractItemView.SelectionMode.ExtendedSelection
            )
            previous_group = None
            for animal_id, label, group, species, project in self._candidates:
                if group and group != previous_group:
                    header = QListWidgetItem(group)
                    header.setData(Qt.ItemDataRole.UserRole, None)
                    header.setData(Qt.ItemDataRole.UserRole + 1, group)
                    header.setFlags(Qt.ItemFlag.ItemIsEnabled)
                    font = header.font()
                    font.setBold(True)
                    header.setFont(font)
                    self._animal_list.addItem(header)
                    previous_group = group
                item = QListWidgetItem(label)
                item.setData(Qt.ItemDataRole.UserRole, animal_id)
                item.setData(Qt.ItemDataRole.UserRole + 1, group)
                item.setData(Qt.ItemDataRole.UserRole + 2, species)
                item.setData(Qt.ItemDataRole.UserRole + 3, project)
                self._animal_list.addItem(item)
                if animal_id == self._current_animal:
                    item.setSelected(True)
            self._search_edit.textChanged.connect(self._apply_prefix_filter)
            self.species_filter.currentIndexChanged.connect(self._apply_filters)
            self.project_filter.currentIndexChanged.connect(self._apply_filters)
            self.select_all_checkbox.toggled.connect(self._select_all_visible)
            animal_layout.addWidget(self._animal_list)

        self._root_layout.addWidget(animal_group)

        options_group = QGroupBox(
            _msg(self.messages, "export.options", "Export options")
        )
        self.options_layout = QFormLayout(options_group)
        self.format_combo = QComboBox()
        self.format_combo.setObjectName("exportFormat")
        self.format_combo.addItems([str(fmt).upper() for fmt in formats])
        self.options_layout.addRow(
            _msg(self.messages, "export.format", "Format"), self.format_combo
        )

        self.signature_checkbox: Optional[QCheckBox] = None
        if show_signatures:
            self.signature_checkbox = QCheckBox(
                _msg(
                    self.messages,
                    "medi_track.export.include_signatures",
                    "Include signatures",
                )
            )
            self.signature_checkbox.setObjectName("exportIncludeSignatures")
            self.signature_checkbox.setChecked(True)
            self.options_layout.addRow("", self.signature_checkbox)

        self.documents_checkbox: Optional[QCheckBox] = None
        if show_documents:
            self.documents_checkbox = QCheckBox(
                _msg(
                    self.messages,
                    "medi_track.export.include_documents",
                    "Copy uploaded documents (XLSX)",
                )
            )
            self.documents_checkbox.setObjectName("exportIncludeDocuments")
            self.documents_checkbox.setChecked(True)
            self.options_layout.addRow("", self.documents_checkbox)
            self.format_combo.currentTextChanged.connect(
                self._sync_document_option
            )
            self._sync_document_option(self.format_combo.currentText())

        self._root_layout.addWidget(options_group)

        self.complete_checkbox: Optional[QCheckBox] = None
        self.from_date: Optional[QDateEdit] = None
        self.to_date: Optional[QDateEdit] = None
        if show_date_range:
            today = date.today()
            initial_to = initial_to or today
            initial_from = initial_from or (today - timedelta(days=31))
            date_group = QGroupBox(
                _msg(self.messages, "export.date_range", "Date range")
            )
            date_layout = QFormLayout(date_group)
            self.complete_checkbox = QCheckBox(
                _msg(self.messages, "export.complete", "Complete")
            )
            self.complete_checkbox.setObjectName("exportComplete")
            date_layout.addRow("", self.complete_checkbox)
            self.from_date = QDateEdit()
            self.from_date.setObjectName("exportFromDate")
            self.from_date.setCalendarPopup(True)
            self.from_date.setDate(QDate(initial_from.year, initial_from.month, initial_from.day))
            self.to_date = QDateEdit()
            self.to_date.setObjectName("exportToDate")
            self.to_date.setCalendarPopup(True)
            self.to_date.setDate(QDate(initial_to.year, initial_to.month, initial_to.day))
            date_layout.addRow(_msg(self.messages, "form.label.from", "From"), self.from_date)
            date_layout.addRow(_msg(self.messages, "form.label.to", "To"), self.to_date)
            self.complete_checkbox.toggled.connect(self._sync_date_range)
            self._root_layout.addWidget(date_group)
        self._extra_options_index = self._root_layout.count()

        self.button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
        )
        self.button_box.setObjectName("exportDialogButtons")
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self._root_layout.addWidget(self.button_box)
        self.resize(480, 520 if mode == self.MODE_FILE_MULTI else 280)

    def add_options_widget(self, widget: QWidget) -> None:
        """Insert context-specific controls before the common button row."""
        self._root_layout.insertWidget(self._root_layout.count() - 1, widget)

    def _apply_prefix_filter(self, text: str) -> None:
        self._apply_filters()

    def _apply_filters(self, *_args) -> None:
        if self._animal_list is None:
            return
        prefix = str(self._search_edit.text() if self._search_edit else "").strip().casefold()
        species = str(self.species_filter.currentData() or "")
        project = str(self.project_filter.currentData() or "")
        visible_groups = set()
        for index in range(self._animal_list.count()):
            item = self._animal_list.item(index)
            if item.data(Qt.ItemDataRole.UserRole) is None:
                continue
            visible = (
                (not prefix or item.text().casefold().startswith(prefix))
                and (not species or item.data(Qt.ItemDataRole.UserRole + 2) == species)
                and (not project or item.data(Qt.ItemDataRole.UserRole + 3) == project)
            )
            item.setHidden(not visible)
            if visible:
                visible_groups.add(str(item.data(Qt.ItemDataRole.UserRole + 1) or ""))
        for index in range(self._animal_list.count()):
            item = self._animal_list.item(index)
            if item.data(Qt.ItemDataRole.UserRole) is None:
                item.setHidden(str(item.data(Qt.ItemDataRole.UserRole + 1) or "") not in visible_groups)

    def _select_all_visible(self, checked: bool) -> None:
        if self._animal_list is None:
            return
        for index in range(self._animal_list.count()):
            item = self._animal_list.item(index)
            if item.data(Qt.ItemDataRole.UserRole) is not None and not item.isHidden():
                item.setSelected(bool(checked))

    def _sync_document_option(self, format_text: str) -> None:
        if self.documents_checkbox is not None:
            self.documents_checkbox.setEnabled(str(format_text).upper() == "XLSX")

    def _sync_date_range(self, complete: bool) -> None:
        if self.from_date is not None:
            self.from_date.setEnabled(not complete)
        if self.to_date is not None:
            self.to_date.setEnabled(not complete)

    def selected_date_range(self) -> Tuple[Optional[date], Optional[date]]:
        if self.complete_checkbox is not None and self.complete_checkbox.isChecked():
            return None, None
        if self.from_date is None or self.to_date is None:
            return None, None
        return self.from_date.date().toPyDate(), self.to_date.date().toPyDate()

    def selected_animal_ids(self) -> List[str]:
        if self.mode == self.MODE_TAB_CURRENT:
            return [self._current_animal] if self._current_animal else []
        if self._animal_list is None:
            return []
        return [
            str(item.data(Qt.ItemDataRole.UserRole))
            for item in self._animal_list.selectedItems()
        ]

    def selected_format(self) -> str:
        return self.format_combo.currentText().strip().lower()

    def include_signatures(self) -> bool:
        return bool(
            self.signature_checkbox is None or self.signature_checkbox.isChecked()
        )

    def include_documents(self) -> bool:
        return bool(
            self.documents_checkbox is not None
            and self.documents_checkbox.isEnabled()
            and self.documents_checkbox.isChecked()
        )


def _safe_name(name: str) -> str:
    """Make animal name safe for use as a folder name."""
    safe = str(name).strip()
    for ch in r'/\:*?"<>|':
        safe = safe.replace(ch, '_')
    return safe or 'unknown'


def _entry_date_in_range(
    value: object,
    date_from: Optional[date],
    date_to: Optional[date],
) -> bool:
    if date_from is None and date_to is None:
        return True
    parsed = None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(str(value or ""), fmt).date()
            break
        except ValueError:
            continue
    if parsed is None:
        return False
    return not ((date_from and parsed < date_from) or (date_to and parsed > date_to))


def _format_neutral_export_label(text: str, fallback: str = "Export") -> str:
    """Remove a legacy PDF-only qualifier while retaining localized wording."""
    label = str(text or "").strip()
    label = label.replace("(.pdf)", "").replace("(.PDF)", "")
    label = label.replace(".pdf", "").replace(".PDF", "")
    label = " ".join(part for part in label.split() if part.casefold() != "pdf")
    return label.strip() or fallback


def _path_identity(path: Path) -> str:
    return os.path.normcase(os.path.abspath(str(path)))


def _document_paths_for_animal(
    animal_name: str,
    store: "MediStore",
    docs_root: Optional[Path] = None,
) -> List[Path]:
    """Return all filesystem documents known for an animal."""
    seen = set()
    paths: List[Path] = []

    def add(path_like: object) -> None:
        if not path_like:
            return
        path = Path(str(path_like))
        try:
            if not path.is_file() or path.name.startswith(".") or path.name.lower() == "thumbs.db":
                return
        except OSError:
            return
        ident = _path_identity(path)
        if ident in seen:
            return
        seen.add(ident)
        paths.append(path)

    backend = getattr(store, "backend", None)
    if backend is not None:
        for record in backend.documents.list_for_owner(
            "animal-medical", animal_name
        ):
            add(backend.documents.payload_path(record))
    elif docs_root is not None:
        folder = Path(docs_root) / _safe_name(animal_name)
        if folder.exists():
            try:
                for path in folder.iterdir():
                    add(path)
            except OSError:
                pass

    try:
        json_docs = store.get_documents(animal_name)
    except Exception:
        json_docs = []
    for doc in json_docs:
        if isinstance(doc, dict):
            add(doc.get("path"))

    return sorted(paths, key=lambda p: p.name.lower())


def _unique_export_destination(destination_dir: Path, source: Path) -> Path:
    candidate = destination_dir / source.name
    counter = 1
    while candidate.exists() and _path_identity(candidate) != _path_identity(source):
        candidate = destination_dir / f"{source.stem}_{counter}{source.suffix}"
        counter += 1
    return candidate


def _copy_document_files_to_directory(
    sources: Iterable[Path],
    destination_dir: Path,
) -> int:
    """Copy document files to a directory without overwriting existing files."""
    destination_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for source in sources:
        if not source.is_file():
            continue
        dest = _unique_export_destination(destination_dir, source)
        if _path_identity(dest) == _path_identity(source):
            continue
        shutil.copy2(str(source), str(dest))
        copied += 1
    return copied


def _display_animal_name(animal_key: str, record: Optional[Dict[str, Any]] = None) -> str:
    display_name = animal_base_name(animal_key, record)
    fallback_name = animal_base_name(animal_key)
    if display_name == str(animal_key or '').strip() or " | " in display_name:
        return fallback_name
    return display_name or fallback_name


def _icon_for_ext(ext: str) -> QIcon:
    """Return a QIcon for the given file extension using icons/*.png files."""
    ext = ext.lower()
    if ext in _IMAGE_EXTS:
        icon_name = 'file_img.png'
    elif ext in _TEXT_EXTS:
        icon_name = 'file_text.png'
    elif ext in _TABLE_EXTS:
        icon_name = 'file_csv.png'
    elif ext in _PDF_EXTS:
        icon_name = 'file_pdf.png'
    else:
        icon_name = 'file_text.png'
    path = ICONS_DIR / icon_name
    return QIcon(str(path)) if path.exists() else QIcon()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _msg(messages: Dict[str, Any], key: str, fallback: str) -> str:
    return messages.get(key, fallback) if isinstance(messages, dict) else fallback


def _today_iso() -> str:
    return date.today().isoformat()


def _snake(label: str) -> str:
    """Convert a display label to a safe key fragment."""
    return (
        label.lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace("'", "")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "_")
        .replace(".", "_")
        .replace(",", "")
        .replace("&", "and")
    )


# ─────────────────────────────────────────────────────────────────────────────
# MediStore — persistence for backend medical-history records
# ─────────────────────────────────────────────────────────────────────────────

class MediStore:
    """Read/write medical-history records in the configured backend."""

    def __init__(self, backend: Any) -> None:
        from Plugins.core.backend_store import BackendJsonStore
        self.backend = backend
        self._backend_store = BackendJsonStore(backend, "medical", "history")
        self._data: Optional[Dict[str, Any]] = None

    # ── internal ──

    def _default(self) -> Dict[str, Any]:
        return {"version": "1.7", "animals": {}}

    def _coerce(self, raw: Any) -> Dict[str, Any]:
        if not isinstance(raw, dict):
            raw = self._default()
        if not isinstance(raw.get("animals"), dict):
            raw["animals"] = {}
        return raw

    # ── public API ──

    def load(self) -> Dict[str, Any]:
        if self._data is not None:
            return self._data
        raw = self._backend_store.load(self._default())
        self._data = self._coerce(raw)
        return self._data

    def save(self) -> None:
        data = self.load()
        self._backend_store.save(data)

    def _block(self, name: str) -> Dict[str, Any]:
        data = self.load()
        key = str(name).strip()
        if key not in data["animals"]:
            data["animals"][key] = {
                "ipid": key,
                "name": animal_base_name(key),
                "animal_id": "",
                "entries": [],
                "documents": [],
            }
        blk = data["animals"][key]
        blk.setdefault("ipid", key)
        blk.setdefault("name", animal_base_name(key))
        if not isinstance(blk.get("entries"), list):
            blk["entries"] = []
        if not isinstance(blk.get("documents"), list):
            blk["documents"] = []
        return blk

    def get_entries(self, name: str) -> List[Dict[str, Any]]:
        return list(self._block(name)["entries"])

    def get_documents(self, name: str) -> List[Dict[str, Any]]:
        return list(self._block(name)["documents"])

    def add_entry(self, name: str, entry: Dict[str, Any]) -> None:
        self._block(name)["entries"].append(entry)
        self.save()

    def add_document(self, name: str, doc: Dict[str, Any]) -> None:
        self._block(name)["documents"].append(doc)
        self.save()

    def get_active_issues(self, name: str, status_type: str) -> List[Dict[str, Any]]:
        """Return issues of *status_type* that have been started but not resolved."""
        entries = self.get_entries(name)
        active: Dict[str, Dict[str, Any]] = {}
        for e in entries:
            if e.get("status_type") != status_type:
                continue
            iid = str(e.get("issue_id", ""))
            etype = e.get("entry_type", "")
            if etype == "issue_start":
                active[iid] = e
            elif etype == "issue_resolution":
                active.pop(iid, None)
        return list(active.values())

    def has_any_of_type(self, name: str, status_type: str) -> bool:
        """True if any entry of *status_type* exists (ever sick / ever abnormal)."""
        return any(
            e.get("status_type") == status_type for e in self.get_entries(name)
        )


# ─────────────────────────────────────────────────────────────────────────────
# ConditionLoader — loads hierarchical condition taxonomy from JSON
# ─────────────────────────────────────────────────────────────────────────────

class ConditionLoader:
    """Load and cache condition taxonomy from lang/conditions_*.json."""

    _CACHE: Dict[str, List[Dict[str, Any]]] = {}

    @classmethod
    def _load_file(cls, lang_code: str) -> List[Dict[str, Any]]:
        path = PLUGIN_DIR / CONDITIONS_SUBDIR / f"conditions_{lang_code}.json"
        try:
            with open(str(path), "r", encoding="utf-8") as fh:
                raw = json.load(fh)
            if isinstance(raw, list):
                return raw
        except (FileNotFoundError, json.JSONDecodeError):
            pass
        return []

    @classmethod
    def get(cls, lang_code: str = "en") -> List[Dict[str, Any]]:
        if lang_code in cls._CACHE:
            return cls._CACHE[lang_code]
        entries = cls._load_file(lang_code)
        if not entries and lang_code != "en":
            entries = cls._load_file("en")
        cls._CACHE[lang_code] = entries
        return entries

    @classmethod
    def label_for_key(cls, key: str, lang_code: str = "en") -> str:
        entries = cls.get(lang_code)
        for e in entries:
            if e.get("key") == key:
                return str(e.get("label", key))
        # fallback to English
        if lang_code != "en":
            entries_en = cls.get("en")
            for e in entries_en:
                if e.get("key") == key:
                    return str(e.get("label", key))
        return key

    @classmethod
    def get_hierarchy_label(cls, key: str, lang_code: str = "en") -> str:
        """Return full ancestor path, e.g. 'Musculoskeletal → Bones → Fracture'."""
        if not key:
            return ""
        entries = cls.get(lang_code)
        if not entries and lang_code != "en":
            entries = cls.get("en")
        key_map = {str(e.get("key", "")): e for e in entries}
        parts: List[str] = []
        current = key
        visited: set = set()
        while current and current not in visited:
            visited.add(current)
            entry = key_map.get(current)
            if not entry:
                parts.insert(0, current)
                break
            parts.insert(0, str(entry.get("label", current)))
            current = entry.get("parent") or ""
        return " \u2192 ".join(parts) if parts else key

    @classmethod
    def compact_hierarchy_label(cls, key: str, lang_code: str = "en") -> str:
        """Hierarchy minus root level, leaf in parentheses.
        E.g. 'General \u2192 Dehydration \u2192 Severe' \u2192 'Dehydration (Severe)'.
        """
        if not key:
            return ""
        entries = cls.get(lang_code)
        if not entries and lang_code != "en":
            entries = cls.get("en")
        key_map = {str(e.get("key", "")): e for e in entries}
        parts: List[str] = []
        current = key
        visited: set = set()
        while current and current not in visited:
            visited.add(current)
            entry = key_map.get(current)
            if not entry:
                parts.insert(0, current)
                break
            parts.insert(0, str(entry.get("label", current)))
            current = entry.get("parent") or ""
        if not parts:
            return key
        if len(parts) > 1:
            parts = parts[1:]  # strip root
        if len(parts) == 1:
            return parts[0]
        return " \u2192 ".join(parts[:-1]) + f" ({parts[-1]})"


# ─────────────────────────────────────────────────────────────────────────────
# ConditionSelector — hierarchical tree UI
# ─────────────────────────────────────────────────────────────────────────────

class ConditionSelector(QWidget):
    """Tree widget for selecting a medical condition."""

    def __init__(self, parent: Optional[QWidget], messages: Dict[str, Any], lang_code: str = "en") -> None:
        super().__init__(parent)
        self._lang = lang_code
        self.messages = messages
        self._selected_key: Optional[str] = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self._search = QLineEdit()
        self._search.setPlaceholderText(_msg(messages, "medi_track.condition.search_placeholder", "Search conditions…"))
        self._search.textChanged.connect(self._apply_filter)
        layout.addWidget(self._search)

        self._tree = QTreeWidget()
        self._tree.setHeaderHidden(True)
        self._tree.itemClicked.connect(self._on_clicked)
        layout.addWidget(self._tree)

        self._populate()

    def _populate(self) -> None:
        self._tree.clear()
        entries = ConditionLoader.get(self._lang)
        items: Dict[str, QTreeWidgetItem] = {}
        for entry in entries:
            key = str(entry.get("key", ""))
            label = str(entry.get("label", key))
            parent_key = entry.get("parent")
            item = QTreeWidgetItem([label])
            item.setData(0, Qt.ItemDataRole.UserRole, key)
            items[key] = item
            if parent_key and parent_key in items:
                items[parent_key].addChild(item)
            else:
                self._tree.addTopLevelItem(item)
        self._tree.expandAll()

    def _apply_filter(self, text: str) -> None:
        text = text.lower().strip()

        def _set_visible(item: QTreeWidgetItem, visible: bool) -> None:
            item.setHidden(not visible)
            for i in range(item.childCount()):
                item.child(i).setHidden(not visible)

        def _matches(item: QTreeWidgetItem) -> bool:
            if not text:
                return True
            label = item.text(0).lower()
            if text in label:
                return True
            for i in range(item.childCount()):
                if _matches(item.child(i)):
                    return True
            return False

        def _apply(item: QTreeWidgetItem) -> bool:
            visible = _matches(item)
            for i in range(item.childCount()):
                child_vis = _apply(item.child(i))
                if child_vis:
                    visible = True
            item.setHidden(not visible)
            return visible

        root = self._tree.invisibleRootItem()
        for i in range(root.childCount()):
            _apply(root.child(i))

    def _on_clicked(self, item: QTreeWidgetItem, _col: int) -> None:
        self._selected_key = item.data(0, Qt.ItemDataRole.UserRole)

    def selected_key(self) -> Optional[str]:
        return self._selected_key

    def selected_label(self) -> str:
        item = self._tree.currentItem()
        return item.text(0) if item else ""


# ─────────────────────────────────────────────────────────────────────────────
# StatusManagementDialog — clicked on active sick/abnormal checkbox
# ─────────────────────────────────────────────────────────────────────────────

class StatusManagementDialog(QDialog):
    """
    Shown when the user clicks on an already-checked Sick or Abnormal checkbox.

    Modes:
    - Add   (medi_track.status_enable) : add a new issue_start for same status
    - Resolve (medi_track.status_manage) : resolve one active issue
    """

    MODE_ADD = "add"
    MODE_RESOLVE = "resolve"

    def __init__(
        self,
        parent: Optional[QWidget],
        messages: Dict[str, Any],
        animal_name: str,
        status_type: str,
        active_issues: List[Dict[str, Any]],
        lang_code: str = "en",
        can_add: bool = True,
        can_resolve: bool = True,
        prefill_signature: str = "",
        readonly_signature: bool = False,
        display_name: str = "",
    ) -> None:
        super().__init__(parent)
        self.messages = messages
        self._lang = lang_code
        self._active_issues = active_issues
        self._result_mode: Optional[str] = None
        self._result_note: str = ""
        self._result_signature: str = ""
        self._result_issue_id: Optional[str] = None
        self._result_condition_key: Optional[str] = None
        self._result_condition_label: str = ""

        title = _msg(
            messages,
            "medi_track.dialog.status_manage.title",
            "Manage Status — {animal}",
        ).replace("{animal}", display_name if display_name else animal_name)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(500)
        self.resize(560, 580)

        layout = QVBoxLayout(self)

        # Date row (non-editable)
        form = QFormLayout()
        date_lbl = QLabel(date.today().strftime("%d.%m.%Y"))
        form.addRow(_msg(messages, "medi_track.dialog.label.date", "Date:"), date_lbl)

        # Mode radios
        mode_lbl = _msg(messages, "medi_track.dialog.label.mode", "Mode:")
        self._rb_add = QRadioButton(
            _msg(messages, "medi_track.dialog.mode.add", "Add new issue"))
        self._rb_resolve = QRadioButton(
            _msg(messages, "medi_track.dialog.mode.resolve", "Resolve issue"))
        self._rb_add.setEnabled(can_add)
        self._rb_resolve.setEnabled(can_resolve and bool(active_issues))
        if can_add:
            self._rb_add.setChecked(True)
        elif can_resolve and active_issues:
            self._rb_resolve.setChecked(True)
        self._rb_add.toggled.connect(self._on_mode_changed)

        mode_row = QHBoxLayout()
        mode_row.addWidget(self._rb_add)
        mode_row.addWidget(self._rb_resolve)
        form.addRow(mode_lbl, mode_row)

        # Note field
        self._note_edit = QLineEdit()
        form.addRow(_msg(messages, "medi_track.dialog.label.note", "Note:"), self._note_edit)

        # Signature field
        self._sig_edit = QLineEdit(prefill_signature)
        if readonly_signature:
            self._sig_edit.setReadOnly(True)
            self._sig_edit.setStyleSheet("QLineEdit { background: #f0f0f0; color: #666; }")
        form.addRow(
            _msg(messages, "medi_track.dialog.label.signature", "Signature:"),
            self._sig_edit)

        layout.addLayout(form)

        # Active issues list (for Resolve mode)
        self._issue_group = QGroupBox(
            _msg(messages, "medi_track.dialog.label.active_issue", "Active issues to resolve:"))
        issue_layout = QVBoxLayout(self._issue_group)
        self._issue_list = QListWidget()
        self._issue_list.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)
        self._issue_list.setMaximumHeight(130)
        for iss in active_issues:
            cond_key = str(iss.get("condition_key", ""))
            cond_label = (ConditionLoader.get_hierarchy_label(cond_key, lang_code)
                          if cond_key else
                          str(iss.get("condition_label_snapshot", "")))
            iss_date = str(iss.get("date", ""))
            display = f"{iss_date}  {cond_label}".strip() or f"Issue {iss.get('issue_id','')[:8]}"
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, iss.get("issue_id", ""))
            self._issue_list.addItem(item)
        if self._issue_list.count():
            self._issue_list.setCurrentRow(0)
        self._issue_list.currentItemChanged.connect(self._on_issue_selection_changed)
        issue_layout.addWidget(self._issue_list)

        # Read-only label showing condition of the selected issue
        self._resolve_cond_lbl = QLabel()
        self._resolve_cond_lbl.setWordWrap(True)
        self._resolve_cond_lbl.setStyleSheet("color: #555; font-style: italic;")
        issue_layout.addWidget(self._resolve_cond_lbl)
        layout.addWidget(self._issue_group)

        # Condition selector (for Add mode)
        self._cond_group = QGroupBox(
            _msg(messages, "medi_track.dialog.label.document", "Condition (optional):"))
        cond_layout = QVBoxLayout(self._cond_group)
        self._condition_selector = ConditionSelector(self._cond_group, messages, lang_code)
        cond_layout.addWidget(self._condition_selector)
        layout.addWidget(self._cond_group, 1)

        # Buttons
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._accept)
        btns.rejected.connect(self.reject)
        # Set explicit translated text
        btns.button(QDialogButtonBox.StandardButton.Ok).setText(
            _msg(messages, "button.ok", "OK"))
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText(
            _msg(messages, "button.cancel", "Cancel"))
        layout.addWidget(btns)

        self._on_mode_changed()
        self._on_issue_selection_changed(
            self._issue_list.currentItem(), None)

    def _on_issue_selection_changed(
        self,
        current: Optional['QListWidgetItem'],
        _prev: Optional['QListWidgetItem'],
    ) -> None:
        """Update the condition label when a different issue is selected."""
        if current is None:
            self._resolve_cond_lbl.setText("")
            return
        issue_id = current.data(Qt.ItemDataRole.UserRole)
        for iss in self._active_issues:
            if iss.get("issue_id") == issue_id:
                ckey = str(iss.get("condition_key", ""))
                if ckey:
                    label = ConditionLoader.get_hierarchy_label(ckey, self._lang)
                else:
                    label = str(iss.get("condition_label_snapshot", ""))
                self._resolve_cond_lbl.setText(
                    f"{_msg(self.messages, 'medi_track.dialog.label.condition_free', 'Condition')}: {label}"
                    if label else "")
                return
        self._resolve_cond_lbl.setText("")

    def _on_mode_changed(self) -> None:
        is_add = self._rb_add.isChecked()
        self._cond_group.setVisible(is_add)
        self._issue_group.setVisible(not is_add)

    def _accept(self) -> None:
        if self._rb_add.isChecked():
            self._result_mode = self.MODE_ADD
            self._result_condition_key = self._condition_selector.selected_key()
            self._result_condition_label = self._condition_selector.selected_label()
        else:
            self._result_mode = self.MODE_RESOLVE
            item = self._issue_list.currentItem()
            self._result_issue_id = item.data(Qt.ItemDataRole.UserRole) if item else None
        self._result_note = self._note_edit.text().strip()
        self._result_signature = self._sig_edit.text().strip()
        self.accept()

    # ── result accessors ──

    def result_mode(self) -> Optional[str]:
        return self._result_mode

    def result_note(self) -> str:
        return self._result_note

    def result_issue_id(self) -> Optional[str]:
        return self._result_issue_id

    def result_condition_key(self) -> Optional[str]:
        return self._result_condition_key

    def result_condition_label(self) -> str:
        return self._result_condition_label

    def result_signature(self) -> str:
        return self._result_signature


# ─────────────────────────────────────────────────────────────────────────────
# MediTrackWidget — main tab widget
# ─────────────────────────────────────────────────────────────────────────────

class MediTrackWidget(QWidget):
    """
    The main Medi_Track tab.

    Layout:
    ┌──────────────────────────────────────────────────────┐
    │  [All] [Sick] [Ever Sick] [Abnormal] [Ever Abnormal]  │  ← filter row
    ├──────────────────────────────────────────────────────┤
    │  Header: name, ID, project, status, birth, diagnoses  │
    ├──────────────────────────────────────────────────────┤
    │  Medical history table (read-only)                    │
    ├──────────────────────────────────────────────────────┤
    │  ▶ Documents (collapsible)                           │
    └──────────────────────────────────────────────────────┘
    """

    FILTER_ALL = "all"
    FILTER_SICK = "sick"
    FILTER_EVER_SICK = "ever_sick"
    FILTER_ABNORMAL = "abnormal"
    FILTER_EVER_ABNORMAL = "ever_abnormal"
    FILTER_IN_EXPERIMENT = "in_experiment"
    FILTER_EVER_EXPERIMENT = "ever_experiment"

    def __init__(
        self,
        app: Any,
        store: MediStore,
        messages: Dict[str, Any],
        lang_code: str = "en",
        data_dir: str = "",
    ) -> None:
        super().__init__()
        self.app = app
        self.store = store
        self.messages = messages
        self._lang = lang_code
        self._data_dir = data_dir
        self._current_animal: Optional[str] = None
        self._active_filter = self.FILTER_ALL

        self._build_ui()

    # ── UI construction ──

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Filter button row — always visible regardless of stack page
        filter_row = QHBoxLayout()
        filter_row.setContentsMargins(8, 4, 8, 4)
        self._filter_btns: Dict[str, QPushButton] = {}
        _filters = [
            (self.FILTER_ALL,           "medi_track.filter.all",           "All"),
            (self.FILTER_SICK,          "medi_track.filter.current_sick",  "Currently Sick"),
            (self.FILTER_EVER_SICK,     "medi_track.filter.ever_sick",     "Ever Sick"),
            (self.FILTER_ABNORMAL,      "medi_track.filter.current_abnormal", "Currently Abnormal"),
            (self.FILTER_EVER_ABNORMAL, "medi_track.filter.ever_abnormal", "Ever Abnormal"),
            (
                self.FILTER_IN_EXPERIMENT,
                "medi_track.filter.current_experiment",
                "Currently in experiment",
            ),
            (
                self.FILTER_EVER_EXPERIMENT,
                "medi_track.filter.ever_experiment",
                "Was in experiment",
            ),
        ]
        filter_icons = {
            self.FILTER_SICK: "medi_track.filter.current_sick",
            self.FILTER_EVER_SICK: "medi_track.filter.ever_sick",
            self.FILTER_ABNORMAL: "medi_track.filter.current_abnormal",
            self.FILTER_EVER_ABNORMAL: "medi_track.filter.ever_abnormal",
            self.FILTER_IN_EXPERIMENT: "medi_track.filter.current_experiment",
            self.FILTER_EVER_EXPERIMENT: "medi_track.filter.ever_experiment",
        }
        for fkey, msg_key, fallback in _filters:
            label = _msg(self.messages, msg_key, fallback)
            btn = QPushButton(label)
            semantic_id = filter_icons.get(fkey)
            if semantic_id:
                apply_icon(btn, semantic_id, fallback=label)
                btn.setIconSize(QSize(27, 27))
            btn.setCheckable(True)
            if fkey in (self.FILTER_IN_EXPERIMENT, self.FILTER_EVER_EXPERIMENT):
                btn.setToolTip(_msg(self.messages, msg_key, fallback))
            btn.setChecked(fkey == self.FILTER_ALL)
            btn.clicked.connect(lambda checked, k=fkey: self._on_filter_clicked(k))
            self._filter_btns[fkey] = btn
            filter_row.addWidget(btn)
        filter_row.addStretch()
        root.addLayout(filter_row)

        # Stacked widget: index 0 = splash (no animal selected), index 1 = content
        self._stack = QStackedWidget()
        root.addWidget(self._stack)

        # ── Splash (index 0) ──────────────────────────────────────────────────
        splash_w = QWidget()
        sl = QVBoxLayout(splash_w)
        sl.addStretch(1)
        disc = QLabel(_msg(self.messages, "footer.rights", "ProgTrack").format(year=datetime.now().year))
        disc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sl.addWidget(disc, alignment=Qt.AlignmentFlag.AlignCenter)
        sl.addSpacing(20)
        img = QLabel()
        pix_path = Path("icons/Splash.png")
        if pix_path.exists():
            pix = QPixmap(str(pix_path))
            pix = pix.scaled(800, 800, Qt.AspectRatioMode.KeepAspectRatio,
                             Qt.TransformationMode.SmoothTransformation)
            img.setPixmap(pix)
        img.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sl.addWidget(img, alignment=Qt.AlignmentFlag.AlignCenter)
        sl.addStretch(1)
        self._stack.addWidget(splash_w)

        # ── Content (index 1) ─────────────────────────────────────────────────
        content_w = QWidget()
        cl = QVBoxLayout(content_w)
        cl.setContentsMargins(8, 4, 8, 8)
        cl.setSpacing(6)

        # Header
        self._header_group = QGroupBox(
            _msg(self.messages, "medi_track.section.animal_info", "Animal Information"))
        hdr_layout = QFormLayout(self._header_group)
        hdr_layout.setContentsMargins(8, 6, 8, 6)

        def _lbl() -> QLabel:
            l = QLabel("–")
            l.setWordWrap(True)
            return l

        self._lbl_name = _lbl()
        self._lbl_id = _lbl()
        self._lbl_id.setTextFormat(Qt.TextFormat.RichText)
        self._lbl_chip_nr = _lbl()
        self._lbl_origin = _lbl()
        self._lbl_genotype = _lbl()
        self._lbl_project = _lbl()
        self._lbl_status = _lbl()
        self._lbl_birth = _lbl()
        self._lbl_death = _lbl()
        self._lbl_diagnoses = _lbl()

        hdr_layout.addRow(_msg(self.messages, "report.header.name", "Name:"), self._lbl_name)
        hdr_layout.addRow(_msg(self.messages, "report.header.id", "ID:"), self._lbl_id)
        hdr_layout.addRow(_msg(self.messages, "reports.header.chip_nr", "Chip Nr.:"), self._lbl_chip_nr)
        hdr_layout.addRow(_msg(self.messages, "reports.header.origin", "Origin:"), self._lbl_origin)
        hdr_layout.addRow(_msg(self.messages, "reports.header.genotype", "Genotype:"), self._lbl_genotype)
        hdr_layout.addRow(_msg(self.messages, "report.header.project", "Project:"), self._lbl_project)
        hdr_layout.addRow(_msg(self.messages, "report.header.status", "Status:"), self._lbl_status)
        hdr_layout.addRow(_msg(self.messages, "report.header.birth_date", "Birth Date:"), self._lbl_birth)
        hdr_layout.addRow(_msg(self.messages, "report.header.death_date", "Death Date:"), self._lbl_death)
        hdr_layout.addRow(
            _msg(self.messages, "medi_track.header.active_diagnoses", "Active Diagnoses:"),
            self._lbl_diagnoses)

        cl.addWidget(self._header_group)

        # History table
        self._hist_group = QGroupBox(
            _msg(self.messages, "medi_track.section.history", "Medical History"))
        hist_layout = QVBoxLayout(self._hist_group)

        # Toolbar row: Add Entry + Export buttons
        hist_toolbar = QHBoxLayout()
        self._btn_add_entry = QPushButton(
            _msg(self.messages, "medi_track.btn.add_entry", "+ Add Entry"))
        self._btn_add_entry.setEnabled(False)
        self._btn_add_entry.clicked.connect(self._on_add_entry_clicked)
        hist_toolbar.addWidget(self._btn_add_entry)
        self._btn_export = QPushButton(
            _format_neutral_export_label(
                _msg(self.messages, "medi_track.btn.export", "Export")))
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._on_export_clicked)
        hist_toolbar.addWidget(self._btn_export)
        hist_toolbar.addStretch()
        hist_layout.addLayout(hist_toolbar)

        self._table = QTableWidget(0, 5)
        self._table.setHorizontalHeaderLabels([
            _msg(self.messages, "medi_track.table.date", "Date"),
            _msg(self.messages, "medi_track.table.type", "Type"),
            _msg(self.messages, "medi_track.table.condition", "Condition"),
            _msg(self.messages, "medi_track.table.details", "Details"),
            _msg(self.messages, "medi_track.table.signature", "Signature"),
        ])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(True)
        hist_layout.addWidget(self._table)
        cl.addWidget(self._hist_group, 1)

        # Documents section (collapsible)
        self._docs_toggle = QPushButton(
            _msg(self.messages, "medi_track.section.documents", "Documents"))
        self._docs_toggle.setCheckable(True)
        self._docs_toggle.setChecked(False)
        self._docs_toggle.setStyleSheet(
            "QPushButton { text-align: left; font-weight: bold; border: none; }")
        self._docs_toggle.toggled.connect(self._on_docs_toggled)
        apply_icon(self._docs_toggle, "toggle.expand", fallback="Documents")

        self._docs_widget = QWidget()
        docs_inner = QVBoxLayout(self._docs_widget)
        docs_inner.setContentsMargins(0, 0, 0, 0)

        # Toolbar: Add Document button
        docs_toolbar = QHBoxLayout()
        self._btn_add_doc = QPushButton(
            _msg(self.messages, "medi_track.btn.add_document", "+ Add Document"))
        self._btn_add_doc.setEnabled(False)
        self._btn_add_doc.clicked.connect(self._on_add_document_clicked)
        docs_toolbar.addWidget(self._btn_add_doc)
        docs_toolbar.addStretch()
        docs_inner.addLayout(docs_toolbar)

        self._docs_list = QListWidget()
        self._docs_list.setMaximumHeight(150)
        self._docs_list.setIconSize(QSize(24, 24))
        self._docs_list.itemDoubleClicked.connect(self._on_doc_item_clicked)
        docs_inner.addWidget(self._docs_list)
        self._docs_widget.setVisible(False)

        cl.addWidget(self._docs_toggle)
        cl.addWidget(self._docs_widget)

        self._stack.addWidget(content_w)
        self._stack.setCurrentIndex(0)

    def _can_docs(self) -> bool:
        fn = getattr(self.app, '_master_can', None)
        return bool(fn('medi_track.add_docs')) if callable(fn) else True

    def _can_export(self) -> bool:
        fn = getattr(self.app, '_master_can', None)
        return bool(fn('medi_track.view')) if callable(fn) else True

    def _can_upload(self) -> bool:
        fn = getattr(self.app, '_master_can', None)
        return bool(fn('medi_track.upload_document')) if callable(fn) else True

    def _set_content_visible(self, visible: bool) -> None:
        self._stack.setCurrentIndex(1 if visible else 0)
        can_docs = self._can_docs()
        can_upload = self._can_upload()
        self._btn_add_entry.setEnabled(visible and can_docs)
        self._btn_export.setEnabled(visible and self._can_export())
        self._btn_add_doc.setEnabled(visible and can_upload)

    def update_language(self, messages: Dict[str, Any], lang_code: str = "") -> None:
        """Refresh all static UI texts after a language change."""
        self.messages = messages
        if lang_code:
            self._lang = lang_code
        # Filter buttons
        _filter_map = {
            self.FILTER_ALL:           ("medi_track.filter.all",              "All"),
            self.FILTER_SICK:          ("medi_track.filter.current_sick",     "Currently Sick"),
            self.FILTER_EVER_SICK:     ("medi_track.filter.ever_sick",        "Ever Sick"),
            self.FILTER_ABNORMAL:      ("medi_track.filter.current_abnormal", "Currently Abnormal"),
            self.FILTER_EVER_ABNORMAL: ("medi_track.filter.ever_abnormal",    "Ever Abnormal"),
            self.FILTER_IN_EXPERIMENT: (
                "medi_track.filter.current_experiment",
                "Currently in experiment",
            ),
            self.FILTER_EVER_EXPERIMENT: (
                "medi_track.filter.ever_experiment",
                "Was in experiment",
            ),
        }
        for fkey, btn in self._filter_btns.items():
            if fkey in _filter_map:
                k, fb = _filter_map[fkey]
                if fkey == self.FILTER_IN_EXPERIMENT:
                    btn.setText(_msg(messages, k, fb))
                    apply_icon(btn, "medi_track.filter.current_experiment", fallback=fb)
                    btn.setToolTip(_msg(messages, k, "Currently in experiment"))
                elif fkey == self.FILTER_EVER_EXPERIMENT:
                    btn.setText(_msg(messages, k, fb))
                    apply_icon(btn, "medi_track.filter.ever_experiment", fallback=fb)
                    btn.setToolTip(_msg(messages, k, "Was in experiment"))
                else:
                    btn.setText(_msg(messages, k, fb))
                    semantic_id = {
                        self.FILTER_SICK: "medi_track.filter.current_sick",
                        self.FILTER_EVER_SICK: "medi_track.filter.ever_sick",
                        self.FILTER_ABNORMAL: "medi_track.filter.current_abnormal",
                        self.FILTER_EVER_ABNORMAL: "medi_track.filter.ever_abnormal",
                    }.get(fkey)
                    if semantic_id:
                        apply_icon(btn, semantic_id, fallback=fb)
        # Group box titles
        self._header_group.setTitle(_msg(messages, "medi_track.section.animal_info", "Animal Information"))
        self._hist_group.setTitle(_msg(messages, "medi_track.section.history", "Medical History"))
        # Form row labels
        _row_keys = [
            ("report.header.name",                  "Name:"),
            ("report.header.id",                    "ID:"),
            ("reports.header.chip_nr",              "Chip Nr.:"),
            ("reports.header.origin",               "Origin:"),
            ("reports.header.genotype",             "Genotype:"),
            ("report.header.project",               "Project:"),
            ("report.header.status",                "Status:"),
            ("report.header.birth_date",            "Birth Date:"),
            ("report.header.death_date",            "Death Date:"),
            ("medi_track.header.active_diagnoses",  "Active Diagnoses:"),
        ]
        hdr_layout = self._header_group.layout()
        if isinstance(hdr_layout, QFormLayout):
            for row_idx, (key, default) in enumerate(_row_keys):
                lbl_item = hdr_layout.itemAt(row_idx, QFormLayout.ItemRole.LabelRole)
                if lbl_item and lbl_item.widget():
                    lbl_item.widget().setText(_msg(messages, key, default))
        # Table header labels
        for col, (key, fb) in enumerate([
            ("medi_track.table.date",      "Date"),
            ("medi_track.table.type",      "Type"),
            ("medi_track.table.condition", "Condition"),
            ("medi_track.table.details",   "Details"),
            ("medi_track.table.signature", "Signature"),
        ]):
            self._table.setHorizontalHeaderItem(col, QTableWidgetItem(_msg(messages, key, fb)))
        # Buttons
        self._btn_add_entry.setText(_msg(messages, "medi_track.btn.add_entry",    "+ Add Entry"))
        self._btn_export.setText(
            _format_neutral_export_label(
                _msg(messages, "medi_track.btn.export", "Export")))
        self._btn_add_doc.setText(  _msg(messages, "medi_track.btn.add_document", "+ Add Document"))
        self._docs_toggle.setText(
            _msg(messages, "medi_track.section.documents", "Documents"))
        apply_icon(self._docs_toggle, "toggle.expand", fallback="Documents")
        # Refresh animal display with updated labels
        if self._current_animal:
            self.show_animal(self._current_animal)

    # ── add-entry dialog ──

    def _on_add_entry_clicked(self) -> None:
        if not self._current_animal:
            return
        if not self._can_docs():
            QMessageBox.warning(self,
                _msg(self.messages, 'medi_track.dialog.permission_denied.title', 'Permission Denied'),
                _msg(self.messages, 'medi_track.dialog.permission_denied.docs',
                     'You do not have permission to add medical history entries.'))
            return
        sig = self._get_signature()
        mt = getattr(self.app, 'master_track', None)
        readonly_sig = bool(mt and getattr(mt, 'is_logged_in', False))
        _app_animals = getattr(self.app, 'animals', {}) or {}
        _rec_entry = (_app_animals.get(self._current_animal) or {})
        _dname_entry = animal_base_name(self._current_animal or '', _rec_entry)
        dlg = ManualEntryDialog(
            self, self.messages, self._current_animal,
            default_signature=sig,
            readonly_signature=readonly_sig,
            lang_code=self._lang,
            display_name=_dname_entry,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        ckey = dlg.result_condition_key()
        clabel = dlg.result_condition_label() if ckey else ""
        entry = {
            "id": str(uuid.uuid4()),
            "date": dlg.result_date(),
            "entry_type": dlg.result_entry_type(),
            "status_type": "",
            "issue_id": "",
            "condition_key": ckey,
            "condition_label_snapshot": clabel,
            "condition_path_keys": [],
            "condition_path_label_snapshot": [],
            "condition_language": self._lang,
            "condition_depth": len(ckey.split(".")) - 1 if ckey else 0,
            "note": dlg.result_note(),
            "signature": dlg.result_signature(),
            "linked_document_ids": [],
        }
        try:
            self.store.add_entry(self._current_animal, entry)
        except Exception as exc:
            logger.error("add_entry failed: %s", exc)
            return
        self.show_animal(self._current_animal)

    def _get_signature(self) -> str:
        mt = getattr(self.app, 'master_track', None)
        if mt and getattr(mt, 'is_logged_in', False):
            dname = getattr(mt, 'current_display_name', None)
            return str(dname or getattr(mt, 'current_username', '') or '') or _msg(
                self.messages, 'medi_track.value.guest', '(guest)')
        return _msg(self.messages, 'medi_track.value.guest', '(guest)')

    def _export_animal_to_pdf(
        self,
        animal_name: str,
        output_path: str,
        lang: Optional[str] = None,
        include_signature: bool = True,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> None:
        """Build and write a Medi Track PDF for *animal_name* to *output_path*.

        *lang* selects the report language (e.g. 'en', 'de').  When omitted
        the widget's current language is used.
        Raises RuntimeError if QPrinter/QTextDocument is unavailable.
        """
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QTextDocument
        except ImportError as exc:
            raise RuntimeError('PDF export requires PyQt6.QtPrintSupport.') from exc

        # Load messages for the requested language
        import json as _json, os as _os
        _lang = lang or self._lang
        _lang_path = _os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)),
            '..', '..', 'lang', f'messages_{_lang}.json')
        try:
            with open(_lang_path, encoding='utf-8') as _lf:
                _messages = _json.load(_lf)
        except Exception:
            _messages = self.messages

        app_animals = self.app.animals if isinstance(
            getattr(self.app, 'animals', None), dict) else {}
        rec = app_animals.get(animal_name, {})

        def _esc(s: object) -> str:
            return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        # ID with chip nr and italic species — matches Reports PDF layout
        _id_raw     = str(rec.get('id',       '') or '').strip() or '\u2013'
        _chip_raw   = str(rec.get('chip_nr',  '') or '').strip()
        _species_raw = str(rec.get('species', '') or '').strip()
        if _species_raw.lower() in ('', 'none', 'null'):
            _species_raw = ''
        _id_part = _esc(_id_raw)
        if _chip_raw and _id_raw != '\u2013':
            _id_part = f"{_esc(_id_raw)} / {_esc(_chip_raw)}"
        elif _chip_raw:
            _id_part = _esc(_chip_raw)
        id_display = (f"{_id_part} <i>{_esc(_species_raw)}</i>"
                      if _species_raw else _id_part)

        status_display = _esc(status_summary_with_death_priority(
            rec,
            _messages,
            projects_track_active=True,
        ))

        # Active diagnoses
        active_sick = self.store.get_active_issues(animal_name, 'sick')
        active_abnormal = self.store.get_active_issues(animal_name, 'abnormal')
        all_active = active_sick + active_abnormal
        if all_active:
            dp: List[str] = []
            for iss in all_active:
                ck = str(iss.get('condition_key', ''))
                lbl = (ConditionLoader.compact_hierarchy_label(ck, _lang)
                       if ck else str(iss.get('condition_label_snapshot', ''))) or '\u2013'
                st = str(iss.get('status_type', ''))
                dp.append(_esc(f'{lbl} [{st}]' if st else lbl))
            diag_display = '; '.join(dp)
        else:
            diag_display = '\u2013'

        entries = sorted(
            (
                entry for entry in self.store.get_entries(animal_name)
                if _entry_date_in_range(entry.get("date"), date_from, date_to)
            ),
            key=lambda e: str(e.get('date', '')))

        rows = ''
        for entry in entries:
            etype = str(entry.get('entry_type', ''))
            stype = str(entry.get('status_type', ''))
            ckey = str(entry.get('condition_key', ''))
            condition = _esc(
                ConditionLoader.compact_hierarchy_label(ckey, _lang)
                if ckey else str(entry.get('condition_label_snapshot', '')))
            type_display = _esc(self._entry_type_label(etype, stype))
            signature_cell = (
                f'<td>{_esc(entry.get("signature", ""))}</td>'
                if include_signature else ""
            )
            rows += (
                f'<tr>'
                f'<td>{_esc(entry.get("date", ""))}</td>'
                f'<td>{type_display}</td>'
                f'<td>{condition}</td>'
                f'<td>{_esc(entry.get("note", ""))}</td>'
                f'{signature_cell}'
                f'</tr>'
            )

        # Pre-compute cell values (backslashes not allowed inside f-string expressions in Python < 3.12)
        _dash = '\u2013'
        _em   = '\u2014'
        _h_history  = _msg(_messages, 'medi_track.section.history', 'Medical History')
        _h_name     = _msg(_messages, 'report.header.name', 'Name')
        _h_id       = _msg(_messages, 'report.header.id', 'ID')
        _h_genotype = _msg(_messages, 'reports.header.genotype', 'Genotype')
        _h_project  = _msg(_messages, 'report.header.project', 'Project')
        _h_status   = _msg(_messages, 'report.header.status', 'Status')
        _h_birth    = _msg(_messages, 'report.header.birth_date', 'Birth Date')
        _h_death    = _msg(_messages, 'report.header.death_date', 'Death Date')
        _h_diag     = _msg(_messages, 'medi_track.header.active_diagnoses', 'Active Diagnoses')
        _h_date_col = _msg(_messages, 'medi_track.table.date', 'Date')
        _h_type_col = _msg(_messages, 'medi_track.table.type', 'Type')
        _h_cond_col = _msg(_messages, 'medi_track.table.condition', 'Condition')
        _h_det_col  = _msg(_messages, 'medi_track.table.details', 'Details')
        _h_sig_col  = _msg(_messages, 'medi_track.table.signature', 'Signature')
        _v_name     = _esc(_display_animal_name(animal_name, rec))
        _v_genotype = _esc(rec.get('genotype') or _dash)
        _proj_sev_fn = getattr(self.app, '_format_project_severity', None)
        _cur_proj = _proj_sev_fn(rec) if callable(_proj_sev_fn) else (rec.get('project') or '')
        _pdf_cur_lbl = _msg(_messages, "medi_track.header.project_current", "Current")
        _pdf_fmr_lbl = _msg(_messages, "medi_track.header.project_former", "Former")
        _pdf_proj_lines: List[str] = []
        if _cur_proj:
            _pdf_proj_lines.append(f"{_pdf_cur_lbl}: {_esc(_cur_proj)}")
        _pdf_hist = rec.get('project_history', [])
        if _pdf_hist:
            _pdf_fmr_items = []
            for _ph in reversed(_pdf_hist):
                _pn2 = _ph.get('project', '')
                _sv2 = _ph.get('severity', '')
                _ed2 = _ph.get('entry_date', '')
                _ld2 = _ph.get('leave_date', '')
                _hs2 = f"{_pn2} ({_sv2})" if _sv2 else _pn2
                if _ed2 or _ld2:
                    _hs2 = f"{_hs2} [{_ed2}\u2013{_ld2}]"
                if _hs2:
                    _pdf_fmr_items.append(_esc(_hs2))
            if _pdf_fmr_items:
                _pdf_proj_lines.append(f"{_pdf_fmr_lbl}: {', '.join(_pdf_fmr_items)}")
        _v_project = '<br>'.join(_pdf_proj_lines) if _pdf_proj_lines else _esc(_dash)
        _v_birth    = _esc(rec.get('birth_date') or _dash)
        _v_death    = _esc(rec.get('sterbedatum') or rec.get('death_date') or _dash)

        signature_header = (
            f"<th>{_msg(self.messages, 'medi_track.table.signature', 'Signature')}</th>"
            if include_signature else ""
        )
        html = f"""<html><head><style>
            body {{ font-family: Arial, sans-serif; font-size: 10pt; margin: 20px; }}
            h2 {{ color: #2c3e6b; font-size: 14pt; margin-bottom: 8px;
                  border-bottom: 2px solid #2c3e6b; padding-bottom: 4px; }}
            .it {{ width: 100%; border-collapse: collapse; margin-bottom: 16px; }}
            .it td {{ padding: 4px 8px; border: 1px solid #ccc; vertical-align: top; }}
            .it .lbl {{ font-weight: bold; background: #eef0f8; width: 140px; white-space: nowrap; }}
            .ht {{ border-collapse: collapse; width: 100%; }}
            .ht th {{ background: #2c3e6b; color: white; padding: 6px 8px;
                      text-align: left; font-size: 9pt; }}
            .ht td {{ padding: 5px 8px; border: 1px solid #ccc; vertical-align: top; font-size: 9pt; }}
            .ht tr:nth-child(even) td {{ background: #f8f8fc; }}
        </style></head><body>
        <h2>{_h_history} {_em} {_v_name}</h2>
        <table class='it'>
          <tr><td class='lbl'>{_h_name}</td><td>{_v_name}</td></tr>
          <tr><td class='lbl'>{_h_id}</td><td>{id_display}</td></tr>
          <tr><td class='lbl'>{_h_genotype}</td><td>{_v_genotype}</td></tr>
          <tr><td class='lbl'>{_h_project}</td><td>{_v_project}</td></tr>
          <tr><td class='lbl'>{_h_status}</td><td>{status_display}</td></tr>
          <tr><td class='lbl'>{_h_birth}</td><td>{_v_birth}</td></tr>
          <tr><td class='lbl'>{_h_death}</td><td>{_v_death}</td></tr>
          <tr><td class='lbl'>{_h_diag}</td><td>{diag_display}</td></tr>
        </table>
        <table class='ht'>
        <tr>
            <th>{_msg(self.messages, 'medi_track.table.date', 'Date')}</th>
            <th>{_msg(self.messages, 'medi_track.table.type', 'Type')}</th>
            <th>{_msg(self.messages, 'medi_track.table.condition', 'Condition')}</th>
            <th>{_msg(self.messages, 'medi_track.table.details', 'Details')}</th>
            {signature_header}
        </tr>
        {rows}
        </table></body></html>"""

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(output_path)
        doc = QTextDocument()
        doc.setHtml(html)
        doc.print(printer)
        # Apply the shared institution header after the Qt PDF has been
        # written, just like the other PDF-producing plugins.  Medi Track is
        # often used as the medical-history export path, so it must not be an
        # unbranded exception when facility branding is enabled.
        from Plugins.core.institution_branding import brand_generated_pdf
        brand_generated_pdf(self, output_path)

    def _copy_documents_for_export(self, animal_name: str, output_path: str) -> int:
        docs = _document_paths_for_animal(animal_name, self.store)
        return _copy_document_files_to_directory(docs, Path(output_path).parent)

    def _export_animal_history_package(
        self,
        animal_name: str,
        output_path: str,
        lang: Optional[str] = None,
    ) -> int:
        self._export_animal_to_pdf(animal_name, output_path, lang=lang)
        return self._copy_documents_for_export(animal_name, output_path)

    def _export_animal_to_xlsx(
        self,
        animal_name: str,
        output_path: str,
        *,
        include_signature: bool = True,
        include_documents: bool = True,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> int:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Medical history"
        headers = ["Date", "Type", "Condition", "Details"]
        if include_signature:
            headers.append("Signature")
        headers.append("Document")
        sheet.append(headers)

        copied = {}
        if include_documents:
            document_dir = Path(output_path).with_suffix("")
            document_dir = document_dir.parent / f"{document_dir.name}_documents"
            document_dir.mkdir(parents=True, exist_ok=True)
            for source in _document_paths_for_animal(animal_name, self.store):
                src = Path(source)
                if not src.is_file():
                    continue
                destination = document_dir / src.name
                counter = 1
                while destination.exists():
                    destination = document_dir / f"{src.stem}_{counter}{src.suffix}"
                    counter += 1
                shutil.copy2(src, destination)
                copied[str(src)] = destination

        entries = sorted(
            (
                entry for entry in self.store.get_entries(animal_name)
                if _entry_date_in_range(entry.get("date"), date_from, date_to)
            ),
            key=lambda entry: str(entry.get("date", "")),
        )
        for entry in entries:
            row = [
                entry.get("date", ""),
                self._entry_type_label(
                    str(entry.get("entry_type", "")),
                    str(entry.get("status_type", "")),
                ),
                entry.get("condition_label_snapshot", ""),
                entry.get("note", ""),
            ]
            if include_signature:
                row.append(entry.get("signature", ""))
            linked = ""
            for source in entry.get("document_paths", []) or []:
                destination = copied.get(str(source))
                if destination:
                    linked = os.path.relpath(destination, Path(output_path).parent)
                    break
            row.append(linked)
            sheet.append(row)
            if linked:
                cell = sheet.cell(sheet.max_row, len(row))
                cell.hyperlink = linked
                cell.style = "Hyperlink"
        for destination in copied.values():
            linked = os.path.relpath(destination, Path(output_path).parent)
            row = ["", "Document", "", destination.name]
            if include_signature:
                row.append("")
            row.append(linked)
            sheet.append(row)
            cell = sheet.cell(sheet.max_row, len(row))
            cell.hyperlink = linked
            cell.style = "Hyperlink"
        workbook.save(output_path)
        return len(copied)

    def _permission_scoped_export_candidates(self) -> List[Tuple[str, str]]:
        """Return user-visible animals as ``(internal key, Name (ID))`` rows."""
        central = getattr(self.app, "_permitted_export_candidates", None)
        if callable(central):
            try:
                return list(central(grouped=True, purpose="medi"))
            except TypeError:
                return list(central())

        app_animals = getattr(self.app, "animals", {}) or {}
        visible_fn = getattr(self.app, "_animal_visible_to_current_user", None)
        candidates: List[Tuple[str, str]] = []
        for animal_id, record in app_animals.items():
            if callable(visible_fn) and not visible_fn(record):
                continue
            display = _display_animal_name(animal_id, record)
            public_id = str(record.get("id") or "").strip()
            label = f"{display} ({public_id})" if public_id else display
            candidates.append((str(animal_id), label))
        return sorted(candidates, key=lambda row: (row[1].casefold(), row[0]))

    def _permission_scope_export_ids(self, requested: Iterable[str]) -> List[str]:
        central = getattr(self.app, "_permission_scope_export_ids", None)
        if callable(central):
            try:
                return list(central(requested, purpose="medi"))
            except TypeError:
                return list(central(requested))
        allowed = {row[0] for row in self._permission_scoped_export_candidates()}
        return [str(animal_id) for animal_id in requested if str(animal_id) in allowed]

    def _build_export_dialog(
        self,
        *,
        mode: str = UnifiedExportDialog.MODE_TAB_CURRENT,
    ) -> UnifiedExportDialog:
        candidates = self._permission_scoped_export_candidates()
        if mode == UnifiedExportDialog.MODE_TAB_CURRENT:
            candidates = [row for row in candidates if row[0] == self._current_animal]
        project_options = None
        scope = getattr(self.app, "_project_visibility_scope", None)
        if callable(scope):
            _unrestricted, visible_projects = scope()
            project_options = sorted(visible_projects)
        return UnifiedExportDialog(
            self,
            title=_msg(
                self.messages,
                "medi_track.dialog.export.title",
                "Export Medical History",
            ),
            candidates=candidates,
            mode=mode,
            current_animal=self._current_animal,
            formats=("PDF", "XLSX"),
            messages=self.messages,
            show_signatures=True,
            show_documents=True,
            show_date_range=True,
            project_options=project_options,
        )

    def _export_animals_to_directory(
        self,
        requested_animals: Iterable[str],
        output_dir: str,
        *,
        export_format: str,
        include_signatures: bool,
        include_documents: bool,
        lang: Optional[str] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> List[str]:
        """Export only if every requested animal is still permission-visible."""
        requested = [str(animal_id) for animal_id in requested_animals]
        selected = self._permission_scope_export_ids(requested)
        if selected != requested:
            raise PermissionError(
                _msg(
                    self.messages,
                    "medi_track.dialog.permission_denied.docs",
                    "You do not have permission to export medical history.",
                )
            )

        suffix = str(export_format).strip().lower()
        if suffix not in ("pdf", "xlsx"):
            raise ValueError(f"Unsupported export format: {export_format}")
        exported: List[str] = []
        for animal_id in selected:
            path = os.path.join(
                output_dir,
                f"{_safe_name(animal_id)}_medical_history.{suffix}",
            )
            if suffix == "xlsx":
                self._export_animal_to_xlsx(
                    animal_id,
                    path,
                    include_signature=include_signatures,
                    include_documents=include_documents,
                    date_from=date_from,
                    date_to=date_to,
                )
            else:
                self._export_animal_to_pdf(
                    animal_id,
                    path,
                    lang=lang or self._lang,
                    include_signature=include_signatures,
                    date_from=date_from,
                    date_to=date_to,
                )
            exported.append(path)
        return exported

    def _on_export_clicked(self) -> None:
        """Export the current tab animal without exposing a batch selector."""
        if not self._current_animal:
            return
        if not self._can_export():
            QMessageBox.warning(
                self,
                _msg(
                    self.messages,
                    "medi_track.dialog.permission_denied.title",
                    "Permission Denied",
                ),
                _msg(
                    self.messages,
                    "medi_track.dialog.permission_denied.docs",
                    "You do not have permission to export medical history.",
                ),
            )
            return

        dialog = self._build_export_dialog(
            mode=UnifiedExportDialog.MODE_TAB_CURRENT
        )
        if not dialog.selected_animal_ids():
            QMessageBox.warning(
                self,
                _msg(self.messages, "error.title", "Error"),
                _msg(
                    self.messages,
                    "medi_track.dialog.permission_denied.docs",
                    "You do not have permission to export medical history.",
                ),
            )
            return
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        output_dir = QFileDialog.getExistingDirectory(
            self,
            _msg(
                self.messages,
                "medi_track.dialog.export.title",
                "Export Medical History",
            ),
            QStandardPaths.writableLocation(
                QStandardPaths.StandardLocation.DesktopLocation
            ),
        )
        if not output_dir:
            return
        try:
            date_from, date_to = dialog.selected_date_range()
            self._export_animals_to_directory(
                dialog.selected_animal_ids(),
                output_dir,
                export_format=dialog.selected_format(),
                include_signatures=dialog.include_signatures(),
                include_documents=dialog.include_documents(),
                lang=self._lang,
                date_from=date_from,
                date_to=date_to,
            )
        except (PermissionError, RuntimeError, ValueError) as exc:
            QMessageBox.warning(
                self, _msg(self.messages, "error.title", "Error"), str(exc)
            )
            return
        except Exception as exc:
            QMessageBox.warning(
                self, _msg(self.messages, "error.title", "Error"), str(exc)
            )
            return
        QMessageBox.information(
            self,
            _msg(self.messages, "title.info", "Info"),
            _msg(
                self.messages,
                "medi_track.export.success",
                "Medical history exported.",
            ),
        )

    # ── add-document helpers ──

    def _animal_docs_folder(self, animal_name: str) -> Path:
        return self.app.backend.paths.managed_documents

    def _on_add_document_clicked(self) -> None:
        if not self._current_animal:
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            _msg(self.messages, "medi_track.dialog.select_file", "Select document"),
        )
        if not path:
            return
        src = Path(path)
        try:
            mt = getattr(self.app, "master_track", None)
            actor = str(
                getattr(mt, "current_username", "") or "medical-user"
            )
            record = self.app.backend.documents.add(
                src,
                owner_type="animal-medical",
                owner_id=self._current_animal,
                actor=actor,
            )
            dest = self.app.backend.documents.payload_path(record)
        except Exception as exc:
            logger.error("copy document failed: %s", exc)
            QMessageBox.warning(
                self,
                _msg(self.messages, "error.title", "Error"),
                str(exc))
            return
        doc = {
            "original_name": src.name,
            "path": str(dest),
            "title": src.name,
            "document_id": str(record["document_id"]),
        }
        try:
            self.store.add_document(self._current_animal, doc)
        except Exception as exc:
            logger.error("add_document failed: %s", exc)
        self._reload_docs(self._current_animal)

    def _reload_docs(self, animal_name: str) -> None:
        """Refresh the documents list for the given animal."""
        self._docs_list.clear()
        found: List[str] = []
        for record in self.app.backend.documents.list_for_owner(
            "animal-medical", animal_name
        ):
            found.append(
                str(self.app.backend.documents.payload_path(record))
            )
        json_docs = self.store.get_documents(animal_name)
        json_paths = {d.get('path', '') for d in json_docs}
        extra = [d for d in json_docs
                 if d.get('path') and d['path'] not in found
                 and Path(d['path']).exists()]
        all_paths = [(p, Path(p).name) for p in found] + [
            (d['path'], d.get('title') or Path(d['path']).name) for d in extra]
        if not all_paths:
            self._docs_list.addItem(
                _msg(self.messages, "medi_track.empty.no_documents", "(no documents)"))
            return
        for fpath, fname in all_paths:
            icon = _icon_for_ext(Path(fpath).suffix)
            item = QListWidgetItem(icon, fname)
            item.setData(Qt.ItemDataRole.UserRole, fpath)
            item.setToolTip(fpath)
            self._docs_list.addItem(item)

    def _on_doc_item_clicked(self, item: QListWidgetItem) -> None:
        path = item.data(Qt.ItemDataRole.UserRole)
        if path and Path(path).exists():
            if not open_local_path(path):
                logger.error("Could not open document path: %s", path)

    # ── filter logic ──

    def _on_filter_clicked(self, fkey: str) -> None:
        self._active_filter = fkey
        for k, btn in self._filter_btns.items():
            btn.setChecked(k == fkey)
        self._refresh_from_filter()

    def _refresh_from_filter(self) -> None:
        """
        The filter applies to which ANIMALS are highlighted in the main list;
        for the history view it just refreshes the current animal view.
        Notify app if available.
        """
        cb = getattr(self.app, '_medi_filter_changed', None)
        if callable(cb):
            cb(self._active_filter)

    # ── docs section ──

    def _on_docs_toggled(self, checked: bool) -> None:
        self._docs_toggle.setText(
            _msg(self.messages, "medi_track.section.documents", "Documents"))
        apply_icon(
            self._docs_toggle,
            "toggle.collapse" if checked else "toggle.expand",
            fallback="Documents",
        )
        self._docs_widget.setVisible(checked)

    # ── public API called by plugin ──

    def show_animal(self, name: Optional[str]) -> None:
        """Called when selection changes. Reload header + history for *name*."""
        self._current_animal = name
        if not name:
            self._set_content_visible(False)
            return

        app_animals = self.app.animals if isinstance(
            getattr(self.app, "animals", None), dict) else {}
        rec = app_animals.get(name, {})

        self._lbl_name.setText(_display_animal_name(name, rec))
        fmt_fn = getattr(self.app, '_format_id_with_species', None)
        if callable(fmt_fn):
            self._lbl_id.setText(fmt_fn(rec, rich_text=True) or "\u2013")
        else:
            self._lbl_id.setText(str(rec.get("id", "\u2013")) or "\u2013")
        self._lbl_chip_nr.setText(str(rec.get("chip_nr", "") or "\u2013"))
        self._lbl_origin.setText(str(rec.get("origin", "") or "\u2013"))
        self._lbl_genotype.setText(str(rec.get("genotype", "") or "\u2013"))
        _proj_sev_fn = getattr(self.app, '_format_project_severity', None)
        cur_txt = _proj_sev_fn(rec) if callable(_proj_sev_fn) else (rec.get('project') or '')
        _cur_lbl = _msg(self.messages, "medi_track.header.project_current", "Current")
        _fmr_lbl = _msg(self.messages, "medi_track.header.project_former", "Former")
        _proj_parts: List[str] = []
        if cur_txt:
            _proj_parts.append(f"{_cur_lbl}: {cur_txt}")
        _hist = rec.get('project_history', [])
        if _hist:
            _fmr_items = []
            for _h in reversed(_hist):
                _pn = _h.get('project', '')
                _sv = _h.get('severity', '')
                _ed = _h.get('entry_date', '')
                _ld = _h.get('leave_date', '')
                _hs = f"{_pn} ({_sv})" if _sv else _pn
                if _ed or _ld:
                    _hs = f"{_hs} [{_ed}\u2013{_ld}]"
                if _hs:
                    _fmr_items.append(_hs)
            if _fmr_items:
                _proj_parts.append(f"{_fmr_lbl}: {', '.join(_fmr_items)}")
        self._lbl_project.setText('\n'.join(_proj_parts) if _proj_parts else '\u2013')

        self._lbl_status.setText(status_summary_with_death_priority(
            rec,
            self.messages,
            projects_track_active=True,
        ))

        self._lbl_birth.setText(str(rec.get("geburtsdatum", rec.get("birth_date", "\u2013"))) or "\u2013")
        death = rec.get("sterbedatum") or rec.get("death_date") or ""
        self._lbl_death.setText(str(death) if death else "\u2013")

        # Active diagnoses — compact hierarchy, strip root level
        active_sick = self.store.get_active_issues(name, "sick")
        active_abnormal = self.store.get_active_issues(name, "abnormal")
        all_active = active_sick + active_abnormal
        if all_active:
            diag_parts = []
            for iss in all_active:
                cond_key = str(iss.get("condition_key", ""))
                if cond_key:
                    label = ConditionLoader.compact_hierarchy_label(cond_key, self._lang) or "\u2013"
                else:
                    label = str(iss.get("condition_label_snapshot", "")) or "\u2013"
                status_t = str(iss.get("status_type", ""))
                tag = f" [{status_t}]" if status_t else ""
                diag_parts.append(f"{label}{tag}")
            self._lbl_diagnoses.setText("; ".join(diag_parts))
        else:
            self._lbl_diagnoses.setText("\u2013")

        # History table
        entries = self.store.get_entries(name)
        self._table.setRowCount(0)
        self._table.setSortingEnabled(False)
        for entry in sorted(entries, key=lambda e: str(e.get("date", ""))):
            row = self._table.rowCount()
            self._table.insertRow(row)

            etype = str(entry.get("entry_type", ""))
            stype = str(entry.get("status_type", ""))
            cond_key = str(entry.get("condition_key", ""))
            condition = (ConditionLoader.compact_hierarchy_label(cond_key, self._lang)
                         if cond_key else
                         str(entry.get("condition_label_snapshot", "")))

            type_display = self._entry_type_label(etype, stype)

            self._table.setItem(row, 0, QTableWidgetItem(str(entry.get("date", ""))))
            self._table.setItem(row, 1, QTableWidgetItem(type_display))
            self._table.setItem(row, 2, QTableWidgetItem(condition))
            self._table.setItem(row, 3, QTableWidgetItem(str(entry.get("note", ""))))
            self._table.setItem(row, 4, QTableWidgetItem(str(entry.get("signature", ""))))
        self._table.setSortingEnabled(True)

        # Documents
        self._reload_docs(name)

        self._set_content_visible(True)

    def _entry_type_label(self, etype: str, stype: str) -> str:
        """Human-readable entry type label."""
        key_map = {
            ("issue_start", "sick"): ("medi_track.report.event.sick_started", "Sick — started"),
            ("issue_start", "abnormal"): ("medi_track.report.event.abnormal_started", "Abnormal — started"),
            ("issue_resolution", "sick"): ("medi_track.report.event.sick_resolved", "Sick — resolved"),
            ("issue_resolution", "abnormal"): ("medi_track.report.event.abnormal_resolved", "Abnormal — resolved"),
            ("note", ""): ("medi_track.entry_type.note", "Note"),
            ("project_assigned", "other"): ("medi_track.report.event.project_assigned", "Project \u2014 assigned"),
            ("project_left", "other"): ("medi_track.report.event.project_left", "Project \u2014 left"),
            ("severity_changed", "other"): ("medi_track.report.event.severity_changed", "Severity \u2014 changed"),
            ("experiment_started", "other"): ("medi_track.report.event.experiment_started", "Experiment \u2014 started"),
            ("experiment_ended",   "other"): ("medi_track.report.event.experiment_ended",   "Experiment \u2014 ended"),
            ("death",              "other"): ("medi_track.report.event.death",              "Death"),
            ("identity_changed",   "other"): ("medi_track.report.event.identity_changed",   "Identity changed"),
        }
        msg_key, fallback = key_map.get((etype, stype), ("medi_track.entry_type.unknown", etype or stype or "\u2013"))
        return _msg(self.messages, msg_key, fallback)

    def active_filter(self) -> str:
        return self._active_filter

    def reset_filter(self) -> None:
        """Reset filter to 'all' (called when the Medi Track tab becomes inactive)."""
        if self._active_filter == self.FILTER_ALL:
            return
        self._active_filter = self.FILTER_ALL
        for k, btn in self._filter_btns.items():
            btn.setChecked(k == self.FILTER_ALL)


# ─────────────────────────────────────────────────────────────────────────────
# ManualEntryDialog — manually add a history line
# ─────────────────────────────────────────────────────────────────────────────

_ENTRY_TYPES = [
    ("note",        "medi_track.entry_type.note",        "Note"),
    ("observation", "medi_track.entry_type.observation",  "Observation"),
    ("treatment",   "medi_track.entry_type.treatment",    "Treatment"),
    ("measurement", "medi_track.entry_type.measurement",  "Measurement"),
    ("other",       "medi_track.entry_type.other",        "Other"),
]


class ManualEntryDialog(QDialog):
    """Simple dialog to manually add a free-form history entry."""

    def __init__(
        self,
        parent: Optional[QWidget],
        messages: Dict[str, Any],
        animal_name: str,
        default_signature: str = "",
        readonly_signature: bool = False,
        lang_code: str = "en",
        display_name: str = "",
    ) -> None:
        super().__init__(parent)
        self.messages = messages
        self.setWindowTitle(
            _msg(messages, "medi_track.dialog.add_entry.title",
                 "Add History Entry — {animal}").replace("{animal}", display_name if display_name else animal_name))
        self.setModal(True)
        self.resize(480, 320)

        layout = QVBoxLayout(self)
        form = QFormLayout()

        # Date
        self._date_edit = QDateEdit()
        self._date_edit.setCalendarPopup(True)
        self._date_edit.setDate(QDate.currentDate())
        self._date_edit.setDisplayFormat("dd.MM.yyyy")
        form.addRow(_msg(messages, "medi_track.dialog.label.date", "Date:"), self._date_edit)

        # Entry type
        self._type_combo = QComboBox()
        for key, msg_key, fallback in _ENTRY_TYPES:
            self._type_combo.addItem(_msg(messages, msg_key, fallback), key)
        form.addRow(_msg(messages, "medi_track.dialog.label.type", "Type:"), self._type_combo)

        layout.addLayout(form)

        form2 = QFormLayout()

        # Note / details
        self._note_edit = QTextEdit()
        self._note_edit.setFixedHeight(70)
        form2.addRow(_msg(messages, "medi_track.dialog.label.note", "Note:"), self._note_edit)

        # Signature
        self._sig_edit = QLineEdit(default_signature)
        if readonly_signature:
            self._sig_edit.setReadOnly(True)
            self._sig_edit.setStyleSheet("QLineEdit { background: #f0f0f0; color: #666; }")
        form2.addRow(
            _msg(messages, "medi_track.dialog.label.signature", "Signature:"),
            self._sig_edit)

        layout.addLayout(form2)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        # Set explicit translated text
        btns.button(QDialogButtonBox.StandardButton.Ok).setText(
            _msg(messages, "button.ok", "OK"))
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText(
            _msg(messages, "button.cancel", "Cancel"))
        layout.addWidget(btns)

    def result_date(self) -> str:
        return self._date_edit.date().toString("yyyy-MM-dd")

    def result_entry_type(self) -> str:
        return self._type_combo.currentData() or "note"

    def result_condition_key(self) -> str:
        return ""

    def result_condition_label(self) -> str:
        return ""

    def result_note(self) -> str:
        return self._note_edit.toPlainText().strip()

    def result_signature(self) -> str:
        return self._sig_edit.text().strip()


# ─────────────────────────────────────────────────────────────────────────────
# MediTrackPlugin — plugin entry point
# ─────────────────────────────────────────────────────────────────────────────

class MediTrackPlugin:
    """Returned by initialize(app). Provides tab widget and status hooks."""

    def __init__(self, app: Any) -> None:
        self.app = app
        self.messages: Dict[str, Any] = getattr(app, "messages", {}) or {}
        self._lang = self._detect_lang()
        self.store = MediStore(app.backend)
        self._widget: Optional[MediTrackWidget] = None

    # ── helpers ──

    def _refresh_current_animal_when_safe(self, animal_name: str) -> None:
        if not self._widget or self._widget._current_animal != animal_name:
            return
        if QApplication.activeModalWidget() is not None:
            QTimer.singleShot(
                250,
                lambda name=animal_name: self._refresh_current_animal_when_safe(name),
            )
            return
        try:
            self._widget.show_animal(animal_name)
        except RuntimeError:
            pass

    def _detect_lang(self) -> str:
        msgs = getattr(self.app, "messages", {}) or {}
        lang = msgs.get("_lang_code", "en")
        return str(lang).lower()[:2] if lang else "en"

    def _detect_data_dir(self) -> str:
        """Return the configured application data directory for attachments."""
        app_file = getattr(self.app, "__file__", None) or ""
        if app_file:
            return str(Path(app_file).parent)
        return str(Path(__file__).parent.parent.parent)

    def _can(self, perm: str) -> bool:
        can_fn = getattr(self.app, "_master_can", None)
        if callable(can_fn):
            return bool(can_fn(perm))
        return True

    # ── tab widget API ──

    def open_status_dialog(
        self,
        parent: QWidget,
        animal_name: str,
        status_type: str,
        mode: str = 'add',
        prefill_signature: str = "",
        can_add: Optional[bool] = None,
        can_resolve: Optional[bool] = None,
    ) -> bool:
        """
        Open StatusManagementDialog for an immediate click event.
        mode='add'     → can_add=True, can_resolve=False (default)
        mode='resolve' → can_add=False, can_resolve=True
        mode='manage'  → both enabled (pass explicit can_add/can_resolve)
        Returns True when the user accepted and the entry was written.
        """
        if can_add is None:
            can_add = mode in ('add', 'manage')
        if can_resolve is None:
            can_resolve = mode in ('resolve', 'manage')
        mt = getattr(self.app, 'master_track', None)
        readonly_sig = bool(mt and getattr(mt, 'is_logged_in', False))
        active = self.store.get_active_issues(animal_name, status_type)
        _app_animals = getattr(self.app, 'animals', {}) or {}
        _dname = animal_base_name(animal_name, _app_animals.get(animal_name) or {})
        dlg = StatusManagementDialog(
            parent, self.messages, animal_name, status_type, active,
            lang_code=self._lang,
            can_add=can_add,
            can_resolve=can_resolve,
            prefill_signature=prefill_signature,
            readonly_signature=readonly_sig,
            display_name=_dname,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return False
        sig = dlg.result_signature() or _msg(
            self.messages, "medi_track.value.guest", "(guest)")
        if dlg.result_mode() == StatusManagementDialog.MODE_ADD:
            entry = self._build_entry("issue_start", status_type, dlg, sig)
        else:
            entry = self._build_resolution_entry(
                status_type, dlg.result_issue_id(), dlg.result_note(), sig,
                animal_name=animal_name)
        try:
            self.store.add_entry(animal_name, entry)
        except Exception as exc:
            logger.error("open_status_dialog add_entry: %s", exc)
        self._refresh_current_animal_when_safe(animal_name)
        return True

    def get_tab_widget(self) -> QWidget:
        if self._widget is None:
            self._widget = MediTrackWidget(
                self.app, self.store, self.messages, self._lang,
                data_dir=self._detect_data_dir())
        return self._widget

    # ── called from ProgTrack when selection changes ──

    def on_animal_selected(self, animal_names: List[str]) -> None:
        if self._widget is None:
            return
        name = animal_names[-1] if animal_names else None
        self._widget.show_animal(name)

    def refresh_view(self) -> None:
        if self._widget is not None and self._widget._current_animal:
            self._widget.show_animal(self._widget._current_animal)

    def update_language(self, messages: Dict[str, Any]) -> None:
        """Update language messages and refresh all Medi Track UI labels."""
        self.messages = messages
        self._lang = self._detect_lang()
        if self._widget is not None:
            self._widget.update_language(messages, self._lang)

    # ── status change hooks (called from animal dialog save paths) ──

    def on_sick_changed(self, animal_name: str, is_sick: bool, signature: str = "") -> None:
        """Called after sick checkbox is toggled AND confirmed. Writes medi_history entry."""
        self._write_status_change(animal_name, "sick", is_sick, signature)

    def on_abnormal_changed(self, animal_name: str, is_abnormal: bool, signature: str = "") -> None:
        """Called after abnormal checkbox is toggled AND confirmed. Writes medi_history entry."""
        self._write_status_change(animal_name, "abnormal", is_abnormal, signature)

    def _write_status_change(
        self,
        animal_name: str,
        status_type: str,
        is_on: bool,
        signature: str,
    ) -> None:
        issue_id = str(uuid.uuid4())
        entry = {
            "id": str(uuid.uuid4()),
            "date": _today_iso(),
            "entry_type": "issue_start" if is_on else "issue_resolution",
            "status_type": status_type,
            "issue_id": issue_id,
            "condition_key": "",
            "condition_label_snapshot": "",
            "condition_path_keys": [],
            "condition_path_label_snapshot": [],
            "condition_language": self._lang,
            "condition_depth": 0,
            "note": "",
            "signature": signature or _msg(
                self.messages, "medi_track.value.guest", "(guest)"),
            "linked_document_ids": [],
        }
        try:
            self.store.add_entry(animal_name, entry)
        except Exception as exc:
            logger.error("MediStore.add_entry failed: %s", exc, exc_info=True)

    def handle_sick_checkbox_clicked(
        self,
        animal_name: str,
        parent_widget: QWidget,
    ) -> Optional[bool]:
        """
        Handle a click on the currently-checked Sick checkbox.
        Returns the new sick state (True=stay sick, False=clear), or None if cancelled.
        """
        if not self._can("medi_track.status_enable") and not self._can("medi_track.status_manage"):
            QMessageBox.warning(
                parent_widget,
                _msg(self.messages, "medi_track.dialog.permission_denied.title", "Permission Denied"),
                _msg(self.messages, "medi_track.dialog.permission_denied.resolve",
                     "You do not have permission to manage medical status."))
            return None

        mt = getattr(self.app, 'master_track', None)
        readonly_sig = bool(mt and getattr(mt, 'is_logged_in', False))
        active = self.store.get_active_issues(animal_name, "sick")
        _app_animals = getattr(self.app, 'animals', {}) or {}
        _dname_sick = animal_base_name(animal_name, _app_animals.get(animal_name) or {})
        dlg = StatusManagementDialog(
            parent_widget, self.messages, animal_name, "sick", active,
            lang_code=self._lang,
            can_add=self._can("medi_track.status_enable"),
            can_resolve=self._can("medi_track.status_manage"),
            prefill_signature=self._get_signature(),
            readonly_signature=readonly_sig,
            display_name=_dname_sick,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return None

        signature = self._get_signature()
        if dlg.result_mode() == StatusManagementDialog.MODE_ADD:
            entry = self._build_entry("issue_start", "sick", dlg, signature)
            self.store.add_entry(animal_name, entry)
            return True  # stay sick

        if dlg.result_mode() == StatusManagementDialog.MODE_RESOLVE:
            issue_id = dlg.result_issue_id()
            entry = self._build_resolution_entry("sick", issue_id, dlg.result_note(), signature,
                                                  animal_name=animal_name)
            self.store.add_entry(animal_name, entry)
            remaining = self.store.get_active_issues(animal_name, "sick")
            return len(remaining) > 0  # stay sick only if issues remain

        return None

    def handle_abnormal_checkbox_clicked(
        self,
        animal_name: str,
        parent_widget: QWidget,
    ) -> Optional[bool]:
        """Same as handle_sick_checkbox_clicked but for abnormal status."""
        if not self._can("medi_track.status_enable") and not self._can("medi_track.status_manage"):
            QMessageBox.warning(
                parent_widget,
                _msg(self.messages, "medi_track.dialog.permission_denied.title", "Permission Denied"),
                _msg(self.messages, "medi_track.dialog.permission_denied.resolve",
                     "You do not have permission to manage medical status."))
            return None

        mt = getattr(self.app, 'master_track', None)
        readonly_sig = bool(mt and getattr(mt, 'is_logged_in', False))
        active = self.store.get_active_issues(animal_name, "abnormal")
        _app_animals = getattr(self.app, 'animals', {}) or {}
        _dname_abn = animal_base_name(animal_name, _app_animals.get(animal_name) or {})
        dlg = StatusManagementDialog(
            parent_widget, self.messages, animal_name, "abnormal", active,
            lang_code=self._lang,
            can_add=self._can("medi_track.status_enable"),
            can_resolve=self._can("medi_track.status_manage"),
            prefill_signature=self._get_signature(),
            readonly_signature=readonly_sig,
            display_name=_dname_abn,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return None

        signature = self._get_signature()
        if dlg.result_mode() == StatusManagementDialog.MODE_ADD:
            entry = self._build_entry("issue_start", "abnormal", dlg, signature)
            self.store.add_entry(animal_name, entry)
            return True

        if dlg.result_mode() == StatusManagementDialog.MODE_RESOLVE:
            issue_id = dlg.result_issue_id()
            entry = self._build_resolution_entry("abnormal", issue_id, dlg.result_note(), signature,
                                                  animal_name=animal_name)
            self.store.add_entry(animal_name, entry)
            remaining = self.store.get_active_issues(animal_name, "abnormal")
            return len(remaining) > 0

        return None

    # ── internal helpers ──

    def _get_signature(self) -> str:
        mt = getattr(self.app, "master_track", None)
        if mt and getattr(mt, "is_logged_in", False):
            dname = getattr(mt, 'current_display_name', None)
            return str(dname or getattr(mt, "current_username", "") or "") or _msg(
                self.messages, "medi_track.value.guest", "(guest)")
        return _msg(self.messages, "medi_track.value.guest", "(guest)")

    def _build_entry(
        self,
        entry_type: str,
        status_type: str,
        dlg: StatusManagementDialog,
        signature: str,
    ) -> Dict[str, Any]:
        ckey = dlg.result_condition_key() or ""
        clabel = dlg.result_condition_label() if ckey else ""
        return {
            "id": str(uuid.uuid4()),
            "date": _today_iso(),
            "entry_type": entry_type,
            "status_type": status_type,
            "issue_id": str(uuid.uuid4()),
            "condition_key": ckey,
            "condition_label_snapshot": clabel,
            "condition_path_keys": [],
            "condition_path_label_snapshot": [],
            "condition_language": self._lang,
            "condition_depth": len(ckey.split(".")) - 1 if ckey else 0,
            "note": dlg.result_note(),
            "signature": signature,
            "linked_document_ids": [],
        }

    def _build_resolution_entry(
        self,
        status_type: str,
        issue_id: Optional[str],
        note: str,
        signature: str,
        animal_name: str = "",
    ) -> Dict[str, Any]:
        # Copy condition info from the original issue_start entry
        cond_key = ""
        cond_label = ""
        if animal_name and issue_id:
            for e in self.store.get_entries(animal_name):
                if (e.get("issue_id") == issue_id
                        and e.get("entry_type") == "issue_start"):
                    cond_key = str(e.get("condition_key", ""))
                    cond_label = str(e.get("condition_label_snapshot", ""))
                    break
        return {
            "id": str(uuid.uuid4()),
            "date": _today_iso(),
            "entry_type": "issue_resolution",
            "status_type": status_type,
            "issue_id": issue_id or "",
            "condition_key": cond_key,
            "condition_label_snapshot": cond_label,
            "condition_path_keys": [],
            "condition_path_label_snapshot": [],
            "condition_language": self._lang,
            "condition_depth": len(cond_key.split(".")) - 1 if cond_key else 0,
            "note": note,
            "signature": signature,
            "linked_document_ids": [],
        }

    def log_severity_change(
        self,
        animal_name: str,
        project: str,
        old_severity: str,
        new_severity: str,
        signature: str = "",
    ) -> None:
        """Record a severity change as a Medi Track 'other' history entry."""
        template = _msg(
            self.messages, 'medi_track.entry.severity_changed',
            'Severity in {project} changed from {old} to {new}')
        note = (template
                .replace('{project}', project)
                .replace('{old}', old_severity)
                .replace('{new}', new_severity))
        entry = {
            "id": str(uuid.uuid4()),
            "date": _today_iso(),
            "entry_type": "severity_changed",
            "status_type": "other",
            "issue_id": "",
            "condition_key": "",
            "condition_label_snapshot": project,
            "condition_path_keys": [],
            "condition_path_label_snapshot": [],
            "condition_language": self._lang,
            "condition_depth": 0,
            "note": note,
            "signature": signature or _msg(
                self.messages, "medi_track.value.guest", "(guest)"),
            "linked_document_ids": [],
        }
        try:
            self.store.add_entry(animal_name, entry)
        except Exception as exc:
            logger.error("log_severity_change: %s", exc)
        self._refresh_current_animal_when_safe(animal_name)

    def log_project_change(
        self,
        animal_name: str,
        old_project: str,
        new_project: str,
        signature: str = "",
        old_severity_lbl: str = "",
        new_severity_lbl: str = "",
    ) -> None:
        """Record a project assignment change as a Medi Track history entry."""
        if old_project == new_project:
            return
        sig = signature or _msg(self.messages, "medi_track.value.guest", "(guest)")

        def _make_label(proj: str, sev_lbl: str) -> str:
            return f"{proj} ({sev_lbl})" if sev_lbl else proj

        def _make_entry(entry_type: str, proj: str, sev_lbl: str, note_prefix: str) -> Dict[str, Any]:
            label = _make_label(proj, sev_lbl)
            return {
                "id": str(uuid.uuid4()),
                "date": _today_iso(),
                "entry_type": entry_type,
                "status_type": "other",
                "issue_id": "",
                "condition_key": "",
                "condition_label_snapshot": label,
                "condition_path_keys": [],
                "condition_path_label_snapshot": [],
                "condition_language": self._lang,
                "condition_depth": 0,
                "note": f"{note_prefix}: {label}",
                "signature": sig,
                "linked_document_ids": [],
            }

        left_prefix = _msg(self.messages, "medi_track.report.event.project_left", "Project \u2014 left")
        asgn_prefix = _msg(self.messages, "medi_track.report.event.project_assigned", "Project \u2014 assigned")

        entries_to_add: List[Dict[str, Any]] = []
        if old_project and new_project:
            entries_to_add.append(_make_entry("project_left", old_project, old_severity_lbl, left_prefix))
            entries_to_add.append(_make_entry("project_assigned", new_project, new_severity_lbl, asgn_prefix))
        elif old_project:
            entries_to_add.append(_make_entry("project_left", old_project, old_severity_lbl, left_prefix))
        elif new_project:
            entries_to_add.append(_make_entry("project_assigned", new_project, new_severity_lbl, asgn_prefix))
        else:
            return

        for entry in entries_to_add:
            try:
                self.store.add_entry(animal_name, entry)
            except Exception as exc:
                logger.error("log_project_change: %s", exc)
        if self._widget and self._widget._current_animal == animal_name:
            self._widget.show_animal(animal_name)

    @staticmethod
    def _lifecycle_date_iso(value: str) -> str:
        """Return an ISO date without silently replacing a supplied valid date."""
        raw = str(value or "").strip()
        if not raw:
            return _today_iso()
        for pattern in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(raw, pattern).date().isoformat()
            except ValueError:
                continue
        return _today_iso()

    def log_lifecycle_change(
        self,
        animal_name: str,
        entry_type: str,
        signature: str = "",
        *,
        details: str = "",
        event_date: str = "",
        role: str = "",
    ) -> None:
        """Record one dated lifecycle event in an animal's medical history."""
        entry = {
            'id': str(uuid.uuid4()),
            'date': self._lifecycle_date_iso(event_date),
            'entry_type': entry_type,
            'status_type': 'other',
            'issue_id': '',
            'condition_key': '',
            'condition_label_snapshot': '',
            'condition_path_keys': [],
            'condition_path_label_snapshot': [],
            'condition_language': self._lang,
            'condition_depth': 0,
            'note': str(details or '').strip(),
            'signature': signature or _msg(
                self.messages, 'medi_track.value.guest', '(guest)'),
            'actor_role': str(role or '').strip(),
            'linked_document_ids': [],
        }
        try:
            self.store.add_entry(animal_name, entry)
        except Exception as exc:
            logger.error('log_experiment_change: %s', exc)
        if self._widget and self._widget._current_animal == animal_name:
            self._widget.show_animal(animal_name)

    def log_experiment_change(
        self,
        animal_name: str,
        entry_type: str,
        signature: str = '',
        *,
        details: str = '',
        event_date: str = '',
        role: str = '',
    ) -> None:
        """Compatibility entry point for experiment lifecycle events."""
        self.log_lifecycle_change(
            animal_name,
            entry_type,
            signature,
            details=details,
            event_date=event_date,
            role=role,
        )

    # ── filter helpers used by ProgTrack._refresh_list ──

    def matches_filter(self, animal_name: str, filter_key: str) -> bool:
        """Return True if the animal passes the active Medi_Track filter."""
        if filter_key == MediTrackWidget.FILTER_ALL:
            return True
        app_animals = self.app.animals if isinstance(
            getattr(self.app, "animals", None), dict) else {}
        rec = app_animals.get(animal_name, {})
        if filter_key == MediTrackWidget.FILTER_SICK:
            return bool(rec.get("sick", False))
        if filter_key == MediTrackWidget.FILTER_EVER_SICK:
            return bool(rec.get("sick", False)) or self.store.has_any_of_type(animal_name, "sick")
        if filter_key == MediTrackWidget.FILTER_ABNORMAL:
            return bool(rec.get("abnormal_current", False))
        if filter_key == MediTrackWidget.FILTER_EVER_ABNORMAL:
            return bool(rec.get("abnormal_ever", False)) or self.store.has_any_of_type(
                animal_name, "abnormal")
        if filter_key == MediTrackWidget.FILTER_IN_EXPERIMENT:
            return bool(rec.get("in_experiment", False))
        if filter_key == MediTrackWidget.FILTER_EVER_EXPERIMENT:
            return ever_in_experiment(rec, self.store.get_entries(animal_name))
        return True
